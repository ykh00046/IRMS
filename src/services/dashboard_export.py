"""운영 대시보드 보고서 Excel 출력 — 배합 실적(blend_records) + 점도 현황 기반.

구 계량 워크플로(recipe_items) 기반 보고서는 데이터가 더 이상 쌓이지 않아
2026-07 배합 기준으로 재작성. openpyxl 만 의존(서버 동작).
"""

import io
import sqlite3
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from . import viscosity_service


def _section(ws, row: int, title: str, headers: list[str] | None, rows: list[list[Any]]) -> int:
    bold = Font(bold=True)
    ws.cell(row=row, column=1, value=title).font = bold
    row += 1
    if headers:
        for col, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=h).font = bold
        row += 1
    for data in rows:
        for col, val in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=val)
        row += 1
    return row + 1  # 섹션 뒤 빈 줄


def build_dashboard_excel(
    connection: sqlite3.Connection,
    *,
    from_date: str,
    to_date: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "대시보드"
    ws.cell(row=1, column=1, value="운영 대시보드 보고서").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"기간: {from_date} ~ {to_date}")

    # 요약
    summary = connection.execute(
        """
        SELECT COUNT(*) AS cnt, COALESCE(SUM(total_amount), 0) AS w,
               COUNT(DISTINCT product_name) AS products,
               COUNT(DISTINCT worker) AS workers
        FROM blend_records
        WHERE status = 'completed' AND COALESCE(is_bulk_regenerated, 0) = 0
          AND work_date BETWEEN ? AND ?
        """,
        (from_date, to_date),
    ).fetchone()

    row = 4
    # 화면과 같은 기준(완료·비-bulk)으로 센다 — 예전엔 여기만 일괄 재생성분을 빼지 않아
    # 같은 기간인데 엑셀 건수가 화면보다 컸다. '결재 대기'는 결재를 현장에서 쓰지 않아
    # 2026-07 에 화면·API 에서 지웠는데 이 보고서에만 남아 있었다(2026-08-08 제거).
    row = _section(ws, row, "[요약]", None, [
        ["배합 건수", int(summary["cnt"] or 0)],
        ["총 배합량(g)", round(float(summary["w"] or 0), 2)],
        ["반제품 종류", int(summary["products"] or 0)],
        ["작업자 수", int(summary["workers"] or 0)],
    ])

    # 반제품별 TOP·작업자 통계는 뺐다(2026-08-08) — 배합 분석 리포트
    # (/api/blend/analysis/export)의 제품·품질 시트와 같은 표였고, 세는 기준만 서로
    # 달랐다. 이 보고서는 운영 스냅샷(요약·점도 현황)만 담는다.
    # 점도 현황 (제품별 최신 연도 기준)
    overview = viscosity_service.overview(connection)
    due = viscosity_service.daily_reading_reminders(
        connection, target_date=date.today().isoformat()
    )
    row = _section(
        ws, row, "[점도 현황 (최신 연도 기준)]",
        ["반제품", "측정수", "평균", "이상", "경고", "최근값", "최근일"],
        [
            [
                it["code"], it["count"], it["mean"], it["anomaly_count"],
                it["warn_count"], it["latest_value"], it["latest_date"],
            ]
            for it in overview["items"]
        ],
    )
    if due:
        row = _section(
            ws, row, "[오늘 점도 미입력 (알림 대상)]", ["반제품"],
            [[it["code"]] for it in due],
        )

    widths = [22, 14, 14, 12, 12, 12, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
