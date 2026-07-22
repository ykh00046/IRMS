"""원료배합일지(DHR) 공식 양식 Excel 출력.

C:\\X\\Program-estimation v3 의 ExcelWriter 를 웹용으로 이식한 것.
공식 양식(src/resources/dhr_template.xlsx, "원 료 배 합 일 지")을 복사·채워서
규제 양식과 동일한 배합일지를 출력한다. openpyxl 만 의존(서버에서 동작).

양식 셀 매핑(원본 config/settings.CELL_MAPPING 동일):
  날짜 A3 · 저울 A4 · 작업자 C3 · 작업시간 E3 · 제품LOT A6 · 총량/100 B6
  데이터 6행~ : 배합원료명 C · 원료LOT D · 배합비율 E · 배합량(g) F · 실제배합량(g) G
서명(작성/검토/승인)은 이미지 합성 단계(별도)에서 올린다.
"""

import io
import json
import os
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "resources", "dhr_template.xlsx")

CELL_MAPPING = {
    "date": "A3",
    "scale": "A4",
    "worker": "C3",
    "work_time": "E3",
    "product_lot": "A6",
    "total_amount": "B6",
    "data_start_row": 6,
    "material_name_col": "C",
    "material_lot_col": "D",
    "ratio_col": "E",
    "theory_amount_col": "F",
    "actual_amount_col": "G",
}

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fmt_amt(value: Any) -> str:
    """총량 표시용 숫자 포맷 — 정수는 소수점 제거, 그 외 2자리."""
    if value is None or value == "":
        return "?"
    try:
        f = float(value)
    except (TypeError, ValueError):
        return str(value)
    return str(int(f)) if f == int(f) else f"{f:.2f}"


def rescale_summary_line(record: dict[str, Any]) -> str:
    """증량(rescale) 이력을 공식 DHR 비고에 실을 한 줄 요약으로 반환(없으면 빈 문자열).

    예) "증량 2회: 1000→1050 (승인: 홍길동); 1050→1100 (부재: 야간 단독)".
    record 는 get_blend_record 가 실어 준 rescale_events_json(정규화 이벤트 JSON)을 쓴다.
    """
    raw = record.get("rescale_events_json")
    if not raw or not record.get("rescale_count"):
        return ""
    try:
        events = json.loads(raw)
    except (ValueError, TypeError):
        return ""
    if not events:
        return ""
    parts: list[str] = []
    for ev in events:
        before = _fmt_amt(ev.get("before_total"))
        after = _fmt_amt(ev.get("after_total"))
        if ev.get("approver"):
            who = f"승인: {ev['approver']}"
        elif ev.get("absence_reason"):
            who = f"부재: {ev['absence_reason']}"
        else:
            who = "승인 없음"
        parts.append(f"{before}→{after} ({who})")
    return f"증량 {len(events)}회: " + "; ".join(parts)


def build_official_dhr_xlsx(
    record: dict[str, Any],
    *,
    include_work_time: bool = True,
    signature_image_path: str | None = None,
    sign_failed: bool = False,
) -> bytes:
    """배합 기록 dict → 원료배합일지 공식 양식 xlsx 바이트.

    record: get_blend_record 반환(product_lot/worker/work_date/work_time/scale/
            total_amount/details[material_name,material_lot,ratio,theory_amount,actual_amount]).
    signature_image_path: 결재 도장(image.jpeg+서명) 이미지를 G2 셀에 삽입(원본 ExcelWriter 동일).
    sign_failed: 서명 합성을 요청했으나 실패한 경우 True — 무언(silent)의 미서명 출력 대신
                 결재칸에 '(서명 합성 실패)' 표식을 남긴다(POLISH-6).
    """
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    m = CELL_MAPPING

    # 인쇄 여백 축소 + 가로 가운데 — 흰 테두리(여백) 최소화
    try:
        ws.page_margins.left = ws.page_margins.right = 0.25
        ws.page_margins.top = ws.page_margins.bottom = 0.3
        ws.page_margins.header = ws.page_margins.footer = 0.2
        ws.print_options.horizontalCentered = True
    except Exception:
        pass

    ws[m["date"]] = f"작업일: {record.get('work_date', '')}"
    ws[m["scale"]] = f"저울: {record.get('scale') or ''}"
    ws[m["worker"]] = f"작업자 : {record.get('worker', '')}"
    if include_work_time:
        ws[m["work_time"]] = f"작업시간 : {record.get('work_time') or ''}"
    ws[m["product_lot"]] = record.get("product_lot", "")
    ws[m["total_amount"]] = (record.get("total_amount") or 0) / 100

    details = record.get("details", []) or []
    start = m["data_start_row"]
    for i, d in enumerate(details):
        row = start + i
        ws[f"{m['material_name_col']}{row}"] = d.get("material_name", "")
        ws[f"{m['material_lot_col']}{row}"] = d.get("material_lot") or ""
        ws[f"{m['ratio_col']}{row}"] = d.get("ratio")
        ws[f"{m['theory_amount_col']}{row}"] = d.get("theory_amount")
        ws[f"{m['actual_amount_col']}{row}"] = d.get("actual_amount")

    end_row = start + max(len(details), 1) - 1

    # 데이터 이후 빈 행 삭제 — 양식 24행 중 채운 만큼만 남김 (원본 _delete_empty_rows)
    for row_num in range(ws.max_row, end_row, -1):
        if all(ws.cell(row=row_num, column=col).value in (None, "") for col in range(1, 8)):
            ws.delete_rows(row_num, 1)

    # 제품 LOT(A)·배합량100g(B)을 데이터 행 전체에 걸쳐 병합 (원본 동작)
    if end_row > start:
        ws.merge_cells(f"A{start}:A{end_row}")
        ws.merge_cells(f"B{start}:B{end_row}")
    ws[f"A{start}"].alignment = _CENTER
    ws[f"B{start}"].alignment = _CENTER

    # 데이터 영역 테두리·정렬
    for row in range(start, end_row + 1):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = _BORDER
            cell.alignment = _CENTER

    # 인쇄 영역을 채운 표까지로 축소 (빈 공간 제거)
    print_end_row = end_row
    ws.print_area = f"A1:G{end_row}"

    # 결재 도장 이미지를 G2 셀에 삽입 (원본 ExcelWriter: 228x65, anchor G2)
    if signature_image_path and os.path.exists(signature_image_path):
        from openpyxl.drawing.image import Image as XLImage
        stamp = XLImage(signature_image_path)
        stamp.width = 228
        stamp.height = 65
        stamp.anchor = "G2"
        ws.add_image(stamp)
    elif sign_failed:
        # 서명 합성 실패를 표면화 — 사용자가 서명본으로 오인해 배포하는 것을 막는다(POLISH-6).
        ws["G2"] = "(서명 합성 실패)"

    # 증량(rescale) 이력을 표 아래 비고 영역에 한 줄로 남긴다(GAP-5). 공식 양식 레이아웃은
    # 건드리지 않고 표 하단(빈 공간)에 append 만 한다.
    summary = rescale_summary_line(record)
    if summary:
        note_row = print_end_row + 2
        cell = ws.cell(row=note_row, column=1, value=summary)
        cell.alignment = _LEFT
        try:
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
        except Exception:
            pass
        ws.print_area = f"A1:G{note_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
