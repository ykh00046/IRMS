"""Excel reading for attendance files.

Cell extraction, header→column-map construction, workbook loading, and the
row→record translator. ``_row_to_record`` delegates issue computation to
``anomaly._row_issue_labels`` via submodule attribute access so that
``patch.object(attendance_excel.anomaly, "_unprocessed_row_issues", ...)``
affects parsing.
"""

from __future__ import annotations

import datetime as _dt
import logging
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
    openpyxl = None  # type: ignore[assignment]
    InvalidFileException = Exception  # type: ignore[assignment,misc]
    _OPENPYXL_IMPORT_ERROR = exc
else:
    _OPENPYXL_IMPORT_ERROR = None

from . import anomaly
from . import files
from .models import AttendanceRow
from .models import (
    FileLocked,
    FileFormatInvalid,
    MonthFileNotFound,
)
from .models import (
    _HEADER_REQUIRED_FIELDS,
    _HEADER_SIMPLE_FIELDS,
    _HEADER_WORKHOUR_SUFFIX,
)

_LOGGER = logging.getLogger(__name__)

DAY_TYPE_VALUES = ("평일", "평일2", "주휴", "무휴", "유휴")

# 헤더에서 검출되길 기대하는 선택(비필수) 필드. 필수는 다 잡혔는데 이들이
# 헤더에서 안 잡히면 구 기본 인덱스로 조용히 폴백되므로 경고 대상이다(GAP-1).
_OPTIONAL_HEADER_FIELDS = frozenset(files.DEFAULT_COLUMNS) - _HEADER_REQUIRED_FIELDS


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M")
    # 숫자형 셀 정규화: ERP가 사번 등을 숫자로 내보내면 openpyxl이 float
    # (171013.0)로 읽는다. 정수값 실수는 정수 문자열로 낮춰 로그인 사번
    # 문자열과 어긋나지 않게 한다(BUG-2). 진짜 소수/그 외 값은 그대로.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_at(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _cell_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _cell_time(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, _dt.datetime):
        return value.strftime("%H:%M")
    text = str(value).strip()
    return text or None


def _build_column_map(
    header_group: tuple[Any, ...] | None,
    header_sub: tuple[Any, ...] | None,
) -> tuple[dict[str, int], list[str], set[str]]:
    """헤더 두 행에서 ``(열맵, 경고목록, 헤더에서 잡힌 필드)`` 를 만든다.

    필수 필드를 모두 찾지 못하면 ``DEFAULT_COLUMNS`` 로 통째 폴백한다(경고
    없음 — 헤더 없는 단위테스트/구버전 레이아웃의 정상 시나리오). 필수는
    잡혔지만 알려진 선택 열(예 ``외출시간``·``조출``)이 헤더에서 안 잡혀 구
    기본 인덱스로 조용히 폴백되면 경고를 남긴다(GAP-1). 어느 경우에도 파싱은
    계속한다.
    """
    if not header_sub:
        return dict(files.DEFAULT_COLUMNS), [], set()

    group = list(header_group or ())
    sub = list(header_sub)

    # 병합된 그룹 헤더(평일근무시간/휴일근무시간 등)를 오른쪽으로 채운다.
    filled_group: list[str] = []
    last = ""
    for cell in group:
        text = _cell_str(cell)
        if text:
            last = text
        filled_group.append(last)

    detected: dict[str, int] = {}
    width = max(len(sub), len(filled_group))
    for idx in range(width):
        sub_name = _cell_str(sub[idx]) if idx < len(sub) else ""
        grp_name = filled_group[idx] if idx < len(filled_group) else ""

        field = _HEADER_SIMPLE_FIELDS.get(sub_name)
        if field and field not in detected:
            detected[field] = idx
            continue

        if sub_name in _HEADER_WORKHOUR_SUFFIX:
            suffix = _HEADER_WORKHOUR_SUFFIX[sub_name]
            if "평일" in grp_name:
                prefix = "weekday"
            elif "휴일" in grp_name:
                prefix = "holiday"
            else:
                # 2026-08 부터 ERP 가 그룹행의 평일/휴일 라벨을 없앴다. 두 묶음의
                # 순서(평일 앞, 휴일 뒤)는 구 양식·고정 인덱스와 동일함을 실측으로
                # 확인했으므로, 라벨이 없으면 등장 순서로 구분한다.
                prefix = "weekday" if f"weekday_{suffix}" not in detected else "holiday"
            detected.setdefault(f"{prefix}_{suffix}", idx)
            continue

        # 비고는 세부 행이 비어 있고 그룹 행에만 '비고'로 존재한다.
        if not sub_name and "비고" in grp_name and "note" not in detected:
            detected["note"] = idx

    if not _HEADER_REQUIRED_FIELDS.issubset(detected):
        return dict(files.DEFAULT_COLUMNS), [], set()

    warnings: list[str] = []
    missing_optional = sorted(_OPTIONAL_HEADER_FIELDS - detected.keys())
    if missing_optional:
        # 실제 헤더 텍스트를 함께 남긴다 — 못 찾은 이유(ERP 가 이름을 뭐라고
        # 바꿨는지)를 로그만 보고 알 수 있어야 스크린샷 왕복 없이 규칙을 고친다.
        group_texts = " | ".join(dict.fromkeys(t for t in filled_group if t))
        sub_texts = " | ".join(t for t in (_cell_str(c) for c in sub) if t)
        warnings.append(
            "선택 열을 헤더에서 찾지 못해 구 기본 인덱스로 폴백: "
            + ", ".join(missing_optional)
            + f" (그룹행 헤더: {group_texts} / 세부행 헤더: {sub_texts})"
        )
    # 세 번째 값은 '헤더에서 실제로 잡힌' 키다 — 병합된 맵은 늘 전 필드를 갖고 있어
    # 무엇이 폴백됐는지 구분할 수 없다(진단이 늘 정상으로 보였다).
    return {**files.DEFAULT_COLUMNS, **detected}, warnings, set(detected)


def _make_column_map(
    header_group: tuple[Any, ...] | None,
    header_sub: tuple[Any, ...] | None,
) -> dict[str, int]:
    """헤더 두 행(그룹 행0 + 세부 행1)에서 논리 필드 → 열 인덱스 맵을 만든다.

    헤더에서 필수 필드를 모두 찾지 못하면 ``DEFAULT_COLUMNS`` 를 그대로
    돌려준다(구버전 레이아웃·헤더 없는 입력에 대한 안전한 폴백). 선택 열이
    헤더에서 안 잡혀 기본 인덱스로 폴백되면 서버 로그로 한 번 경고한다(GAP-1).
    """
    colmap, warnings, _detected = _build_column_map(header_group, header_sub)
    for message in warnings:
        _LOGGER.warning("attendance 헤더 매핑: %s", message)
    return colmap


def _cell_day_type(row: tuple[Any, ...], colmap: dict[str, int]) -> str:
    day_type_idx = colmap["day_type"]
    primary = _cell_str(_cell_at(row, day_type_idx))
    if primary in DAY_TYPE_VALUES:
        return primary
    shifted = _cell_str(_cell_at(row, day_type_idx + 1))
    if shifted in DAY_TYPE_VALUES:
        return shifted
    return primary or shifted


def _load_workbook(path: Path):
    if not path.exists():
        raise MonthFileNotFound(str(path))
    if openpyxl is None:
        raise RuntimeError(
            "openpyxl is required to load attendance Excel files. "
            "Install test dependencies with `pip install -r requirements-dev.txt`."
        ) from _OPENPYXL_IMPORT_ERROR
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError as exc:  # pragma: no cover - windows-specific
        raise FileLocked(str(path)) from exc
    except InvalidFileException as exc:
        raise FileFormatInvalid(str(path)) from exc


def _iter_data_rows(ws, colmap: dict[str, int] | None = None):
    date_idx = (colmap or files.DEFAULT_COLUMNS)["date"]
    emp_idx = (colmap or files.DEFAULT_COLUMNS)["emp_id"]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx < 2:
            continue
        if not row:
            continue
        if _cell_at(row, date_idx) is None and _cell_at(row, emp_idx) is None:
            continue
        yield row


def _row_to_record(
    row: tuple[Any, ...], colmap: dict[str, int] | None = None
) -> dict[str, Any]:
    cols = colmap or files.DEFAULT_COLUMNS
    emp_id = _cell_str(_cell_at(row, cols["emp_id"]))
    if not emp_id:
        return {}
    shift_time = _cell_str(_cell_at(row, cols["shift_time"]))
    row_data = AttendanceRow(
        date=_cell_str(_cell_at(row, cols["date"])),
        weekday=_cell_str(_cell_at(row, cols["weekday"])),
        day_type=_cell_day_type(row, cols),
        check_in=_cell_time(_cell_at(row, cols["check_in"])),
        check_out=_cell_time(_cell_at(row, cols["check_out"])),
        next_day=bool(_cell_float(_cell_at(row, cols["next_day"]))),
        weekday_early=_cell_float(_cell_at(row, cols["weekday_early"])),
        weekday_normal=_cell_float(_cell_at(row, cols["weekday_normal"])),
        weekday_overtime=_cell_float(_cell_at(row, cols["weekday_overtime"])),
        weekday_night=_cell_float(_cell_at(row, cols["weekday_night"])),
        holiday_early=_cell_float(_cell_at(row, cols["holiday_early"])),
        holiday_normal=_cell_float(_cell_at(row, cols["holiday_normal"])),
        holiday_overtime=_cell_float(_cell_at(row, cols["holiday_overtime"])),
        holiday_night=_cell_float(_cell_at(row, cols["holiday_night"])),
        late_hours=_cell_float(_cell_at(row, cols["late"])),
        early_leave_hours=_cell_float(_cell_at(row, cols["early_leave"])),
        outing_hours=_cell_float(_cell_at(row, cols["outing"])),
        note=_cell_str(_cell_at(row, cols["note"])),
        attendance_code=_cell_str(_cell_at(row, cols["attendance_code"])),
    )
    row_data.issues = anomaly._row_issue_labels(row_data, shift_time)
    row_data.has_issue = bool(row_data.issues)
    return {
        "emp_id": emp_id,
        "name": _cell_str(_cell_at(row, cols["name"])),
        "department": _cell_str(_cell_at(row, cols["department"])),
        "factory": _cell_str(_cell_at(row, cols["factory"])),
        "shift_time": shift_time,
        "shift_group": _cell_str(_cell_at(row, cols["shift_group"])),
        "job_type": _cell_str(_cell_at(row, cols["job_type"])),
        "gender": _cell_str(_cell_at(row, cols["gender"])),
        "row": row_data,
    }


def _column_map_from_ws(ws) -> dict[str, int]:
    """워크시트의 헤더 두 행으로 열 맵을 만든다.

    헤더를 읽을 수 없으면(예: 헤더가 없는 입력) ``DEFAULT_COLUMNS`` 폴백.
    """
    try:
        header_rows = list(ws.iter_rows(values_only=True, max_row=2))
    except (AttributeError, TypeError):
        return dict(files.DEFAULT_COLUMNS)
    if len(header_rows) < 2:
        return dict(files.DEFAULT_COLUMNS)
    return _make_column_map(header_rows[0], header_rows[1])


def header_diagnostics(year_month: str) -> list[dict[str, Any]]:
    """월 파일별 헤더 인식 상태 — 책임자에게 보여 주기 위한 진단.

    ERP 는 2026-06 에 신원/근무정보 블록의 열 순서를 바꿔 내보내기 시작했고, 고정
    인덱스로 읽던 코드가 이름을 '근무공장' 값으로 잡는 등 조용히 어긋났다. 헤더 자동
    매핑이 그 대책이고 폴백 경고가 안전망인데, 그 경고는 서버 로그에만 남는다 — 로그를
    보는 사람이 없으면 안전망이 아니다(2026-08-08). 이 함수는 데이터 행은 건드리지 않고
    헤더 두 줄만 다시 읽어 상태를 돌려준다(조회 경로에 영향 없음).

    반환: [{file, ok, fallback, missing_optional:[...]}]
      fallback=True  → 필수 열을 못 찾아 파일 전체를 구 기본 인덱스로 읽고 있다(위험).
      missing_optional → 그 열들만 기본 인덱스로 폴백했다(값이 어긋날 수 있다).
    """
    out: list[dict[str, Any]] = []
    for path in files.month_file_paths(year_month):
        entry: dict[str, Any] = {
            "file": path.name, "ok": True, "fallback": False, "missing_optional": [],
        }
        try:
            wb = _load_workbook(path)
        except (MonthFileNotFound, FileLocked, FileFormatInvalid) as exc:
            entry.update(ok=False, error=type(exc).__name__)
            out.append(entry)
            continue
        try:
            ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
            header_rows = list(ws.iter_rows(values_only=True, max_row=2))
        except Exception:  # noqa: BLE001 — 진단이 조회를 막지 않는다
            entry.update(ok=False, error="HEADER_UNREADABLE")
            out.append(entry)
            continue
        finally:
            try:
                wb.close()
            except Exception:  # noqa: BLE001
                pass
        if len(header_rows) < 2:
            entry.update(fallback=True)
            out.append(entry)
            continue
        _colmap, warnings, detected = _build_column_map(header_rows[0], header_rows[1])
        # 필수 열을 못 찾으면 _build_column_map 이 DEFAULT_COLUMNS 를 통째로 돌려준다
        # (경고 없이) — 그 경우를 폴백으로 표면화한다.
        # 필수 열을 못 찾으면 잡힌 필드가 하나도 없다(통째 폴백).
        entry["fallback"] = not detected
        entry["missing_optional"] = sorted(_OPTIONAL_HEADER_FIELDS - detected)
        entry["ok"] = not entry["fallback"] and not entry["missing_optional"]
        out.append(entry)
    return out


# (path 문자열) → (mtime, records) 파싱 캐시. 종전에는 요청마다 워크북을 다시
# 열었는데(명시적 무캐시 설계), 화면 한 번에 최대 12개월 파일을 읽는 구조라 비용이
# 컸다 — 책임자 진입 1회에 워크북 15개+(2026-08-14 검토 5번). ERP 야간 배치가 파일을
# 덮어쓰면 mtime 이 바뀌어 자동 무효화된다(erp_lot_service 의 캐시와 같은 규칙).
# 반환 records 는 공유 객체이므로 호출부는 읽기 전용으로만 다룬다(현재 전 호출부 확인).
_PARSE_CACHE: dict[str, tuple[tuple[int, int], list[dict[str, Any]]]] = {}


def reset_parse_cache() -> None:
    """테스트용 — 파싱 캐시 초기화(파일을 같은 스탬프로 다시 쓰는 픽스처 대비)."""
    _PARSE_CACHE.clear()


def _cache_stamp(path: Path) -> tuple[int, int] | None:
    """캐시 무효화 키 — ns 단위 mtime + 크기. 초 단위 mtime 만 보면 같은 초 안에
    파일을 다시 쓰는 테스트 픽스처(또는 빠른 연속 배치)가 낡은 캐시를 받는다."""
    try:
        st = path.stat()
    except OSError:
        return None
    return (st.st_mtime_ns, st.st_size)


def _records_from_path(path: Path) -> list[dict[str, Any]]:
    key = str(path)
    stamp = _cache_stamp(path)
    if stamp is not None:
        cached = _PARSE_CACHE.get(key)
        if cached is not None and cached[0] == stamp:
            return cached[1]
    wb = _load_workbook(path)
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        colmap = _column_map_from_ws(ws)
        records: list[dict[str, Any]] = []
        for raw in _iter_data_rows(ws, colmap):
            rec = _row_to_record(raw, colmap)
            if rec:
                records.append(rec)
        if stamp is not None:
            _PARSE_CACHE[key] = (stamp, records)
        return records
    finally:
        wb.close()
