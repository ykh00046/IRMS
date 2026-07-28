"""ERP 원재료 LOT 검증 서비스.

C:/ErpExcel/ERP_YYYY-MM-DD.xlsx (단일 시트 'Sheet') 를 읽어 자재별 유효 LOT 목록을
만든다. 배합 화면은 무로그인 공용 단말이므로, 이 서비스는 엑셀 파일이 없거나 읽기에
실패해도 현장을 막지 않는다(fail-open — valid=True, file_ok=False).

파일 구조(헤더 14열):
    창고 · 구분 · 품목코드 · 품목명 · 대분류 · 중분류 · 규격 · Lot.No
    · 기초 · 입고 · 출고 · 재고 · 검사대기 · 단위
마지막 행은 합계행(창고 열이 비어 있음) — load_lots 가 버린다.

캐시: (경로, mtime) 단위 모듈 캐시. 같은 파일이면 재파싱하지 않는다.
"""

import datetime as _dt
import glob
import os
import sqlite3
from typing import Any

from openpyxl import load_workbook

# 헤더 순서(0-base) — 실제 ERP 파일 기준. 창고 열이 비었으면 합계/빈행으로 본다.
_COL_WAREHOUSE = 0    # 창고
_COL_GUBUN = 1        # 구분
_COL_CODE = 2         # 품목코드
_COL_NAME = 3         # 품목명
_COL_LOT = 7          # Lot.No
_COL_STOCK = 11       # 재고

# 모듈 캐시: {(path, mtime): {"file_name", "file_date", "read_at", "lots"}}
# lots 구조: {code(str): {"name": str, "lots": {lot(str): {"stock": float, "gubun": str}}}}
_CACHE: dict[tuple[str, float], dict[str, Any]] = {}


def _erp_dir() -> str:
    return os.environ.get("IRMS_ERP_EXCEL_DIR") or r"C:\ErpExcel"


def _parse_file_date(file_name: str) -> _dt.date | None:
    """ERP_YYYY-MM-DD.xlsx 파일명에서 날짜 추출. 파싱 불가면 None."""
    base = os.path.basename(file_name)
    # ERP_YYYY-MM-DD.xlsx → YYYY-MM-DD
    if not base.startswith("ERP_") or not base.endswith(".xlsx"):
        return None
    stamp = base[len("ERP_"):-len(".xlsx")]
    try:
        return _dt.datetime.strptime(stamp, "%Y-%m-%d").date()
    except ValueError:
        return None


def latest_erp_file() -> str | None:
    """ERP_*.xlsx 중 파일명 날짜(YYYY-MM-DD)가 가장 최근인 것. 없으면 None.

    날짜 파싱이 안 되는 파일명(예: ERP_기타.xlsx)은 무시한다.
    """
    pattern = os.path.join(_erp_dir(), "ERP_*.xlsx")
    best: tuple[_dt.date, str] | None = None
    for path in glob.glob(pattern):
        d = _parse_file_date(path)
        if d is None:
            continue
        if best is None or d > best[0]:
            best = (d, path)
    return best[1] if best else None


def _to_float(value: Any) -> float:
    """재고 숫자 변환 — None/빈칸/문자는 0.0."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _parse_file(path: str, mtime: float) -> dict[str, Any]:
    """단일 ERP 파일을 파싱해 캐시값을 만든다. 예외는 호출자가 처리(fail-open)."""
    file_name = os.path.basename(path)
    file_date = _parse_file_date(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        # 첫 행(헤더) 건너뛴다.
        try:
            next(rows)
        except StopIteration:
            rows = iter(())
        lots: dict[str, dict[str, Any]] = {}
        for row in rows:
            if not row:
                continue
            # 창고 열이 비었으면 합계/빈행 → 버린다.
            warehouse = row[_COL_WAREHOUSE] if len(row) > _COL_WAREHOUSE else None
            if warehouse is None or str(warehouse).strip() == "":
                continue
            # Lot.No 가 비었거나 '*' 이면 무시(유효 LOT 가 아님).
            lot_raw = row[_COL_LOT] if len(row) > _COL_LOT else None
            if lot_raw is None:
                continue
            lot = str(lot_raw).strip()
            if lot == "" or lot == "*":
                continue
            code_raw = row[_COL_CODE] if len(row) > _COL_CODE else None
            if code_raw is None:
                continue
            code = str(code_raw).strip()
            if code == "":
                continue
            name_raw = row[_COL_NAME] if len(row) > _COL_NAME else None
            name = str(name_raw).strip() if name_raw is not None else ""
            gubun_raw = row[_COL_GUBUN] if len(row) > _COL_GUBUN else None
            gubun = str(gubun_raw).strip() if gubun_raw is not None else ""
            stock = _to_float(row[_COL_STOCK] if len(row) > _COL_STOCK else None)
            entry = lots.setdefault(
                code, {"name": name, "lots": {}}
            )
            # 품목명은 가장 마지막에 본 값으로 갱신(비어있으면 기존 유지).
            if name:
                entry["name"] = name
            entry["lots"][lot] = {"stock": stock, "gubun": gubun}
    finally:
        wb.close()
    return {
        "file_name": file_name,
        "file_date": file_date.isoformat() if file_date else None,
        "read_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "lots": lots,
    }


def load_lots() -> dict[str, Any] | None:
    """최신 ERP 파일을 (경로, mtime) 캐시로 파싱. 파일이 없거나 실패면 None.

    반환: {"file_name", "file_date", "read_at", "lots"} 또는 None.
    """
    path = latest_erp_file()
    if path is None or not os.path.exists(path):
        return None
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return None
    key = (path, mtime)
    cached = _CACHE.get(key)
    if cached is not None:
        return cached
    try:
        parsed = _parse_file(path, mtime)
    except Exception:
        # 엑셀 읽기 실패 — fail-open. 캐시하지 않고 None.
        return None
    # 디렉터리가 바뀌면 옛 키가 남지 않도록 캐시를 비운다(단일 최신 파일 가정).
    _CACHE.clear()
    _CACHE[key] = parsed
    return parsed


def reset_cache() -> None:
    """테스트용 — 캐시를 비운다(monkeypatch 로 디렉터리를 바꾼 뒤 호출)."""
    _CACHE.clear()


def _manual_lot_exists(connection: sqlite3.Connection, code: str, lot: str) -> bool:
    """manual_material_lots 테이블에 (code, lot) 이 있는지."""
    try:
        row = connection.execute(
            "SELECT 1 FROM manual_material_lots "
            "WHERE material_code = ? AND lot = ? LIMIT 1",
            (code, lot),
        ).fetchone()
    except sqlite3.OperationalError:
        # 마이그 전 DB(테이블 없음) — 수동 LOT 없음으로 처리.
        return False
    return row is not None


def check_lot(connection: sqlite3.Connection, code: str, lot: str) -> dict[str, Any]:
    """자재 코드/LOT 검증 → {valid, source, stock, file_name, file_ok}.

    규칙:
      - 수동 LOT 테이블에 (code,lot) → valid=True source='manual'
      - 엑셀에 있고 재고>0 → valid=True source='erp'
      - 엑셀에 있는데 재고<=0 → valid=False source='erp'(소진)
      - 엑셀에 없음 → valid=False source=None
      - 파일 없음/읽기 실패 → valid=True file_ok=False (fail-open)
    LOT 비교는 양끝 공백 strip 후 정확 일치.
    """
    norm_code = (code or "").strip()
    norm_lot = (lot or "").strip()

    # 파일이 아예 없거나 읽기 실패 — 엑셀 문제로 현장을 막지 않는다(fail-open).
    parsed = load_lots()
    if parsed is None:
        return {
            "valid": True,
            "source": None,
            "stock": None,
            "file_name": None,
            "file_ok": False,
        }

    # 수동 LOT 우선 — 즉석 책임자 인증으로 추가한 LOT 는 엑셀 재고와 무관.
    if _manual_lot_exists(connection, norm_code, norm_lot):
        return {
            "valid": True,
            "source": "manual",
            "stock": None,
            "file_name": parsed["file_name"],
            "file_ok": True,
        }

    entry = parsed["lots"].get(norm_code)
    if entry is None:
        return {
            "valid": False,
            "source": None,
            "stock": None,
            "file_name": parsed["file_name"],
            "file_ok": True,
        }
    lot_entry = entry["lots"].get(norm_lot)
    if lot_entry is None:
        return {
            "valid": False,
            "source": None,
            "stock": None,
            "file_name": parsed["file_name"],
            "file_ok": True,
        }
    stock = float(lot_entry.get("stock") or 0.0)
    if stock > 0:
        return {
            "valid": True,
            "source": "erp",
            "stock": stock,
            "file_name": parsed["file_name"],
            "file_ok": True,
        }
    return {
        "valid": False,
        "source": "erp",
        "stock": stock,
        "file_name": parsed["file_name"],
        "file_ok": True,
    }


def list_manual_lots(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    """수동 LOT 전체 목록(id 포함) — status 병합·삭제용."""
    try:
        rows = connection.execute(
            "SELECT id, material_code, lot, note, created_by, created_at "
            "FROM manual_material_lots ORDER BY id"
        ).fetchall()
    except sqlite3.OperationalError:
        return []
    return [
        {
            "id": int(r["id"]),
            "material_code": r["material_code"],
            "lot": r["lot"],
            "note": r["note"],
            "created_by": r["created_by"],
            "created_at": r["created_at"],
        }
        for r in rows
    ]
