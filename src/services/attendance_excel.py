"""Excel parser for monthly attendance files.

Source: ``C:\\ErpExcel\\monthly_attendance_YYYY-MM.xlsx`` (ERP exports nightly
at 18:00). The file is read on every request because it is small (~85KB for
~500 rows) and using a cache would risk showing stale data right after the
nightly refresh.

Columns (0-index on Sheet1, headers are on rows 0-1 so data starts at row 2):

    0  근무일자           11  부서명              22  평일 정상
    1  요일               12  근무조구분          23  평일 연장 (급여 1.5배)
    2  구분 (평일/휴일 등) 14  근무타임            24  평일 야근 (급여 1.5배)
    4  사번               17  출근 HH:MM          25  휴일 조출
    6  남여               18  퇴근 HH:MM          26  휴일 정상
    7  근무공장           19  익일 (0/1)          27  휴일 연장 (급여 1.5배)
    8  성명               20  총시간 (표시 제외)  28  휴일 야근 (급여 1.5배)
    10 근무직구분         21  평일 조출           29  지각시간
                                                  30  조퇴시간
                                                  31  외출시간
                                                  33  비고
"""

from __future__ import annotations

import datetime as _dt
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
    openpyxl = None  # type: ignore[assignment]
    InvalidFileException = Exception  # type: ignore[assignment,misc]
    _OPENPYXL_IMPORT_ERROR = exc
else:
    _OPENPYXL_IMPORT_ERROR = None

ATTENDANCE_DIR = Path(r"C:\ErpExcel")
FILENAME_PATTERN = "monthly_attendance_{year_month}.xlsx"
FILENAME_REGEX = re.compile(r"^monthly_attendance(?:_.+)?_(\d{4}-\d{2})\.xlsx$")
ANNUAL_LEAVE_KEYWORDS = ("연차", "월차", "휴가", "반차", "유급", "공가")
HALF_DAY_LEAVE_KEYWORDS = ("반차", "오전", "오후")

# 부분 휴가가 출/퇴근 기준 시각을 이동시키는 시간(시간 단위, 주간 9-18시 기준).
# 반차는 근무 4시간 + 점심 1시간을 덜어내므로 baseline 이 5시간 이동하고,
# 반반차는 2시간짜리라 점심과 겹치지 않아 2시간만 이동한다. 반반차가 반차의
# 부분 문자열이므로 탐지 시에는 반드시 "반반차"를 먼저 검사해야 한다.
#
# 검증 (주간 기준):
#   반차   오전 → 09:00 + 5h = 14:00 출근,  오후 → 18:00 - 5h = 13:00 퇴근
#   반반차 오전 → 09:00 + 2h = 11:00 출근,  오후 → 18:00 - 2h = 16:00 퇴근
PARTIAL_LEAVE_SHIFT_HOURS = (("반반차", 2), ("반차", 5))
PARTIAL_LEAVE_SHIFT_HOURS_BY_SHIFT = {
    "주간": {"반반차": 2, "반차": 5},
    "2교대(주간)": {"반반차": 2, "반차": 8},
    "2교대(야간)": {"반반차": 2, "반차": 8},
}

# Baseline (출근, 퇴근) per shift_time, expressed in minutes-from-midnight.
# 퇴근 baseline이 1440 이상이면 익일 표기(ERP는 19:00 → 다음 날 07:00 근무를
# 19:00 → 31:00으로 기록함). 빈 shift_time은 교대조의 비번 날을 뜻하므로
# anomaly 감지 대상에서 제외된다.
SHIFT_BASELINES = {
    "주간":        (9 * 60,  18 * 60),   # 09:00 ~ 18:00 (평일만)
    "2교대(주간)": (7 * 60,  19 * 60),   # 07:00 ~ 19:00
    "2교대(야간)": (19 * 60, 31 * 60),   # 19:00 ~ 익일 07:00 (ERP 표기 31:00)
}
ALERT_WORKDAY_TYPES = ("평일", "평일2")
DAY_TYPE_VALUES = ("평일", "평일2", "주휴", "무휴", "유휴")
DAY_SHIFT_SHORT_LUNCH_DAY_TYPE = "평일2"
DAY_SHIFT_SHORT_LUNCH_CHECKOUT_OFFSET_MINUTES = 30
ALERT_GRACE_MINUTES = 15

# Full-day leaves that exclude a row from anomaly detection entirely.
# (반차/반반차는 이 목록에 들어가지 않는다 — baseline이 이동할 뿐 여전히
# 감지 대상.)
# "휴직"은 육아휴직·병가휴직 등 장기 전일 부재를 포괄한다. 휴직 기간에는
# 출근/퇴근 기록이 없는 것이 정상이므로 "출근 누락" 등으로 잡으면 안 된다.
# "결근"도 근태코드로 명시되면 담당자가 이미 처리한 전일 부재이므로,
# 출근/퇴근 기록이 없는 것이 정상 — "미타각/출근 누락"으로 잡지 않는다.
FULL_DAY_LEAVE_KEYWORDS = ("연차", "월차", "휴가", "휴직", "유급", "공가", "훈련", "예비군", "교육", "결근")
ATTENDANCE_ISSUE_CODE_KEYWORDS = ("미타각", "누락")

COL_DATE = 0
COL_WEEKDAY = 1
COL_DAY_TYPE = 2
COL_EMP_ID = 4
COL_GENDER = 6
COL_FACTORY = 7
COL_NAME = 8
COL_JOB_TYPE = 10
COL_DEPARTMENT = 11
COL_SHIFT_GROUP = 12
COL_SHIFT_TIME = 14
COL_ATTENDANCE_CODE = 16
COL_CHECK_IN = 17
COL_CHECK_OUT = 18
COL_NEXT_DAY = 19
COL_WD_EARLY = 21
COL_WD_NORMAL = 22
COL_WD_OVERTIME = 23
COL_WD_NIGHT = 24
COL_HD_EARLY = 25
COL_HD_NORMAL = 26
COL_HD_OVERTIME = 27
COL_HD_NIGHT = 28
COL_LATE = 29
COL_EARLY_LEAVE = 30
COL_OUTING = 31
COL_NOTE = 33

# 위 COL_* 상수는 2026-05 까지의 ERP 내보내기 열 순서를 기준으로 한다.
# 2026-06 부터 ERP가 신원/근무정보 블록(성명·근무타임·부서명 등)의 열 순서를
# 바꿔 내보내기 시작했다. 고정 인덱스로 읽으면 이름이 '근무공장' 값으로,
# 근무타임이 엉뚱한 값으로 잡혀 근태 이상 감지가 무력화된다. 그래서 실제
# 파일을 읽을 때는 헤더 행에서 열 위치를 자동으로 찾고(_make_column_map),
# 헤더를 못 찾으면 아래 기본값으로 폴백한다. (단위 테스트는 헤더 없이 위치
# 기반으로 행을 만들어 _row_to_record 에 직접 넘기므로 이 기본값을 쓴다.)
DEFAULT_COLUMNS: dict[str, int] = {
    "date": COL_DATE,
    "weekday": COL_WEEKDAY,
    "day_type": COL_DAY_TYPE,
    "emp_id": COL_EMP_ID,
    "gender": COL_GENDER,
    "factory": COL_FACTORY,
    "name": COL_NAME,
    "job_type": COL_JOB_TYPE,
    "department": COL_DEPARTMENT,
    "shift_group": COL_SHIFT_GROUP,
    "shift_time": COL_SHIFT_TIME,
    "attendance_code": COL_ATTENDANCE_CODE,
    "check_in": COL_CHECK_IN,
    "check_out": COL_CHECK_OUT,
    "next_day": COL_NEXT_DAY,
    "weekday_early": COL_WD_EARLY,
    "weekday_normal": COL_WD_NORMAL,
    "weekday_overtime": COL_WD_OVERTIME,
    "weekday_night": COL_WD_NIGHT,
    "holiday_early": COL_HD_EARLY,
    "holiday_normal": COL_HD_NORMAL,
    "holiday_overtime": COL_HD_OVERTIME,
    "holiday_night": COL_HD_NIGHT,
    "late": COL_LATE,
    "early_leave": COL_EARLY_LEAVE,
    "outing": COL_OUTING,
    "note": COL_NOTE,
}

# 헤더 행1(세부 명칭)에서 바로 매핑되는 단순 컬럼.
_HEADER_SIMPLE_FIELDS = {
    "근무일자": "date",
    "요일": "weekday",
    "구분": "day_type",
    "사번": "emp_id",
    "성명": "name",
    "남여": "gender",
    "근무공장": "factory",
    "근무직구분": "job_type",
    "부서명": "department",
    "근무조구분": "shift_group",
    "근무타임": "shift_time",
    "근태코드": "attendance_code",
    "출근": "check_in",
    "퇴근": "check_out",
    "익일": "next_day",
    "지각시간": "late",
    "조퇴시간": "early_leave",
    "외출시간": "outing",
}
# 평일/휴일 근무시간 그룹은 '조출/정상/연장/야근' 세부 명칭이 두 번 나오므로
# 그룹 헤더(행0: 평일근무시간/휴일근무시간)로 구분한다.
_HEADER_WORKHOUR_SUFFIX = {"조출": "early", "정상": "normal", "연장": "overtime", "야근": "night"}
# 헤더 자동 매핑을 신뢰하려면 최소한 이 필드들이 모두 잡혀야 한다.
_HEADER_REQUIRED_FIELDS = frozenset(
    {"date", "emp_id", "name", "attendance_code", "check_in", "check_out", "shift_time", "day_type"}
)


class AttendanceError(Exception):
    """Base class for attendance parsing errors."""


class MonthFileNotFound(AttendanceError):
    pass


class FileLocked(AttendanceError):
    pass


class FileFormatInvalid(AttendanceError):
    pass


@dataclass
class AttendanceRow:
    date: str
    weekday: str
    day_type: str
    check_in: str | None
    check_out: str | None
    next_day: bool
    weekday_early: float
    weekday_normal: float
    weekday_overtime: float
    weekday_night: float
    holiday_early: float
    holiday_normal: float
    holiday_overtime: float
    holiday_night: float
    late_hours: float
    early_leave_hours: float
    outing_hours: float
    note: str
    attendance_code: str = ""
    issues: list[str] = field(default_factory=list)
    has_issue: bool = False


@dataclass
class AttendanceProfile:
    emp_id: str
    name: str
    department: str
    factory: str
    shift_time: str
    shift_group: str
    job_type: str
    gender: str


@dataclass
class AttendanceSummary:
    work_days: int = 0
    late_count: int = 0
    late_total: float = 0.0
    early_leave_count: int = 0
    early_leave_total: float = 0.0
    outing_count: int = 0
    outing_total: float = 0.0
    weekday_early: float = 0.0
    weekday_normal: float = 0.0
    weekday_overtime: float = 0.0
    weekday_night: float = 0.0
    holiday_early: float = 0.0
    holiday_normal: float = 0.0
    holiday_overtime: float = 0.0
    holiday_night: float = 0.0


@dataclass
class AttendanceAnnualSummary:
    year: int
    available_months_count: int = 0
    months_count: int = 0
    skipped_months: list[str] = field(default_factory=list)
    late_count: int = 0
    late_total: float = 0.0
    annual_leave_count: int = 0
    annual_leave_days: float = 0.0
    annual_leave_full_days: float = 0.0
    annual_leave_half_days: float = 0.0
    annual_leave_quarter_days: float = 0.0


def _canonical_month_file_path(year_month: str) -> Path:
    return ATTENDANCE_DIR / FILENAME_PATTERN.format(year_month=year_month)


def _year_month_from_filename(name: str) -> str | None:
    match = FILENAME_REGEX.match(name)
    if not match:
        return None
    return match.group(1)


def month_file_paths(year_month: str) -> list[Path]:
    if not ATTENDANCE_DIR.exists():
        return []
    canonical_name = FILENAME_PATTERN.format(year_month=year_month)
    matches = [
        entry
        for entry in ATTENDANCE_DIR.iterdir()
        if _year_month_from_filename(entry.name) == year_month
    ]
    return sorted(
        matches,
        key=lambda entry: (0 if entry.name == canonical_name else 1, entry.name.lower()),
    )


def month_file_path(year_month: str) -> Path:
    paths = month_file_paths(year_month)
    if paths:
        return paths[0]
    return _canonical_month_file_path(year_month)


def _month_file_paths_or_raise(year_month: str) -> list[Path]:
    paths = month_file_paths(year_month)
    if paths:
        return paths
    raise MonthFileNotFound(str(_canonical_month_file_path(year_month)))


def available_months() -> list[str]:
    if not ATTENDANCE_DIR.exists():
        return []
    months: set[str] = set()
    for entry in ATTENDANCE_DIR.iterdir():
        year_month = _year_month_from_filename(entry.name)
        if year_month:
            months.add(year_month)
    return sorted(months, reverse=True)


def current_year_month() -> str:
    today = _dt.date.today()
    return f"{today.year:04d}-{today.month:02d}"


def current_date() -> str:
    return _dt.date.today().isoformat()


def _hhmm_to_minutes(text: str | None) -> int | None:
    """ERP 시각 문자열을 '자정 기준 분'으로 변환.

    ERP는 야간조 퇴근을 익일 07:04처럼 ``31:04``로 기록하므로 단순
    문자열 비교가 불가능하다. 시:분을 읽어 분 단위 정수로 반환하고,
    파싱이 실패하면 ``None``을 돌려준다.
    """
    if not text:
        return None
    s = str(text).strip()
    if not s:
        return None
    try:
        h_part, _, m_part = s.partition(":")
        return int(h_part) * 60 + int(m_part or 0)
    except ValueError:
        return None


def _alert_reference_datetime() -> _dt.datetime:
    return _dt.datetime.now()


def alert_year_month() -> str:
    current = current_year_month()
    if month_file_paths(current):
        return current
    months = available_months()
    return months[0] if months else current


def _format_hours(value: float) -> str:
    return f"{value:g}"


def _is_full_day_leave(day_type: str, note: str, attendance_code: str = "") -> bool:
    """감지 대상에서 아예 제외할 '하루 전체 휴가' 판정.

    반차/반반차는 여기 포함되지 않는다. 하루 중 일부만 빠지는 휴가는
    baseline 만 이동시키고 감지는 계속 수행한다.
    """
    text = f"{day_type or ''} {note or ''} {attendance_code or ''}"
    if any(keyword in text for keyword, _hours in PARTIAL_LEAVE_SHIFT_HOURS):
        return False
    return any(keyword in text for keyword in FULL_DAY_LEAVE_KEYWORDS)


def _partial_leave_shift(
    day_type: str, note: str, attendance_code: str = ""
) -> tuple[str, str, int] | None:
    """반차·반반차 이동량(분)과 오전/오후 구분을 반환.

    반환 형식: ``(leave_kind, half, shift_minutes)``
      - leave_kind: "반차" 또는 "반반차"
      - half: "오전" / "오후" / "unknown"
      - shift_minutes: 해당 근무형태의 baseline을 이동시킬 분 단위

    해당 행에 부분 휴가 키워드가 없으면 ``None`` 반환. 반반차는 반차의
    부분 문자열이므로 반반차 먼저 검사한다.
    """
    text = f"{day_type or ''} {note or ''} {attendance_code or ''}"
    for keyword, hours in PARTIAL_LEAVE_SHIFT_HOURS:
        if keyword in text:
            if "오전" in text:
                half = "오전"
            elif "오후" in text:
                half = "오후"
            else:
                half = "unknown"
            return keyword, half, hours * 60
    return None


def _compute_anomaly_baseline(
    shift_time: str, day_type: str, note: str, attendance_code: str = ""
) -> tuple[int, int] | None:
    """shift_time과 부분 휴가를 반영한 ``(출근분, 퇴근분)`` baseline을 반환.

    shift_time이 알려진 근무형태가 아니면 ``None`` (감지 제외).
    부분 휴가(반차/반반차)는 주간 시프트에만 적용한다. 2교대는 근무
    구조가 달라 동일 규칙이 맞지 않으므로 일단 기본 baseline 그대로 쓴다.
    """
    baseline = SHIFT_BASELINES.get(shift_time)
    if baseline is None:
        return None
    base_in, base_out = baseline

    if shift_time == "주간":
        if day_type == DAY_SHIFT_SHORT_LUNCH_DAY_TYPE:
            base_out -= DAY_SHIFT_SHORT_LUNCH_CHECKOUT_OFFSET_MINUTES
        leave = _partial_leave_shift(day_type, note, attendance_code)
        if leave:
            _kind, half, shift_minutes = leave
            if half == "오전":
                base_in += shift_minutes
            elif half == "오후":
                base_out -= shift_minutes
            # half == "unknown" → 구분 모호. 보수적으로 기본 baseline 유지.
    return base_in, base_out


def _partial_leave_shift_minutes(
    shift_time: str,
    leave_kind: str,
    default_minutes: int,
) -> int:
    hours_by_kind = PARTIAL_LEAVE_SHIFT_HOURS_BY_SHIFT.get(shift_time)
    if not hours_by_kind:
        return default_minutes
    return int(hours_by_kind.get(leave_kind, default_minutes // 60)) * 60


def _infer_partial_leave_half(
    row: AttendanceRow,
    base_in: int,
    base_out: int,
    shift_minutes: int,
) -> str:
    in_mins = _hhmm_to_minutes(row.check_in)
    out_mins = _hhmm_to_minutes(row.check_out)
    morning_in = base_in + shift_minutes
    afternoon_out = base_out - shift_minutes

    if in_mins is not None and in_mins >= morning_in:
        return "오전"
    if out_mins is not None and out_mins <= afternoon_out + ALERT_GRACE_MINUTES:
        return "오후"
    return "unknown"


def _compute_row_anomaly_baseline(
    row: AttendanceRow,
    shift_time: str,
) -> tuple[int, int] | None:
    baseline = SHIFT_BASELINES.get(shift_time)
    if baseline is None:
        return None
    base_in, base_out = baseline

    if shift_time in PARTIAL_LEAVE_SHIFT_HOURS_BY_SHIFT:
        if row.day_type == DAY_SHIFT_SHORT_LUNCH_DAY_TYPE:
            base_out -= DAY_SHIFT_SHORT_LUNCH_CHECKOUT_OFFSET_MINUTES
        leave = _partial_leave_shift(row.day_type, row.note, row.attendance_code)
        if leave:
            leave_kind, half, default_shift_minutes = leave
            shift_minutes = _partial_leave_shift_minutes(
                shift_time, leave_kind, default_shift_minutes
            )
            if half == "unknown":
                half = _infer_partial_leave_half(row, base_in, base_out, shift_minutes)
            if half == "오전":
                base_in += shift_minutes
            elif half == "오후":
                base_out -= shift_minutes
    return base_in, base_out


def _row_is_future(row: AttendanceRow, reference: _dt.datetime) -> bool:
    try:
        row_date = _dt.date.fromisoformat(row.date)
    except ValueError:
        return False
    return row_date > reference.date()


def _baseline_has_passed(row: AttendanceRow, baseline_minutes: int, reference: _dt.datetime) -> bool:
    try:
        row_date = _dt.date.fromisoformat(row.date)
    except ValueError:
        return True
    if row_date < reference.date():
        return True
    if row_date > reference.date():
        return False
    reference_minutes = reference.hour * 60 + reference.minute
    return reference_minutes >= baseline_minutes + ALERT_GRACE_MINUTES


def _append_issue(issues: list[str], issue: str) -> None:
    if issue not in issues:
        issues.append(issue)


def _attendance_code_issues(row: AttendanceRow) -> list[str]:
    code = str(row.attendance_code or "").strip()
    if not code:
        return []
    if not any(keyword in code for keyword in ATTENDANCE_ISSUE_CODE_KEYWORDS):
        return []

    issues: list[str] = []
    if "출" in code:
        _append_issue(issues, "출근 누락")
    if "퇴" in code:
        _append_issue(issues, "퇴근 누락")
    if not issues:
        _append_issue(issues, f"근태코드 확인: {code}")
    return issues


def _deduction_code_mismatch_issues(row: AttendanceRow) -> list[str]:
    code = str(row.attendance_code or "")
    has_deduction = (
        row.late_hours > 0
        or row.early_leave_hours > 0
        or row.outing_hours > 0
    )
    if not has_deduction:
        return []

    if any(keyword in code for keyword in ANNUAL_LEAVE_KEYWORDS):
        return ["공제시간 불일치"]

    issues: list[str] = []
    if row.late_hours > 0 and "지각" not in code:
        _append_issue(issues, "근태코드 누락(지각)")
    if row.early_leave_hours > 0 and "조퇴" not in code:
        _append_issue(issues, "근태코드 누락(조퇴)")
    if row.outing_hours > 0 and "외출" not in code:
        _append_issue(issues, "근태코드 누락(외출)")
    return issues


def _unprocessed_row_issues(
    row: AttendanceRow,
    shift_time: str,
    *,
    reference: _dt.datetime | None = None,
) -> list[str]:
    reference = reference or _alert_reference_datetime()
    if _row_is_future(row, reference):
        return []
    if row.day_type not in ALERT_WORKDAY_TYPES:
        return []

    issues = _attendance_code_issues(row)
    for issue in _deduction_code_mismatch_issues(row):
        _append_issue(issues, issue)

    if _is_full_day_leave(row.day_type, row.note, row.attendance_code):
        return issues

    baseline = _compute_row_anomaly_baseline(row, shift_time)
    if baseline is None:
        return issues
    base_in, base_out = baseline

    if row.late_hours <= 0 and not row.check_in and _baseline_has_passed(row, base_in, reference):
        _append_issue(issues, "출근 누락")
    elif row.late_hours <= 0:
        in_mins = _hhmm_to_minutes(row.check_in)
        if in_mins is not None and in_mins > base_in:
            _append_issue(issues, "지각 미처리")

    if row.early_leave_hours <= 0 and not row.check_out and _baseline_has_passed(row, base_out, reference):
        _append_issue(issues, "퇴근 누락")
    elif row.early_leave_hours <= 0:
        out_mins = _hhmm_to_minutes(row.check_out)
        if out_mins is not None and out_mins < base_out:
            _append_issue(issues, "조퇴 미처리")

    return issues


def _row_issue_labels(row: AttendanceRow, shift_time: str) -> list[str]:
    return _unprocessed_row_issues(row, shift_time)


def _display_date(date_text: str) -> str:
    try:
        return _dt.date.fromisoformat(date_text).strftime("%m-%d")
    except ValueError:
        return date_text


def _row_alert_category(row: AttendanceRow, issues: list[str]) -> tuple[str, str]:
    joined = " / ".join(issues)
    has_check_in_missing = "\uCD9C\uADFC \uB204\uB77D" in issues
    has_check_out_missing = "\uD1F4\uADFC \uB204\uB77D" in issues

    if has_check_in_missing and has_check_out_missing:
        return "1", "\uCD9C/\uD1F4\uADFC \uBBF8\uD0C0\uAC01"
    if has_check_in_missing:
        return "2", "\uCD9C\uADFC \uBBF8\uD0C0\uAC01"
    if has_check_out_missing:
        return "3", "\uD1F4\uADFC \uBBF8\uD0C0\uAC01"

    has_code_deduction_mismatch = any(
        issue.startswith("\uADFC\uD0DC\uCF54\uB4DC \uB204\uB77D")
        or issue == "\uACF5\uC81C\uC2DC\uAC04 \uBD88\uC77C\uCE58"
        for issue in issues
    )
    if has_code_deduction_mismatch:
        return "0", "\uADFC\uD0DC \uC774\uC0C1"

    if any(issue.startswith("\uC9C0\uAC01 ") for issue in issues):
        return "4", "\uADFC\uD0DC \uC774\uC0C1"
    if any(issue.startswith("\uC870\uD1F4 ") for issue in issues):
        return "4", "\uADFC\uD0DC \uC774\uC0C1"

    has_unprocessed_late_or_leave = any(
        "\uBBF8\uCC98\uB9AC" in issue for issue in issues
    )
    if has_unprocessed_late_or_leave:
        basis = f"{row.day_type} {row.note} {row.attendance_code}"
        if "\uBC18\uBC18\uCC28" in basis:
            return "6", "\uADFC\uD0DC\uCF54\uB4DC \uB204\uB77D(\uBC18\uBC18\uCC28/\uC870\uD1F4 \uC608\uC0C1)"
        if "\uBC18\uCC28" in basis or "\uC870\uD1F4" in joined:
            return "5", "\uADFC\uD0DC\uCF54\uB4DC \uB204\uB77D(\uBC18\uCC28/\uC870\uD1F4 \uC608\uC0C1)"
        return "0", "\uAE30\uD0C0"

    return "0", "\uAE30\uD0C0"


def _hide_detail_issue_text(content: str, issues: list[str]) -> bool:
    return content == "\uADFC\uD0DC \uC774\uC0C1"


def _anomaly_detail(row: AttendanceRow, issues: list[str]) -> dict[str, Any]:
    code, content = _row_alert_category(row, issues)
    extra_content = " / ".join(issues)
    if extra_content == content or _hide_detail_issue_text(content, issues):
        extra_content = ""
    return {
        "date": row.date,
        "display_date": _display_date(row.date),
        "code": code,
        "content": content,
        "extra_content": extra_content,
        "status": "",
        "issues": list(issues),
    }


def _merge_anomaly_record(
    anomalies_by_emp: dict[str, dict[str, Any]],
    rec: dict[str, Any],
    issues: list[str],
    *,
    include_dates: bool = False,
) -> None:
    row: AttendanceRow = rec["row"]
    detail = _anomaly_detail(row, issues) if include_dates else None
    existing = anomalies_by_emp.get(rec["emp_id"])
    if existing is None:
        item = {
            "emp_id": rec["emp_id"],
            "name": rec["name"],
            "department": rec["department"],
            "shift_time": rec.get("shift_time", "") or "",
            "issues": list(issues),
        }
        if include_dates:
            item["dates"] = [row.date]
            item["details"] = [detail]
        anomalies_by_emp[rec["emp_id"]] = item
        return

    for issue in issues:
        if issue not in existing["issues"]:
            existing["issues"].append(issue)

    if include_dates:
        dates = existing.setdefault("dates", [])
        if row.date not in dates:
            dates.append(row.date)
        details = existing.setdefault("details", [])
        if detail is not None:
            detail_key = (detail["date"], detail["code"], detail["content"], detail["extra_content"])
            existing_keys = {
                (
                    existing_detail.get("date"),
                    existing_detail.get("code"),
                    existing_detail.get("content"),
                    existing_detail.get("extra_content"),
                )
                for existing_detail in details
            }
            if detail_key not in existing_keys:
                details.append(detail)


def detect_today_anomalies(
    year_month: str, target_date: str
) -> tuple[str, list[dict[str, Any]]]:
    """``target_date``의 근태 이상을 감지해 반환.

    알림은 이미 ERP에 산출된 지각/조퇴 시간과, 아직 코드/시간 값으로
    정리되지 않았지만 기준 시각을 벗어난 미처리 이상을 함께 보여준다.
    연차/휴가/훈련처럼 하루 전체 부재로 승인된 행은 제외한다.

    감지 규칙:
      - day_type == '평일' 또는 '평일2'만 대상. '평일2'는 주간 기준
        점심시간 30분 단축일로 17:30 퇴근을 정상 기준으로 본다.
      - 연차/월차/휴가/유급/공가 키워드가 day_type 또는 비고에 있으면
        전체 감지 대상에서 제외. 반차/반반차는 제외가 아니라 baseline
        만 이동(주간 한정)시켜 감지를 계속 수행한다.
      - shift_time이 ``SHIFT_BASELINES``에 없으면 감지 제외
        (빈 값 = 교대조 비번 날; 알 수 없는 근무형태).

    이상 항목:
      - 공제시간의 지각/조퇴/외출 값은 같은 근태코드가 있을 때만 정상
      - 반차/반반차/연차 등 휴가 코드에 지각/조퇴/외출 공제시간이 있으면
        "공제시간 불일치"
      - 기준 출근+grace 이후 check_in 없음 → "출근 누락"
      - 기준 퇴근+grace 이후 check_out 없음 → "퇴근 누락"
      - check_in 시각 > 기준 출근  AND late_hours == 0        → "지각 미처리"
      - check_out 시각 < 기준 퇴근 AND early_leave_hours == 0 → "조퇴 미처리"

    Raises ``MonthFileNotFound`` or ``FileLocked`` on I/O issues.
    """

    day_type = ""
    anomalies_by_emp: dict[str, dict[str, Any]] = {}
    reference = _alert_reference_datetime()
    for path in _month_file_paths_or_raise(year_month):
        for rec in _records_from_path(path):
            row: AttendanceRow = rec["row"]
            if row.date != target_date:
                continue
            if not day_type:
                day_type = row.day_type
            if row.day_type not in ALERT_WORKDAY_TYPES:
                continue
            if _is_full_day_leave(row.day_type, row.note, row.attendance_code):
                continue

            shift_time = rec.get("shift_time", "") or ""
            issues = _unprocessed_row_issues(row, shift_time, reference=reference)
            if not issues:
                continue

            _merge_anomaly_record(anomalies_by_emp, rec, issues)

    return day_type, list(anomalies_by_emp.values())


def detect_month_anomalies(year_month: str) -> list[dict[str, Any]]:
    """Return attendance anomalies for the selected month up to the current time."""

    anomalies_by_emp: dict[str, dict[str, Any]] = {}
    reference = _alert_reference_datetime()
    for path in _month_file_paths_or_raise(year_month):
        for rec in _records_from_path(path):
            row: AttendanceRow = rec["row"]
            shift_time = rec.get("shift_time", "") or ""
            issues = _unprocessed_row_issues(row, shift_time, reference=reference)
            if not issues:
                continue
            _merge_anomaly_record(anomalies_by_emp, rec, issues, include_dates=True)

    items = list(anomalies_by_emp.values())
    for item in items:
        if "dates" in item:
            item["dates"] = sorted(item["dates"])
        if "details" in item:
            item["details"] = sorted(
                item["details"],
                key=lambda detail: (
                    str(detail.get("date", "")),
                    str(detail.get("code", "")),
                    str(detail.get("content", "")),
                ),
            )
    return items


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M")
    return str(value).strip()


def _cell_at(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _cell_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _cell_time(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, _dt.datetime):
        return value.strftime("%H:%M")
    text = str(value).strip()
    return text or None


def _make_column_map(
    header_group: tuple[Any, ...] | None,
    header_sub: tuple[Any, ...] | None,
) -> dict[str, int]:
    """헤더 두 행(그룹 행0 + 세부 행1)에서 논리 필드 → 열 인덱스 맵을 만든다.

    헤더에서 필수 필드를 모두 찾지 못하면 ``DEFAULT_COLUMNS`` 를 그대로
    돌려준다(구버전 레이아웃·헤더 없는 입력에 대한 안전한 폴백).
    """
    if not header_sub:
        return dict(DEFAULT_COLUMNS)

    group = list(header_group or ())
    sub = list(header_sub)

    # 병합된 그룹 헤더(평일근무시간/휴일근무시간 등)를 오른쪽으로 채운다.
    filled_group: list[str] = []
    last = ""
    for cell in group:
        text = _cell_str(cell)
        if text:
            last = text
        filled_group.append(last)

    detected: dict[str, int] = {}
    width = max(len(sub), len(filled_group))
    for idx in range(width):
        sub_name = _cell_str(sub[idx]) if idx < len(sub) else ""
        grp_name = filled_group[idx] if idx < len(filled_group) else ""

        field = _HEADER_SIMPLE_FIELDS.get(sub_name)
        if field and field not in detected:
            detected[field] = idx
            continue

        if sub_name in _HEADER_WORKHOUR_SUFFIX:
            prefix = "weekday" if "평일" in grp_name else "holiday" if "휴일" in grp_name else ""
            if prefix:
                key = f"{prefix}_{_HEADER_WORKHOUR_SUFFIX[sub_name]}"
                detected.setdefault(key, idx)
            continue

        # 비고는 세부 행이 비어 있고 그룹 행에만 '비고'로 존재한다.
        if not sub_name and "비고" in grp_name and "note" not in detected:
            detected["note"] = idx

    if not _HEADER_REQUIRED_FIELDS.issubset(detected):
        return dict(DEFAULT_COLUMNS)
    return {**DEFAULT_COLUMNS, **detected}


def _cell_day_type(row: tuple[Any, ...], colmap: dict[str, int]) -> str:
    day_type_idx = colmap["day_type"]
    primary = _cell_str(_cell_at(row, day_type_idx))
    if primary in DAY_TYPE_VALUES:
        return primary
    shifted = _cell_str(_cell_at(row, day_type_idx + 1))
    if shifted in DAY_TYPE_VALUES:
        return shifted
    return primary or shifted


def _load_workbook(path: Path):
    if not path.exists():
        raise MonthFileNotFound(str(path))
    if openpyxl is None:
        raise RuntimeError(
            "openpyxl is required to load attendance Excel files. "
            "Install test dependencies with `pip install -r requirements-dev.txt`."
        ) from _OPENPYXL_IMPORT_ERROR
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError as exc:  # pragma: no cover - windows-specific
        raise FileLocked(str(path)) from exc
    except InvalidFileException as exc:
        raise FileFormatInvalid(str(path)) from exc


def _iter_data_rows(ws, colmap: dict[str, int] | None = None):
    date_idx = (colmap or DEFAULT_COLUMNS)["date"]
    emp_idx = (colmap or DEFAULT_COLUMNS)["emp_id"]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx < 2:
            continue
        if not row:
            continue
        if _cell_at(row, date_idx) is None and _cell_at(row, emp_idx) is None:
            continue
        yield row


def _row_to_record(
    row: tuple[Any, ...], colmap: dict[str, int] | None = None
) -> dict[str, Any]:
    cols = colmap or DEFAULT_COLUMNS
    emp_id = _cell_str(_cell_at(row, cols["emp_id"]))
    if not emp_id:
        return {}
    shift_time = _cell_str(_cell_at(row, cols["shift_time"]))
    row_data = AttendanceRow(
        date=_cell_str(_cell_at(row, cols["date"])),
        weekday=_cell_str(_cell_at(row, cols["weekday"])),
        day_type=_cell_day_type(row, cols),
        check_in=_cell_time(_cell_at(row, cols["check_in"])),
        check_out=_cell_time(_cell_at(row, cols["check_out"])),
        next_day=bool(_cell_float(_cell_at(row, cols["next_day"]))),
        weekday_early=_cell_float(_cell_at(row, cols["weekday_early"])),
        weekday_normal=_cell_float(_cell_at(row, cols["weekday_normal"])),
        weekday_overtime=_cell_float(_cell_at(row, cols["weekday_overtime"])),
        weekday_night=_cell_float(_cell_at(row, cols["weekday_night"])),
        holiday_early=_cell_float(_cell_at(row, cols["holiday_early"])),
        holiday_normal=_cell_float(_cell_at(row, cols["holiday_normal"])),
        holiday_overtime=_cell_float(_cell_at(row, cols["holiday_overtime"])),
        holiday_night=_cell_float(_cell_at(row, cols["holiday_night"])),
        late_hours=_cell_float(_cell_at(row, cols["late"])),
        early_leave_hours=_cell_float(_cell_at(row, cols["early_leave"])),
        outing_hours=_cell_float(_cell_at(row, cols["outing"])),
        note=_cell_str(_cell_at(row, cols["note"])),
        attendance_code=_cell_str(_cell_at(row, cols["attendance_code"])),
    )
    row_data.issues = _row_issue_labels(row_data, shift_time)
    row_data.has_issue = bool(row_data.issues)
    return {
        "emp_id": emp_id,
        "name": _cell_str(_cell_at(row, cols["name"])),
        "department": _cell_str(_cell_at(row, cols["department"])),
        "factory": _cell_str(_cell_at(row, cols["factory"])),
        "shift_time": shift_time,
        "shift_group": _cell_str(_cell_at(row, cols["shift_group"])),
        "job_type": _cell_str(_cell_at(row, cols["job_type"])),
        "gender": _cell_str(_cell_at(row, cols["gender"])),
        "row": row_data,
    }


def _column_map_from_ws(ws) -> dict[str, int]:
    """워크시트의 헤더 두 행으로 열 맵을 만든다.

    헤더를 읽을 수 없으면(예: 헤더가 없는 입력) ``DEFAULT_COLUMNS`` 폴백.
    """
    try:
        header_rows = list(ws.iter_rows(values_only=True, max_row=2))
    except (AttributeError, TypeError):
        return dict(DEFAULT_COLUMNS)
    if len(header_rows) < 2:
        return dict(DEFAULT_COLUMNS)
    return _make_column_map(header_rows[0], header_rows[1])


def _records_from_path(path: Path) -> list[dict[str, Any]]:
    wb = _load_workbook(path)
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        colmap = _column_map_from_ws(ws)
        records: list[dict[str, Any]] = []
        for raw in _iter_data_rows(ws, colmap):
            rec = _row_to_record(raw, colmap)
            if rec:
                records.append(rec)
        return records
    finally:
        wb.close()


def _attendance_row_key(row: AttendanceRow) -> tuple[Any, ...]:
    data = asdict(row)
    normalized: list[tuple[str, Any]] = []
    for key, value in data.items():
        if isinstance(value, list):
            normalized.append((key, tuple(value)))
        else:
            normalized.append((key, value))
    return tuple(normalized)


def _record_to_profile(rec: dict[str, Any]) -> AttendanceProfile:
    return AttendanceProfile(
        emp_id=rec["emp_id"],
        name=rec["name"],
        department=rec["department"],
        factory=rec["factory"],
        shift_time=rec["shift_time"],
        shift_group=rec["shift_group"],
        job_type=rec["job_type"],
        gender=rec["gender"],
    )


def _annual_leave_text(row: AttendanceRow) -> str:
    return f"{row.day_type} {row.attendance_code} {row.note}".strip()


def _is_annual_leave_row(row: AttendanceRow) -> bool:
    text = _annual_leave_text(row)
    return any(keyword in text for keyword in ANNUAL_LEAVE_KEYWORDS)


def _annual_leave_bucket(row: AttendanceRow) -> tuple[str, float]:
    text = _annual_leave_text(row)
    if "반반차" in text:
        return "quarter", 0.25
    if any(keyword in text for keyword in HALF_DAY_LEAVE_KEYWORDS):
        return "half", 0.5
    return "full", 1.0


def _annual_leave_days(row: AttendanceRow) -> float:
    _bucket, days = _annual_leave_bucket(row)
    return days


def employee_list(year_month: str) -> list[dict[str, str]]:
    """Return distinct employees present in the given month's file."""
    seen: dict[str, dict[str, str]] = {}
    for path in _month_file_paths_or_raise(year_month):
        for rec in _records_from_path(path):
            if rec["emp_id"] not in seen:
                seen[rec["emp_id"]] = {
                    "emp_id": rec["emp_id"],
                    "name": rec["name"],
                    "department": rec["department"],
                    "factory": rec["factory"],
                }
    return sorted(seen.values(), key=lambda x: (x["department"], x["name"], x["emp_id"]))


def employee_exists_in_any_month(emp_id: str) -> bool:
    """Used during first login: accept any month that has this sa-beon."""
    emp_id = (emp_id or "").strip()
    if not emp_id:
        return False
    for ym in available_months():
        for path in month_file_paths(ym):
            try:
                records = _records_from_path(path)
            except (MonthFileNotFound, FileLocked, FileFormatInvalid):
                continue
            for rec in records:
                if rec["emp_id"] == emp_id:
                    return True
    return False


def employee_profile_from_any_month(emp_id: str) -> AttendanceProfile | None:
    """Find employee identity even when the selected month has no rows for them."""
    emp_id = (emp_id or "").strip()
    if not emp_id:
        return None
    for ym in available_months():
        for path in month_file_paths(ym):
            try:
                records = _records_from_path(path)
            except (MonthFileNotFound, FileLocked, FileFormatInvalid):
                continue
            for rec in records:
                if rec and rec["emp_id"] == emp_id:
                    return _record_to_profile(rec)
    return None


def _load_month_rows_for_employee(
    year_month: str, emp_id: str
) -> tuple[AttendanceProfile | None, list[AttendanceRow]]:
    """Return the rows present in one month without searching other months."""
    profile: AttendanceProfile | None = None
    rows: list[AttendanceRow] = []
    seen_rows: set[tuple[Any, ...]] = set()
    for path in _month_file_paths_or_raise(year_month):
        for rec in _records_from_path(path):
            if rec["emp_id"] != emp_id:
                continue
            if profile is None:
                profile = _record_to_profile(rec)
            row = rec["row"]
            row_key = _attendance_row_key(row)
            if row_key in seen_rows:
                continue
            seen_rows.add(row_key)
            rows.append(row)
    rows.sort(key=lambda row: (row.date, row.check_in or "", row.check_out or "", row.note))
    return profile, rows


def load_month_for_employee(
    year_month: str, emp_id: str
) -> tuple[AttendanceProfile | None, list[AttendanceRow], AttendanceSummary]:
    """Return (profile, rows, summary) for one employee in one month."""
    profile, rows = _load_month_rows_for_employee(year_month, emp_id)
    if profile is None:
        profile = employee_profile_from_any_month(emp_id)
    summary = _summarize(rows)
    return profile, rows, summary


def load_year_summary_for_employee(year: int, emp_id: str) -> AttendanceAnnualSummary:
    """Aggregate yearly late/annual-leave totals from available monthly files."""
    summary = AttendanceAnnualSummary(year=year)
    prefix = f"{year:04d}-"
    year_months = [ym for ym in available_months() if ym.startswith(prefix)]
    summary.available_months_count = len(year_months)
    for ym in year_months:
        try:
            _profile, rows = _load_month_rows_for_employee(ym, emp_id)
        except (MonthFileNotFound, FileLocked, FileFormatInvalid):
            summary.skipped_months.append(ym)
            continue
        if not rows:
            continue
        month_summary = _summarize(rows)
        summary.months_count += 1
        summary.late_count += month_summary.late_count
        summary.late_total += month_summary.late_total
        for row in rows:
            if _is_annual_leave_row(row):
                bucket, days = _annual_leave_bucket(row)
                summary.annual_leave_count += 1
                summary.annual_leave_days += days
                if bucket == "quarter":
                    summary.annual_leave_quarter_days += days
                elif bucket == "half":
                    summary.annual_leave_half_days += days
                else:
                    summary.annual_leave_full_days += days
    return summary


def _summarize(rows: list[AttendanceRow]) -> AttendanceSummary:
    s = AttendanceSummary()
    for r in rows:
        if r.check_in or r.check_out:
            s.work_days += 1
        if r.late_hours > 0:
            s.late_count += 1
            s.late_total += r.late_hours
        if r.early_leave_hours > 0:
            s.early_leave_count += 1
            s.early_leave_total += r.early_leave_hours
        if r.outing_hours > 0:
            s.outing_count += 1
            s.outing_total += r.outing_hours
        s.weekday_early += r.weekday_early
        s.weekday_normal += r.weekday_normal
        s.weekday_overtime += r.weekday_overtime
        s.weekday_night += r.weekday_night
        s.holiday_early += r.holiday_early
        s.holiday_normal += r.holiday_normal
        s.holiday_overtime += r.holiday_overtime
        s.holiday_night += r.holiday_night
    return s


def serialize_profile(profile: AttendanceProfile | None) -> dict[str, str] | None:
    return asdict(profile) if profile else None


def serialize_rows(rows: list[AttendanceRow]) -> list[dict[str, Any]]:
    return [asdict(r) for r in rows]


def serialize_summary(summary: AttendanceSummary) -> dict[str, Any]:
    return asdict(summary)


def serialize_annual_summary(summary: AttendanceAnnualSummary) -> dict[str, Any]:
    return asdict(summary)
