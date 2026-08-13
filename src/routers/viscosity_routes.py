"""합성 점도 등록·추세·이상 분석 라우트.

접근(정책 ⓑ, 2026-07-22):
  - op_router(열람·측정 등록): 로그인 없이 누구나 — 사내 공용 단말 현장 등록 편의.
    등록 시 로그인 사용자가 있으면 그 이름으로, 없으면 '현장' 으로 created_by/audit 기록.
  - mgr_router(제품 생성/수정, 측정 삭제, Excel export): 관리(management) 성격의 쓰기라
    책임자 권한을 서버에서 강제한다. 의존성은 api.py 에서 mgr_router include 시
    require_access_level("manager") 로 건다(라우터 단위 wiring). 화면은 설정/삭제 버튼을
    can_manage 로 숨기므로, 관리 세션에서는 이 강제가 투명하다.

Plan:   docs/01-plan/features/viscosity-analysis.plan.md
Design: docs/02-design/features/viscosity-analysis.design.md

Endpoints:
    GET    /viscosity/overview                  (개방)
    GET    /viscosity/products                  (개방)
    GET    /viscosity/products/{id}             분석 포함 (개방)
    POST   /viscosity/readings                  (개방 — 현장 등록)
    POST   /viscosity/products                  (책임자)
    PATCH  /viscosity/products/{id}             (책임자)
    DELETE /viscosity/readings/{id}             (책임자)
    POST   /viscosity/readings/{id}/exclude     통계 제외 (책임자)
    POST   /viscosity/readings/{id}/include     제외 해제 (책임자)
    GET    /viscosity/products/{id}/export      Excel (책임자)
"""

import io
import sqlite3
from datetime import date
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

# 판정 라벨(화면 STATUS_LABEL 과 동일) — Excel '측정 원본' 판정 열에 쓴다.
_STATUS_LABEL = {"normal": "정상", "warn": "경고", "anomaly": "이상"}
_ALLOWED_GRANULARITY = ("day", "week", "month", "quarter", "year")


def _parse_reactor(raw):
    """reactor 쿼리 파라미터 해석 — "1"~"4" 는 정수, "none" 은 미지정(IS NULL) 뷰,
    그 외/빈 값은 필터 없음. 미지정 항목은 반응기 도입 전 과거 데이터 분리 목적(2026-07-23)."""
    if raw is None or raw == "":
        return None
    if raw == "none":
        return "none"
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return None
    return value if value in (1, 2, 3, 4) else None

_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _xlsx_safe(value: Any) -> Any:
    # 수식 인젝션 방지 — CSV 와 동일하게 =,+,-,@,\t,\r 로 시작하는 문자열은 앞에 ' 를 붙여
    # Excel 이 수식으로 해석하지 못하게 한다. 숫자/None 은 그대로 둔다.
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value

from ..auth import get_current_user
from ..db import get_db, local_today_text, utc_now_text, write_audit_log
from ..services import viscosity_service
from .models import (
    ViscosityExcludeBody,
    ViscosityProductCreateBody,
    ViscosityProductUpdateBody,
    ViscosityReadingBody,
    actor_name,
)


def _require_product(connection: sqlite3.Connection, product_id: int) -> dict[str, Any]:
    product = viscosity_service.get_product(connection, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="점도 제품을 찾을 수 없습니다.")
    return product


def build_router() -> tuple[APIRouter, APIRouter]:
    # op_router: 열람·측정 등록은 로그인 없이 누구나(현장 편의). 등록 시 로그인 사용자가
    # 있으면 그 이름으로, 없으면 '현장' 으로 기록. mgr_router: 제품 설정·측정 삭제·export 는
    # 책임자 전용(Excel export 포함) — api.py 에서 require_access_level("manager") 의존성을 라우터에 건다(정책 ⓑ).
    op_router = APIRouter()
    mgr_router = APIRouter()

    # ---- 조회 + 등록 -----------------------------------------------------
    @op_router.get("/viscosity/overview")
    def viscosity_overview(
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        return viscosity_service.overview(connection)

    @op_router.get("/viscosity/products")
    def viscosity_products(
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        return {"items": viscosity_service.list_products(connection)}

    @op_router.get("/viscosity/products/{product_id}")
    def viscosity_product_detail(
        product_id: int,
        granularity: str = "quarter",
        year: int | None = None,
        reactor: str | None = None,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        product = _require_product(connection, product_id)
        if granularity not in ("day", "week", "month", "quarter", "year"):
            granularity = "quarter"
        reactor = _parse_reactor(reactor)
        return viscosity_service.analyze_product(
            connection, product, granularity=granularity, year=year, reactor=reactor
        )

    @op_router.get("/viscosity/products/{product_id}/blend-records")
    def viscosity_blend_records(
        product_id: int,
        unregistered: str | None = None,
        q: str | None = None,
        reactor: str | None = None,
        limit: int = 20,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """등록 패널용 배합 기록 목록 — 한 번의 요청으로 등록 여부까지 함께.

        종전에는 화면이 /blend/records 검색 후 20건을 낱개 상세 조회(총 22회)했고,
        '미등록만' 필터는 잘라온 20건에만 적용돼 더 오래된 미등록 LOT 이 있어도
        빈 목록이 나왔다(2026-08-13 검토 4·10번). 여기서는 서버가 점도 등록 여부를
        LEFT JOIN 으로 함께 계산해 필터·집계까지 끝낸다.
        """
        product = _require_product(connection, product_id)
        limit = max(1, min(int(limit or 20), 200))
        names = {product["name"], product["code"]}
        placeholders = ",".join("?" for _n in names)
        params: list[Any] = list(names)
        where = [
            f"br.product_name IN ({placeholders})",
            "br.status = 'completed'",
            "COALESCE(br.is_bulk_regenerated, 0) = 0",
        ]
        query = (q or "").strip()
        if query:
            like = f"%{query}%"
            where.append("(br.product_lot LIKE ? OR br.worker LIKE ?)")
            params.extend([like, like])
        # 반응기 필터 — 화면 select 와 동일 규약(_parse_reactor). 클라이언트에서
        # 걸면 받아온 limit 건 안에서만 걸러져 반응기별 옛 기록을 놓친다.
        reactor_value = _parse_reactor(reactor)
        if reactor_value == "none":
            where.append("br.reactor IS NULL")
        elif reactor_value is not None:
            where.append("br.reactor = ?")
            params.append(reactor_value)
        # 등록 여부는 EXISTS 로 — 한 기록에 측정이 둘 이상 연결돼도 행이 불지 않는다.
        registered_sql = (
            "EXISTS (SELECT 1 FROM viscosity_readings vr "
            "        WHERE vr.blend_record_id = br.id)"
        )
        base_sql = f"FROM blend_records br WHERE {' AND '.join(where)}"
        total = connection.execute(
            f"SELECT COUNT(*) {base_sql}", params
        ).fetchone()[0]
        unregistered_total = connection.execute(
            f"SELECT COUNT(*) {base_sql} AND NOT {registered_sql}", params
        ).fetchone()[0]
        only_unregistered = unregistered == "1"
        rows = connection.execute(
            f"""
            SELECT br.id, br.product_lot, br.work_date, br.worker,
                   br.total_amount, br.reactor,
                   {registered_sql} AS registered,
                   (SELECT vr2.viscosity FROM viscosity_readings vr2
                     WHERE vr2.blend_record_id = br.id
                     ORDER BY vr2.id DESC LIMIT 1) AS reading_value
            {base_sql}
            {f"AND NOT {registered_sql}" if only_unregistered else ""}
            ORDER BY br.work_date DESC, br.id DESC
            LIMIT ?
            """,
            [*params, limit],
        ).fetchall()
        return {
            "items": [
                {
                    "id": int(r["id"]),
                    "product_lot": r["product_lot"],
                    "work_date": r["work_date"],
                    "worker": r["worker"],
                    "total_amount": r["total_amount"],
                    "reactor": r["reactor"],
                    "registered": bool(r["registered"]),
                    "viscosity": (
                        float(r["reading_value"])
                        if r["reading_value"] is not None
                        else None
                    ),
                }
                for r in rows
            ],
            "total": int(total),
            "unregistered_total": int(unregistered_total),
            "limit": limit,
        }

    @op_router.get("/viscosity/blend-records/{record_id}/used-pb")
    def viscosity_used_pb_preview(
        record_id: int,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """선택한 배합 기록의 '사용한 PB' 감지 미리보기 — 등록 전에 확인·수정 기회.

        종전에는 서버가 상세 첫 행을 몰래 PB 로 저장했고 화면에 보이지 않았다
        (2026-08-13 검토 2번). method 가 'first_row' 면 화면이 "자동 감지 실패 -
        확인 필요" 를 표시한다. pb_viscosity 는 그 PB LOT 의 최신 점도(없으면 null).
        """
        from ..services import blend_service

        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        lot, method = viscosity_service.detect_source_pb_lot(
            record.get("details") or []
        )
        pb_map = viscosity_service._pb_viscosity_map(connection)
        pb_value = pb_map.get(viscosity_service._lot_digits(lot)) if lot else None
        return {"lot": lot, "method": method, "pb_viscosity": pb_value}

    @op_router.post("/viscosity/readings")
    def viscosity_add_reading(
        body: ViscosityReadingBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        current_user = get_current_user(request, required=False)
        product = _require_product(connection, body.product_id)
        # 반응기는 배합 실적에서 지정하고 점도는 실적에서 물려받는다. 이 직접 등록
        # 경로(수동/임포트)는 reactor 를 선택적으로만 받는다.
        # 입력 전 '같은 연도' 표본 기준으로 판정 (연도별 기준 + 자기 자신 평균 오염 방지)
        resolved_date = (
            body.measured_date
            or viscosity_service.parse_lot_date(body.lot_no)
            # 측정일 기본값은 현장 기준 '오늘'(로컬) — UTC 로 자르면 자정 전
            # 등록이 어제로 밀려 '오늘 미입력' 리마인더와 어긋난다.
            or local_today_text()
        )
        reading_year = int(resolved_date[:4]) if resolved_date[:4].isdigit() else None
        verdict = viscosity_service.classify_value(
            connection, product, body.viscosity, year=reading_year, reactor=body.reactor
        )
        try:
            reading_id = viscosity_service.add_reading(
                connection,
                product_id=product["id"],
                lot_no=body.lot_no,
                viscosity=body.viscosity,
                # 판정에 쓴 날짜를 그대로 저장한다 — 원본(body.measured_date)을 넘기면
                # 서비스가 같은 폴백을 **다시** 돌려, 자정 경계에서 "A년 표본으로 판정 →
                # B년으로 저장"이 될 수 있다(감사 F-9). 폴백은 한 번만.
                measured_date=resolved_date,
                memo=body.memo,
                recipe_material=body.recipe_material,
                material_lot=body.material_lot,
                created_by=actor_name(current_user) if current_user else "현장",
                created_at=utc_now_text(),
                reactor=body.reactor,
            )
        except sqlite3.IntegrityError:
            raise HTTPException(
                status_code=409,
                detail=f"이미 등록된 LOT 입니다: {body.lot_no}",
            )
        write_audit_log(
            connection,
            action="viscosity_reading_add",
            actor=current_user,
            target_type="viscosity_reading",
            target_id=str(reading_id),
            target_label=f"{product['code']}/{body.lot_no}",
            details={"viscosity": body.viscosity, "lot_no": body.lot_no},
        )
        connection.commit()
        result = viscosity_service.analyze_product(
            connection, product, year=reading_year
        )
        result["new_reading"] = {
            "lot_no": body.lot_no,
            "viscosity": body.viscosity,
            "year": reading_year,
            "status": verdict["status"],
            "side": verdict["side"],
            "reasons": verdict["reasons"],
        }
        return result

    # ---- manager: 제품 설정 + 측정 삭제 + export -------------------------
    @mgr_router.post("/viscosity/products")
    def viscosity_create_product(
        body: ViscosityProductCreateBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        current_user = get_current_user(request, required=False)
        if viscosity_service.get_product_by_code(connection, body.code):
            raise HTTPException(status_code=409, detail=f"이미 존재하는 코드입니다: {body.code}")
        # 반제품 코드는 레시피 제품명과 연동되는 키 — 자유 입력 금지.
        # (배합 기록에 점도를 처음 등록하면 자동 생성되므로, 수동 추가는
        #  '첫 배합 전에 기준값·반응기 필수를 미리 세팅'하는 용도.)
        recipe_exists = connection.execute(
            "SELECT 1 FROM recipes WHERE product_name = ? LIMIT 1",
            (body.code.strip(),),
        ).fetchone()
        if not recipe_exists:
            raise HTTPException(
                status_code=400,
                detail=f"레시피에 없는 제품입니다: {body.code}. 레시피를 먼저 등록하세요.",
            )
        cur = connection.execute(
            """
            INSERT INTO viscosity_products
                (code, name, target, lower_limit, upper_limit, sigma_k, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 1, ?)
            """,
            (
                body.code.strip(),
                body.name.strip(),
                body.target,
                body.lower_limit,
                body.upper_limit,
                body.sigma_k,
                utc_now_text(),
            ),
        )
        product_id = int(cur.lastrowid)
        write_audit_log(
            connection,
            action="viscosity_product_create",
            actor=current_user,
            target_type="viscosity_product",
            target_id=str(product_id),
            target_label=body.code,
            details={"name": body.name},
        )
        connection.commit()
        return viscosity_service.get_product(connection, product_id)

    @mgr_router.patch("/viscosity/products/{product_id}")
    def viscosity_update_product(
        product_id: int,
        body: ViscosityProductUpdateBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        current_user = get_current_user(request, required=False)
        product = _require_product(connection, product_id)
        # reactor-ownership: use_reactor 의 소유가 recipes 로 이전되어 이 라우트는 더 이상
        # use_reactor 를 쓰지 않는다. 본문 필드는 API 호환을 위해 그대로 받되 무시한다
        # — 반응기 설정은 레시피(PUT /recipes/{id}/use-reactor)에서 변경한다.
        connection.execute(
            """
            UPDATE viscosity_products
            SET name = ?, target = ?, lower_limit = ?, upper_limit = ?,
                sigma_k = ?, rpm = ?, temperature = ?, remind_daily = ?,
                is_active = ?
            WHERE id = ?
            """,
            (
                body.name.strip(),
                body.target,
                body.lower_limit,
                body.upper_limit,
                body.sigma_k,
                body.rpm,
                body.temperature,
                1 if body.remind_daily else 0,
                1 if body.is_active else 0,
                product_id,
            ),
        )
        write_audit_log(
            connection,
            action="viscosity_product_update",
            actor=current_user,
            target_type="viscosity_product",
            target_id=str(product_id),
            target_label=product["code"],
            details={
                "target": body.target,
                "lower_limit": body.lower_limit,
                "upper_limit": body.upper_limit,
                "sigma_k": body.sigma_k,
                "remind_daily": body.remind_daily,
                # use_reactor 는 이제 recipes 소유 — 여기서 받아도 기록하지 않는다(무시).
                "is_active": body.is_active,
            },
        )
        connection.commit()
        return viscosity_service.get_product(connection, product_id)

    @mgr_router.delete("/viscosity/readings/{reading_id}")
    def viscosity_delete_reading(
        reading_id: int,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        current_user = get_current_user(request, required=False)
        row = connection.execute(
            "SELECT id, product_id, lot_no FROM viscosity_readings WHERE id = ?",
            (reading_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="측정 기록을 찾을 수 없습니다.")
        connection.execute("DELETE FROM viscosity_readings WHERE id = ?", (reading_id,))
        write_audit_log(
            connection,
            action="viscosity_reading_delete",
            actor=current_user,
            target_type="viscosity_reading",
            target_id=str(reading_id),
            target_label=str(row["lot_no"]),
        )
        connection.commit()
        return {"deleted": reading_id}

    @mgr_router.post("/viscosity/readings/{reading_id}/exclude")
    def viscosity_exclude_reading(
        reading_id: int,
        body: ViscosityExcludeBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """측정 1건을 통계에서 제외(삭제 아님) — 책임자 전용. 이상 하나가 σ 를 스스로
        오염시키지 않도록, 기록은 남기되 평균/σ/관리한계/추세/집계에서 뺀다."""
        current_user = get_current_user(request, required=False)
        try:
            result = viscosity_service.exclude_reading(
                connection,
                reading_id,
                body.reason,
                by=current_user,
                now=utc_now_text(),
            )
        except ValueError:
            raise HTTPException(status_code=400, detail="제외 사유를 입력하세요.")
        if result is None:
            raise HTTPException(status_code=404, detail="측정 기록을 찾을 수 없습니다.")
        connection.commit()
        return {"ok": True, "id": reading_id}

    @mgr_router.post("/viscosity/readings/{reading_id}/include")
    def viscosity_include_reading(
        reading_id: int,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """측정 1건의 통계 제외를 해제 — 책임자 전용. 다시 통계에 포함된다."""
        current_user = get_current_user(request, required=False)
        result = viscosity_service.include_reading(
            connection,
            reading_id,
            by=current_user,
            now=utc_now_text(),
        )
        if result is None:
            raise HTTPException(status_code=404, detail="측정 기록을 찾을 수 없습니다.")
        connection.commit()
        return {"ok": True, "id": reading_id}

    @mgr_router.get("/viscosity/products/{product_id}/export")
    def viscosity_export(
        product_id: int,
        granularity: str = "quarter",
        year: int | None = None,
        reactor: str | None = None,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        product = _require_product(connection, product_id)
        # GAP-2: Excel 의 판정·기간 요약은 화면과 같은 필터(단위/연도/반응기)로 계산해야 한다.
        # granularity/year/reactor 없이 전 연도·기본 단위로 판정하면 같은 측정이 화면=정상,
        # Excel=경고/이상으로 갈리는 불일치가 생긴다(제품별로 연도마다 점도 대역이 다르다는
        # 이 도메인 전제와 상충). product 상세 라우트와 동일하게 검증한다.
        if granularity not in _ALLOWED_GRANULARITY:
            granularity = "quarter"
        reactor = _parse_reactor(reactor)
        analysis = viscosity_service.analyze_product(
            connection, product, granularity=granularity, year=year, reactor=reactor
        )

        workbook = Workbook()

        # 시트 1) 측정 원본 — 화면 배합 기록 표와 같은 필드(한글 헤더).
        ws_readings = workbook.active
        ws_readings.title = "측정 원본"
        ws_readings.append(
            ["LOT", "측정일", "점도", "판정", "반응기", "메모", "배합 원료", "원료 LOT", "작성자"]
        )
        for it in analysis["readings"]:
            ws_readings.append([
                _xlsx_safe(it["lot_no"]),
                it["measured_date"],
                it["viscosity"],
                _STATUS_LABEL.get(it["status"], it["status"]),
                it["reactor"],
                _xlsx_safe(it["memo"] or ""),
                _xlsx_safe(it["recipe_material"] or ""),
                _xlsx_safe(it["material_lot"] or ""),
                _xlsx_safe(it["created_by"] or ""),
            ])

        # 시트 2) 기간 요약 — 요청한 단위/연도/반응기의 전체 기간 집계(화면은 60개만 표시하나
        # Excel 은 전체). 컬럼은 화면 기간별 표와 동일.
        ws_periods = workbook.create_sheet("기간 요약")
        ws_periods.append(
            ["기간", "건수", "평균", "전기대비", "표준편차", "최소", "최대", "이상", "경고"]
        )
        for p in analysis["periods"]:
            ws_periods.append([
                p["period"], p["count"], p["mean"], p["mean_delta"], p["std"],
                p["min"], p["max"], p["anomaly_count"], p["warn_count"],
            ])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        filename = f"viscosity_{product['code']}_{date.today().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type=_XLSX_MEDIA_TYPE,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @mgr_router.get("/viscosity/export-all")
    def viscosity_export_all(
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        """전체 반제품의 점도 측정을 한 flat 시트로 내보낸다 — 책임자 전용.

        판정(status)은 화면과 **반드시 같아야** 한다. GAP-2(제품별 export 주석 참고):
        점도 기준이 연도별로 다르므로 전 연도를 한데 섞어 재계산하면 같은 측정이
        화면=정상 / Excel=경고(이상)으로 갈린다. 따라서 list_products 로 전 반제품을
        돌고, 각 반제품마다 available_years 의 각 연도별로 analyze_product(year=연도)
        를 호출해 그 readings 에 붙은 status 를 그대로 가져온다(연도 없는 측정은
        available_years 에 걸리지 않으므로 자연 누락 — 화면의 연도 탭 기준과 동일).
        """
        current_user = get_current_user(request, required=False)
        workbook = Workbook()
        ws = workbook.active
        ws.title = "전체 측정"
        ws.append([
            "반제품 코드", "반제품명", "연도", "LOT", "측정일",
            "점도", "판정", "반응기", "메모", "등록자", "등록일시",
        ])
        products = viscosity_service.list_products(connection)
        row_count = 0
        for product in products:
            years = viscosity_service.available_years(connection, product["id"])
            for yr in years:
                analysis = viscosity_service.analyze_product(
                    connection, product, year=yr
                )
                # analyze_product 의 item 에는 created_at 이 없다(서비스 공개 키 아님).
                # 같은 연도 표본의 created_at 만 별도로 모아 매핑한다.
                ids = [r["id"] for r in analysis["readings"]]
                created_map: dict[int, str] = {}
                if ids:
                    placeholders = ",".join("?" for _ in ids)
                    crows = connection.execute(
                        f"SELECT id, created_at FROM viscosity_readings WHERE id IN ({placeholders})",
                        ids,
                    ).fetchall()
                    created_map = {int(r["id"]): (r["created_at"] or "") for r in crows}
                for r in analysis["readings"]:
                    ws.append([
                        product["code"],
                        product["name"],
                        yr,
                        _xlsx_safe(r["lot_no"]),
                        r["measured_date"],
                        r["viscosity"],
                        _STATUS_LABEL.get(r["status"], ""),
                        r["reactor"],
                        _xlsx_safe(r["memo"] or ""),
                        _xlsx_safe(r["created_by"] or ""),
                        created_map.get(int(r["id"]), ""),
                    ])
                    row_count += 1

        write_audit_log(
            connection,
            action="viscosity_exported_all",
            actor=current_user,
            target_type="viscosity_reading",
            target_id=None,
            target_label=f"전체 {row_count}건",
            details={"row_count": row_count, "product_count": len(products)},
        )
        connection.commit()

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        filename = f"viscosity_all_{date.today().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type=_XLSX_MEDIA_TYPE,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return op_router, mgr_router
