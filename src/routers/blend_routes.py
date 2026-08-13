"""배합 실적(잉크 계량 재구축) 라우트 — DHR Generator 이식.

접근: 점도 화면과 동일하게 로그인 없이 사내 공용 단말에서 사용. 작성자는 로그인
사용자가 있으면 그 이름, 없으면 '현장' 으로 기록.

Plan:   docs/01-plan/features/blend-overhaul.plan.md
Design: docs/02-design/features/blend-overhaul.design.md

Endpoints:
    GET    /blend/recipes                     배합용 레시피 목록 (최신 개정판만)
    GET    /blend/recipes/{id}?total=...      비율·이론량 환산 (개정 자동 귀결)
    GET    /blend/next-lot                    저장 시 부여될 제품 LOT 미리보기
    GET    /blend/workers                     작업자 목록(필터용)
    GET    /blend/analysis                    배합 분석 통합 집계(/insight 전용)
    GET    /blend/analysis/export             배합 분석 리포트 Excel(5시트)
    GET    /blend/material-usage              자재별 사용량 집계
    GET    /blend/product-usage               제품별 배치 빈도
    GET    /blend/batch-details[/export]      배치 상세(+Excel)
    POST   /blend/records                     배합 실적 저장 (작업자 세션 필요)
    POST   /blend/records/bulk                일괄 생성
    GET    /blend/records                     기록 조회(필터)
    GET    /blend/records/export-all          전체 Excel 백업
    GET    /blend/records/dhr-batch           배합일지 일괄 PDF (한 파일로 병합)
    GET    /blend/records/dhr-zip             배합일지 ZIP (반제품명 폴더로 묶음)
    GET    /blend/records/{id}                상세(배합상세+편차+점도)
    PUT    /blend/records/{id}                전체 수정 (책임자 전용)
    DELETE /blend/records/{id}                기록 취소/삭제(soft/hard 모두 책임자 전용)
    POST   /blend/records/{id}/restore        soft 취소 복원 (책임자 전용)
    POST   /blend/records/{id}/viscosity      점도 등록(배합 연계 — 점도 화면의 저장 경로)
    POST   /blend/records/{id}/approve        결재 기록 (책임자 전용, 현장 미사용)
    GET    /blend/records/{id}/export         실적서 Excel
    GET    /blend/records/{id}/pdf            배합일지 PDF(?sign=1 서명 합성)
"""

import io
import json
import logging
import sqlite3
from datetime import datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse

from ..auth import (
    authenticate_manager_worker,
    authenticate_user,
    get_current_user,
    has_access_level,
    require_access_level,
)
from ..blend_session import current_blend_worker, touch_worker_session
from ..config import CANCELED_RETENTION_DAYS
from ..db import get_db, utc_now_text, write_audit_log
from ..limiter import limiter
from ..services import blend_service, dhr_cache, dhr_excel, dhr_pdf, record_delete_service, viscosity_service
from .models import (
    BlendApprovalBody,
    BlendBatchDiscardBody,
    BlendBulkBody,
    BlendContinuousBody,
    BlendCreateBody,
    BlendViscosityBody,
    actor_name,
)


_ZIP_INVALID_CHARS = '\\/:*?"<>|'


def _family_folder_name(connection, product_name):
    """ZIP 폴더용 가족 이름 — 1차(-1)/최종(2차) 레시피를 한 폴더로 묶는다(2026-07-23).

    우선순위: ①이 반제품을 stage1 로 참조하는 2차 레시피가 있으면 그 이름
    ②이름이 "-1" 로 끝나고 접미사 제거한 이름의 레시피가 실존하면 그 이름
    ③그 외는 자기 이름. 현황 탭의 1차/2차 가족 묶음과 같은 기준.
    """
    name = (product_name or "").strip()
    if not name:
        return name
    try:
        row = connection.execute(
            """
            SELECT s2.product_name FROM recipes s1
            JOIN recipes s2 ON s2.stage1_recipe_id = s1.id
            WHERE s1.product_name = ? LIMIT 1
            """,
            (name,),
        ).fetchone()
        if row and row["product_name"]:
            return str(row["product_name"])
        if name.endswith("-1"):
            base = name[:-2]
            hit = connection.execute(
                "SELECT 1 FROM recipes WHERE product_name = ? LIMIT 1", (base,)
            ).fetchone()
            if hit:
                return base
    except sqlite3.OperationalError:
        pass
    return name


def _sanitize_zip_name(value: Any, fallback: str = "") -> str:
    """zip 경로용 파일/폴더명 정리 — Windows 금지문자(\\ / : * ? " < > |)를 '_' 로 치환.

    앞뒤 공백을 제거하고, 그래도 비면 fallback 을 돌려준다(빈 폴더/파일명 방지).
    """
    text = str(value if value is not None else "").strip()
    for ch in _ZIP_INVALID_CHARS:
        text = text.replace(ch, "_")
    text = text.strip()
    return text or fallback


logger = logging.getLogger(__name__)

# 저장 멱등 키의 엔드포인트 구분 — 같은 request_id 라도 단건/다중 계량은 서로의
# 결과를 되돌려주지 않는다(blend_save_requests.endpoint 에 그대로 저장된다).
_SAVE_EP_SINGLE = "blend_create"
_SAVE_EP_CONTINUOUS = "blend_create_continuous"


def _flag_total_anomalies(
    connection: sqlite3.Connection,
    *,
    record_id: int,
    product_lot: str,
    total_amount: Any,
    recipe_id: int | None,
    rescale_count: int,
    current_user: dict[str, Any] | None,
) -> dict[str, Any]:
    """총량 플래그 2종을 기록에 남기고(저장 차단 없음) 감사 로그까지 쓴다 → 플래그 dict.

    ① oversize_total       — 현장 1회 배합 상한(25,000 g) 초과. 폐기 권장을 무시한
                             '그래도 증량' 경로는 계속 저장돼야 하므로 표시만 남긴다.
    ② total_bypass_suspect — 증량 이력이 없는데 총량이 레시피 기준 배합량을 크게 상회
                             (증량 승인 우회 의심). 판정 규칙은
                             blend_service.detect_total_bypass 참조.

    단건 생성·수정(PUT)·다중 계량이 같은 함수를 쓴다. 어느 쪽도 저장을 막지 않는다.
    """
    flags = blend_service.apply_total_flags(
        connection,
        record_id,
        total_amount,
        recipe_id=recipe_id,
        rescale_count=rescale_count,
    )
    if flags["oversize_total"]:
        saved_total = float(total_amount)
        write_audit_log(
            connection,
            action="blend_total_oversize",
            actor=current_user,
            target_type="blend_record",
            target_id=str(record_id),
            target_label=product_lot,
            details={
                "total_amount": saved_total,
                "limit_g": blend_service.BLEND_OVERSIZE_FLAG_G,
                "over_g": round(saved_total - blend_service.BLEND_OVERSIZE_FLAG_G, 2),
            },
        )
    if flags["total_bypass_suspect"]:
        base = float(flags["total_bypass_base"] or 0.0)
        total = float(total_amount)
        write_audit_log(
            connection,
            action="blend_total_bypass_suspect",
            actor=current_user,
            target_type="blend_record",
            target_id=str(record_id),
            target_label=product_lot,
            details={
                "total_amount": total,
                "base_total": base,
                "excess_g": round(total - base, 2),
                "excess_pct": round((total - base) / base * 100, 2) if base else None,
                "recipe_id": recipe_id,
            },
        )
    return flags


def build_router() -> APIRouter:
    router = APIRouter()

    def require_blend_worker(request: Request) -> str:
        worker = current_blend_worker(request)
        if not worker:
            raise HTTPException(status_code=401, detail="BLEND_WORKER_REQUIRED")
        touch_worker_session(request)
        return worker

    def _mask_manual_entry(request: Request, record: dict[str, Any]) -> dict[str, Any]:
        """수동 입력 표시(manual_entry)는 책임자 전용 — 비책임자 응답에서는 False 로 가린다.

        화면 가림이 아니라 응답 자체를 가려 API 직접 조회로도 노출되지 않는다.
        저장·감사 로그의 원본 값은 불변(조회 표시만 제한)."""
        user = get_current_user(request, required=False)
        if user and has_access_level(user, "manager"):
            return record
        record["manual_entry"] = False
        # 부재 사유·미확인 플래그도 책임자 전용 통제 정보 — 같은 기준으로 가린다.
        if "manual_absence_reason" in record:
            record["manual_absence_reason"] = None
        if "manual_unacked" in record:
            record["manual_unacked"] = None
        for d in record.get("details", []) or []:
            d["manual_entry"] = False
        return record

    def _mask_worker_manual_stats(request: Request, data: dict[str, Any]) -> dict[str, Any]:
        """작업자별 수동 입력 통계도 책임자 전용 — 기록 목록과 같은 기준으로 가린다.

        기록 단위에서는 '누가 손으로 넣었는지'를 _mask_manual_entry 가 가리는데, 분석
        화면의 품질 탭은 같은 사실을 이름과 함께 집계로 보여 주고 있었다(로그인 없이도).
        한쪽만 잠그면 통제가 반쪽이라 같은 선으로 맞춘다.

        0 이 아니라 None 으로 비운다 — 0 은 '수동 입력이 없었다'는 뜻이라 거짓말이 되고,
        None 은 '말할 수 없다'로 화면·엑셀·도구가 모두 '—' 로 낸다. 기간 전체 합계
        (summary.manual_records)와 추세의 저울 계량률은 사람을 지목하지 않으므로 그대로
        둔다 — 그건 이 화면이 존재하는 이유인 지표다.
        """
        quality = data.get("quality") or {}
        visible = bool(
            (user := get_current_user(request, required=False))
            and has_access_level(user, "manager")
        )
        quality["manual_visible"] = visible
        if not visible:
            for row in quality.get("by_worker") or []:
                row["manual_records"] = None
                row["manual_rate"] = None
        return data

    def _log_duplicate_save(request_id: str, record_ids: list[int], label: str | None) -> None:
        """중복 저장을 막고 첫 결과를 돌려줬음을 서버 로그에 남긴다.

        작업자에게는 성공으로 보이므로(의도한 동작), 무엇이 억제됐는지 흔적이 없으면
        "두 번 저장했는데 한 건뿐"이라는 현장 신고를 사후에 확인할 수 없다.
        DB 쪽 근거는 blend_save_requests 행(request_id → record_ids, created_at)이 남긴다.
        감사 로그가 아니라 서버 로그를 쓰는 이유: 새 감사 action 은 한글 라벨표
        (static/js/admin_users.js AUDIT_GROUPS)와 짝을 맞춰야 하고, 그 표는 이 작업의
        수정 범위 밖이다(tests/test_audit_action_labels.py 가 짝을 강제한다).
        """
        logger.info(
            "blend save deduplicated: request_id=%s records=%s lot=%s",
            request_id, record_ids, label,
        )

    def _resolve_duplicate_save(
        connection: sqlite3.Connection,
        request: Request,
        request_id: str,
        endpoint: str,
    ) -> dict[str, Any]:
        """동시 요청 경합에서 진 쪽의 응답 — 먼저 커밋된 저장 결과를 그대로 돌려준다.

        이 저장은 이미 롤백된 상태로 호출된다. 같은 request_id 인데 결과를 찾을 수 없으면
        (다른 엔드포인트가 그 id 를 선점한 경우 등) 409 로 되돌려 중복 생성을 막는다.
        """
        prior = blend_service.lookup_save_request(connection, request_id, endpoint)
        if not prior:
            raise HTTPException(
                status_code=409,
                detail="이미 처리 중이거나 처리된 저장 요청입니다. 배합 기록을 확인하세요.",
            )
        if endpoint == _SAVE_EP_SINGLE:
            existing = blend_service.get_blend_record(connection, prior[0])
            if not existing:
                raise HTTPException(
                    status_code=409,
                    detail="이미 처리된 저장 요청입니다. 배합 기록을 확인하세요.",
                )
            return _mask_manual_entry(request, existing)
        return _continuous_result(connection, prior)

    def _continuous_result(
        connection: sqlite3.Connection, ids: list[int]
    ) -> dict[str, Any]:
        """다중 계량 저장 응답(기존 필드 그대로) — 멱등 재응답에서도 동일 형태를 쓴다."""
        lots = []
        for rid in ids:
            rec = blend_service.get_blend_record(connection, rid)
            lots.append(rec["product_lot"] if rec else None)
        return {"created": len(ids), "ids": ids, "product_lots": lots}

    def _audit_dhr_export(
        connection: sqlite3.Connection,
        request: Request,
        *,
        fmt: str,
        record_ids: list[int],
        target_id: str | None = None,
        target_label: str | None = None,
    ) -> None:
        """규제 문서(DHR) 출력·다운로드 행위를 감사한다(GAP-4).

        누가(책임자 또는 현장) 언제 어떤 배합일지를 어느 형식으로 출력·배포했는지 남긴다.
        actor 는 로그인 사용자(없으면 현장=None). 감사만 커밋한다(응답은 스트리밍 산출물).
        """
        current_user = get_current_user(request, required=False)
        write_audit_log(
            connection,
            action="dhr_exported",
            actor=current_user,
            target_type="blend_record",
            target_id=target_id,
            target_label=target_label,
            details={
                "format": fmt,
                "count": len(record_ids),
                "record_ids": record_ids[:200],
                "actor_name": actor_name(current_user) if current_user else "현장",
            },
        )
        connection.commit()

    @router.get("/blend/recipes")
    def blend_recipes(
        dhr: bool = Query(default=False),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        # dhr=True: DHR 전용 레시피(일괄 배합일지 생성용). 기본은 일반 레시피.
        return {"items": blend_service.list_blend_recipes(connection, dhr=dhr)}

    @router.get("/blend/recipes/{recipe_id}")
    def blend_recipe_detail(
        recipe_id: int,
        total: float | None = Query(default=None, gt=0),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        result = blend_service.get_recipe_for_blend(connection, recipe_id, total)
        if not result:
            raise HTTPException(status_code=404, detail="레시피를 찾을 수 없습니다.")
        return result

    @router.get("/blend/next-lot")
    def blend_next_lot(
        product: str = Query(..., min_length=1),
        date: str | None = Query(default=None),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """저장 시 실제 부여될 product_lot 미리보기({제품명}{YYMMDD}{순번:02d})."""
        work_date = date or utc_now_text()[:10]
        return {"next_lot": blend_service.generate_product_lot(connection, product, work_date)}

    @router.get("/blend/material-usage")
    def blend_material_usage(
        start_date: str = "",
        end_date: str = "",
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """배합 기록 기반 자재 사용 분석(기간별 자재별 실제/이론 사용량·건수)."""
        return blend_service.material_usage(connection, start_date or None, end_date or None)

    @router.get("/blend/analysis")
    def blend_analysis(
        request: Request,
        start_date: str = "",
        end_date: str = "",
        bucket: str = "month",
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """배합 분석 화면 한 판 — 지표(전기 대비)·기간 추세·제품·자재·품질을 한 번에."""
        return _mask_worker_manual_stats(request, blend_service.analysis(
            connection, start_date or None, end_date or None, bucket
        ))

    @router.get("/blend/analysis/export")
    def blend_analysis_export(
        request: Request,
        start_date: str = "",
        end_date: str = "",
        bucket: str = "month",
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        """배합 분석 리포트 Excel — 화면에서 본 그대로 5시트(요약/추세/제품/자재/품질)."""
        # 엑셀도 같은 기준으로 가린다 — 화면만 잠그면 내려받기로 그대로 새어 나간다.
        data = _mask_worker_manual_stats(request, blend_service.analysis(
            connection, start_date or None, end_date or None, bucket
        ))
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        head = Font(bold=True)
        title_font = Font(bold=True, size=14)
        wb = Workbook()

        def _sheet(ws, headers: list[str], widths: list[int]) -> None:
            ws.append(headers)
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).font = head
            for col, w in enumerate(widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        s = data["summary"]
        rng = data["range"]
        period = f"{rng['start'] or '전체'} ~ {rng['end'] or '전체'}"
        prev = data.get("previous")

        ws = wb.active
        ws.title = "요약"
        ws["A1"] = "배합 분석 리포트"
        ws["A1"].font = title_font
        ws["A2"] = f"분석 기간: {period}" + (
            f" ({rng['days']}일)" if rng.get("days") else ""
        )
        ws["A3"] = (
            f"비교 기간: {prev['start']} ~ {prev['end']}"
            if prev
            else "비교 기간: 없음 (시작·종료를 모두 지정하면 전기 대비를 계산합니다)"
        )
        ws["A4"] = f"생성 시각: {utc_now_text()} (UTC)"
        if data.get("scale_since"):
            ws["A5"] = (
                f"저울 계량률 기준: 저울 도입 {data['scale_since']} 이후 "
                f"{s.get('scale_base_records', 0)}건 — 그 전 기록에는 수동 입력 표시가 없어"
                " 저울로 잰 것과 구분되지 않습니다."
            )
        ws.append([])
        # 지표는 '값 / 전기 / 증감'을 나란히 둔다 — 리포트로 뽑았을 때 숫자 하나만
        # 남으면 읽는 사람이 많고 적음을 판단할 근거가 없다.
        ws.append(["지표", "이번 기간", "직전 기간", "증감"])
        for c in range(1, 5):
            ws.cell(row=ws.max_row, column=c).font = head

        def _delta(cur: Any, prv: Any, unit: str = "") -> str:
            if prv is None or not isinstance(cur, (int, float)):
                return "-"
            diff = round(float(cur) - float(prv), 1)
            sign = "+" if diff > 0 else ""
            return f"{sign}{diff:g}{unit}"

        rows = [
            ("배합 건수", s["records"], s["records_prev"], "건"),
            ("총 생산량(kg)", round(s["total_weight_g"] / 1000, 2),
             None if s["total_weight_prev"] is None else round(s["total_weight_prev"] / 1000, 2),
             "kg"),
            ("제품 종수", s["product_count"], s["product_count_prev"], "종"),
            ("자재 종수", s["material_count"], s["material_count_prev"], "종"),
            # 저울 도입 전 기록에는 수동 입력 표시가 없어 계량률이 뜻을 갖지 못한다 —
            # 표본이 없으면 숫자를 지어내지 않고 '해당 없음'으로 적는다.
            ("저울 계량률(%)",
             s["scale_rate"] if s["scale_rate"] is not None else "해당 없음",
             s["scale_rate_prev"], "%p"),
            ("취소율(%)", s["cancel_rate"], s["cancel_rate_prev"], "%p"),
        ]
        for label, cur, prv, unit in rows:
            ws.append([label, cur, prv if prv is not None else "-", _delta(cur, prv, unit)])
        ws.append([])
        ws.append(["보조 지표", "값"])
        for c in (1, 2):
            ws.cell(row=ws.max_row, column=c).font = head
        for label, value in [
            ("수동 입력 배합", s["manual_records"]),
            ("취소 배합", s["canceled_records"]),
            ("증량 적용 배합", s["rescale_records"]),
            ("1회 상한 초과 배합", s["oversize_records"]),
            ("투입 로스 보정 누계(g)", s["loss_comp_total_g"]),
        ]:
            ws.append([label, value])
        for col, w in [("A", 24), ("B", 16), ("C", 16), ("D", 12)]:
            ws.column_dimensions[col].width = w
        ws["A1"].alignment = Alignment(vertical="center")

        ws = wb.create_sheet("추세")
        _sheet(
            ws,
            ["구간", "배합 건수", "생산량(kg)", "수동 입력", "취소", "저울 계량률(%)", "비고"],
            [14, 12, 14, 12, 10, 16, 20],
        )
        for t in data["trend"]:
            ws.append([
                t["bucket"], t["records"], round(t["weight_g"] / 1000, 2),
                t["manual_records"], t["canceled_records"],
                t["scale_rate"] if t["scale_rate"] is not None else "해당 없음",
                # 양 끝 구간은 대개 잘려 있다 — 표시가 없으면 생산 급감으로 읽힌다.
                "기간이 잘린 구간(다른 구간과 길이가 다름)" if t.get("partial") else "",
            ])

        ws = wb.create_sheet("제품")
        _sheet(
            ws,
            ["제품", "배치 수", "총 생산량(kg)", "생산 비중(%)", "최근 작업일"],
            [20, 10, 16, 14, 14],
        )
        for p in data["products"]:
            ws.append([
                p["product_name"], p["batch_count"], round(p["total_amount"] / 1000, 2),
                p["share"], p["last_work_date"] or "",
            ])

        ws = wb.create_sheet("자재")
        _sheet(
            ws,
            ["자재", "실제 사용량(kg)", "이론량(kg)", "차이(g)", "투입 배치 수",
             "소비 비중(%)", "로스 보정 누계(g)"],
            [20, 18, 16, 12, 14, 14, 18],
        )
        for m in data["materials"]:
            ws.append([
                m["material_name"], round(m["total_actual"] / 1000, 3),
                round(m["total_theory"] / 1000, 3),
                round(m["total_actual"] - m["total_theory"], 2),
                m["usage_count"], m["share"], m["loss_comp_g"],
            ])

        ws = wb.create_sheet("품질")
        ws.append(["작업자별 실적·이상"])
        ws.cell(row=ws.max_row, column=1).font = head
        manual_visible = data["quality"].get("manual_visible", True)
        if not manual_visible:
            ws.append(["수동 입력 열은 책임자만 볼 수 있습니다 — 담당자 계정으로 내려받아 비어 있습니다."])
        _sheet(
            ws,
            ["작업자", "완료", "생산량(kg)", "제품 종수", "수동 입력", "수동 비율(%)", "취소"],
            [16, 10, 14, 12, 12, 14, 10],
        )
        # 가려진 값은 0 이 아니라 빈 칸 — 0 으로 채우면 '수동 입력이 없었다'로 읽힌다.
        blank = "" if not manual_visible else 0
        # 생산 실적과 이상 통계는 대상이 조금 다르다(취소만 있는 사람 / 완료만 있는 사람)
        # — 화면과 같은 합집합 기준으로 쓴다.
        production = {w["worker"]: w for w in data["workers"]}
        seen: set[str] = set()
        for w in data["quality"]["by_worker"]:
            name = w["worker"] or ""
            seen.add(name)
            p = production.get(name, {})
            ws.append([
                name, w["records"], round(p.get("total_amount", 0) / 1000, 2),
                p.get("product_count", 0),
                w["manual_records"] if w["manual_records"] is not None else "",
                w["manual_rate"] if w["manual_rate"] is not None else "",
                w["canceled_records"],
            ])
        for name, p in production.items():
            if name not in seen:
                ws.append([
                    name, p["records"], round(p["total_amount"] / 1000, 2),
                    p["product_count"], blank, blank, 0,
                ])
        ws.append([])
        ws.append(["자재별 (수동 입력이 있었던 자재만)"])
        ws.cell(row=ws.max_row, column=1).font = head
        ws.append(["자재", "계량 행", "수동 입력", "수동 비율(%)"])
        for c in range(1, 5):
            ws.cell(row=ws.max_row, column=c).font = head
        for m in data["quality"]["by_material"]:
            ws.append([m["material_name"], m["rows"], m["manual_rows"], m["manual_rate"]])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from datetime import date as _date

        filename = f"blend-analysis-{_date.today().isoformat()}.xlsx"
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @router.get("/blend/product-usage")
    def blend_product_usage(
        start_date: str = "",
        end_date: str = "",
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """제품별 배합 빈도 분석(기간 내 제품별 배치 수·총 배합량·최근 작업일)."""
        return blend_service.product_usage(connection, start_date or None, end_date or None)

    @router.get("/blend/mistake-stats")
    def blend_mistake_stats(
        start_date: str = "",
        end_date: str = "",
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """작업자·자재별 이상(수동 입력·취소) 통계 — 편차 강제로 편차 대신 이 신호를 본다."""
        return blend_service.mistake_stats(connection, start_date or None, end_date or None)

    @router.get("/blend/batch-details")
    def blend_batch_details(
        start_date: str = "",
        end_date: str = "",
        product: str = "",
        limit: int = Query(default=2000, ge=1, le=10000),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """배치 상세(자재별 비율·이론량·실제량·편차 평면 목록, 작업일 역순)."""
        return blend_service.batch_details(
            connection, start_date or None, end_date or None, product or None, limit,
        )

    @router.get("/blend/batch-details/export")
    def blend_batch_details_export(
        start_date: str = "",
        end_date: str = "",
        product: str = "",
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        """배치 상세를 Excel 한 시트로 — 배합 분석·데이터 이관용."""
        result = blend_service.batch_details(
            connection, start_date or None, end_date or None, product or None, limit=10000,
        )
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "배합 상세"
        headers = [
            "작업일", "제품LOT", "제품", "작업자", "자재코드", "자재명",
            "자재LOT", "비율(%)", "이론량(g)", "실제량(g)", "편차(g)",
        ]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        for it in result["items"]:
            ws.append([
                it["work_date"], it["product_lot"], it["product_name"], it["worker"],
                it.get("material_code") or "", it["material_name"], it.get("material_lot") or "",
                it["ratio"], it["theory_amount"], it["actual_amount"], it["variance"],
            ])
        widths = [12, 18, 16, 10, 12, 18, 14, 9, 11, 11, 9]
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = w
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from datetime import date as _date
        filename = f"blend-batch-details-{_date.today().isoformat()}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @router.get("/blend/workers")
    def blend_workers(connection: sqlite3.Connection = Depends(get_db)) -> dict[str, Any]:
        return {"items": blend_service.list_workers(connection)}

    @router.get("/blend/material-lot-trace")
    def blend_material_lot_trace(
        lot: str = Query(..., min_length=1, max_length=100),
        limit: int = Query(default=500, ge=1, le=2000),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """자재 LOT 역추적 — 이 LOT 이 들어간 배합 기록·상세(리콜 추적)."""
        return blend_service.trace_material_lot(connection, lot, limit=limit)

    @router.get("/blend/recent-product-lots")
    def blend_recent_product_lots(
        names: str = Query(default=""),
        limit: int = Query(default=5),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """반제품 원료 LOT 자동 제안용 — names 에 든 제품(반제품)명별 최근 product_lot.

        2단 제조(1차 중간체 → 2차 최종)에서 2차 배합의 원료 행(=1차 반제품명) 자재 LOT 칸에
        1차 배합의 제품 LOT 을 넣어 1차→2차 LOT 연결을 남기는 데 쓴다. 완료(completed) 기록만,
        최신순(id DESC), 반제품별 limit 개, NULL/빈 LOT·중복 제거. 기록 없는 이름은 키 자체 제외.

        각 항목은 {lot, total} — total 은 그 1차 배합 기록의 total_amount(1차 배치 총량)로,
        반응기 이월(carry-over) 입력이 채울 기준값으로 화면에 같이 보여준다.

        names: 쉼표 구분 제품명(빈 항목 제거, 최대 50 — 초과분 무시).
        limit: 반제품당 LOT 개수(1~20 클램프, 기본 5). 0 이하→1, 20 초과→20.
        """
        # limit 클램프 — FastAPI ge/le 가 422 로 거부하는 대신 1~20 으로 끌어온다(스펙 C.2).
        limit = max(1, min(20, limit))
        # names 파싱: strip → 빈 항목 제거 → 순서 보존 중복 제거 → 50개 초과분 무시.
        raw_names = [n.strip() for n in (names or "").split(",")]
        seen: set[str] = set()
        name_list: list[str] = []
        for n in raw_names:
            if n and n not in seen:
                seen.add(n)
                name_list.append(n)
            if len(name_list) >= 50:
                break
        items: dict[str, list[dict[str, Any]]] = {}
        if not name_list:
            return {"items": items}
        # IN (?, ?, ...) 자리표시자 — 제품명 수만큼. total_amount 까지 함께 가져온다(이월 채움용).
        placeholders = ",".join("?" for _ in name_list)
        rows = connection.execute(
            f"SELECT product_name, product_lot, total_amount FROM blend_records "
            f"WHERE product_name IN ({placeholders}) AND status = 'completed' "
            f"ORDER BY id DESC",
            name_list,
        ).fetchall()
        # 반제품별 최신순(id DESC 로 이미 정렬됨)로 순회하며 LOT 단위 중복 제거해 limit 개씩 채운다.
        for r in rows:
            lot_val = (r["product_lot"] or "").strip()
            if not lot_val:
                continue
            lots = items.setdefault(r["product_name"], [])
            if any(it["lot"] == lot_val for it in lots):
                continue
            if len(lots) < limit:
                lots.append({"lot": lot_val, "total": float(r["total_amount"] or 0)})
        return {"items": items}

    @router.get("/blend/product-lot-exists")
    def blend_product_lot_exists(
        name: str = Query(default=""),
        lot: str = Query(default=""),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """반제품 원료 LOT 확인용 — 주어진 제품명/LOT 의 완료 기록 존재 여부.

        2단 제조(1차 중간체 → 2차 최종)에서 2차 배합의 원료 행(=1차 반제품명) 자재 LOT 칸에
        입력된 값이 실제 1차 배합의 완료(completed) 기록에 존재하는 LOT 인지 확인한다.
        **차단용이 아니다**(2026-08-04) — 화면은 없을 때 가벼운 확인 창만 띄우고, 진행하면
        서버가 blend_lot_acks 에 대사용 기록을 남긴다. 양쪽 값은 strip 후 정확히 일치하는
        product_name·product_lot 이며 status='completed' 인 행이 있어야 exists=true.

        name: 검증할 제품(반제품)명(빈 값 → exists=false).
        lot: 검증할 LOT(빈 값 → exists=false).
        """
        name = (name or "").strip()
        lot = (lot or "").strip()
        if not name or not lot:
            return {"exists": False}
        # 정확 일치 — strip 된 name/lot 으로 파라미터화 WHERE. status='completed' 한정.
        row = connection.execute(
            "SELECT 1 FROM blend_records "
            "WHERE product_name = ? AND product_lot = ? AND status = 'completed' LIMIT 1",
            (name, lot),
        ).fetchone()
        return {"exists": row is not None}

    @router.get("/blend/records")
    def blend_records(
        request: Request,
        start_date: str | None = None,
        end_date: str | None = None,
        worker: str | None = None,
        product: str | None = None,
        search: str | None = None,
        limit: int = Query(default=500, ge=1, le=1000),
        include_canceled: bool = Query(default=False),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        # 최신 limit 건만 반환(기본 500). 날짜·작업자·검색 필터가 범위를 좁히는 도구다.
        # total_available(전체 M)로 상한 도달 여부를 표면화 — '표시 N / 전체 M' 안내용.
        # include_canceled: 취소된 기록까지 함께 조회(취소분을 다시 열어 복원하는 유일한 경로).
        items = blend_service.list_blend_records(
            connection,
            start_date=start_date,
            end_date=end_date,
            worker=worker,
            product=product,
            search=search,
            limit=limit,
            include_canceled=include_canceled,
        )
        for item in items:
            _mask_manual_entry(request, item)
        total_available = blend_service.count_blend_records(
            connection,
            start_date=start_date,
            end_date=end_date,
            worker=worker,
            product=product,
            search=search,
            include_canceled=include_canceled,
        )
        # 취소분을 숨기고 보는 게 기본인데, 몇 건이 숨겨졌는지 화면이 말할 방법이 없었다.
        # ('취소된 기록 포함'을 켜 보기 전에는 0건인지 12건인지 알 수 없다.)
        canceled_hidden = 0
        if not include_canceled:
            canceled_hidden = blend_service.count_blend_records(
                connection,
                start_date=start_date,
                end_date=end_date,
                worker=worker,
                product=product,
                search=search,
                include_canceled=True,
            ) - total_available
        return {
            "items": items,
            "total": len(items),
            "total_available": total_available,
            "truncated": total_available > len(items),
            "limit": limit,
            "canceled_hidden": max(0, canceled_hidden),
        }

    @router.get("/blend/records/export-all")
    def blend_export_all(
        request: Request,
        start_date: str | None = None,
        end_date: str | None = None,
        worker: str | None = None,
        # 목록 조회와 동일한 제품 필터 — 없어서 화면은 걸러 놓고 파일은 전 제품이 나왔다
        # (2026-08-05 전수 감사 R-12, '전체'는 기간·필터 안의 전체를 뜻해야 한다).
        product: str | None = None,
        search: str | None = None,
        include_canceled: bool = False,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        """전체(필터) 배합 기록을 한 시트로 — 데이터 백업·이관용.

        include_canceled=False(기본) 면 화면 기본 목록과 같이 취소 기록을 빼고,
        True 면 /status 의 '취소 포함' 체크와 같이 취소 기록까지 내려준다 — 화면과
        파일의 정합. 취소 행이 섞이면 구분이 안 되므로 '상태' 열 값은 한글로 표기.
        """
        records = blend_service.list_blend_records(
            connection, start_date=start_date, end_date=end_date,
            worker=worker, product=product, search=search, limit=10000,
            include_canceled=include_canceled,
        )
        _audit_dhr_export(
            connection, request, fmt="xlsx_all",
            record_ids=[int(r["id"]) for r in records],
            target_label=f"전체 {len(records)}건",
        )
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "배합기록"
        # 열 내용은 레거시 ink_name 컬럼 — 사용자 노출 라벨에 '잉크'는 금칙어(CLAUDE.md).
        headers = ["작업일", "제품LOT", "제품", "세부 품명", "작업자", "총량(g)", "저울", "상태", "비고"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        _status_label = {"completed": "완료", "canceled": "취소"}
        for r in records:
            ws.append([
                r["work_date"], r["product_lot"], r["product_name"], r.get("ink_name") or "",
                r["worker"], r["total_amount"], r.get("scale") or "",
                _status_label.get(r["status"], r["status"]), r.get("note") or "",
            ])
        widths = [12, 18, 16, 14, 10, 10, 10, 10, 24]
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = w
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from datetime import date as _date
        filename = f"blend-records-{_date.today().isoformat()}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @router.get("/blend/records/dhr-batch")
    def blend_dhr_batch(
        request: Request,
        ids: str = Query(...),
        sign: bool = Query(default=False),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        """선택한(또는 전체) 배합 기록의 배합일지를 한 PDF로 일괄 출력(최대 200건).

        sign=True 면 서명 합성, 기본은 빈 결재칸(서명 없음).
        """
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()][:200]
        fetched = [
            r for r in (blend_service.get_blend_record(connection, i) for i in id_list) if r
        ]
        # 취소된 기록은 일괄 출력에서 제외한다. '취소 포함'으로 조회한 뒤 전체 선택하면
        # 취소분이 정상 기록과 한 문서로 인쇄됐고, PIL 폴백 렌더러에는 취소 표식이 없어
        # 정상 문서처럼 보였다. 단건 출력은 의도적으로 계속 허용(표식이 찍힌다).
        records = [r for r in fetched if r.get("status") != "canceled"]
        skipped_canceled = len(fetched) - len(records)
        if not records:
            raise HTTPException(
                status_code=404,
                detail=("선택한 기록이 모두 취소된 기록입니다 — 일괄 출력 대상이 없습니다."
                        if skipped_canceled else "배합 기록을 찾을 수 없습니다."),
            )
        _audit_dhr_export(
            connection, request, fmt="pdf_batch" + ("_signed" if sign else ""),
            record_ids=[int(r["id"]) for r in records],
            target_label=f"{len(records)}건",
        )
        pdf_bytes = dhr_pdf.build_batch_dhr_pdf(records, sign=sign)
        from urllib.parse import quote
        utf8_name = quote(f"배합일지-{len(records)}건.pdf")
        disposition = (
            f"attachment; filename=\"dhr-batch.pdf\"; filename*=UTF-8''{utf8_name}"
        )
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": disposition},
        )

    @router.get("/blend/records/dhr-zip")
    def blend_dhr_zip(
        request: Request,
        ids: str = Query(...),
        sign: bool = Query(default=False),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        """선택한 배합 기록의 배합일지를 반제품(레시피)명 폴더로 묶어 ZIP 으로 내려준다(최대 200건).

        각 기록의 PDF 는 단건 출력(/pdf)과 동일한 경로로 생성한다 — 비서명본은 캐시 재사용,
        sign=True 면 서명 합성(캐시 미사용). 폴더는 반제품명 — 단, 1차(-1)와 최종(2차)은 가족으로 묶어 최종 이름 폴더 하나에 담는다. 파일은
        {제품LOT}.pdf. 한 폴더 안에서 LOT 이 겹치면 _2, _3… 을 붙인다. Windows 압축 풀기에서
        한글이 깨지지 않도록 UTF-8 로 기록한다. 존재하지 않는 id 는 전체를 실패시키지 않고
        누락.txt 에 모아 남긴다.
        """
        import zipfile

        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()][:200]
        records: list[dict[str, Any]] = []
        missing: list[int] = []
        canceled_lots: list[str] = []
        for i in id_list:
            rec = blend_service.get_blend_record(connection, i)
            if rec and rec.get("status") == "canceled":
                # 취소분은 넣지 않되, 무엇이 빠졌는지 ZIP 안에 남긴다(조용한 누락 금지).
                canceled_lots.append(str(rec.get("product_lot") or i))
            elif rec:
                records.append(rec)
            else:
                missing.append(i)
        if not records:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")

        _audit_dhr_export(
            connection, request, fmt="zip" + ("_signed" if sign else ""),
            record_ids=[int(r["id"]) for r in records],
            target_label=f"{len(records)}건",
        )

        def _pdf_bytes(record: dict[str, Any]) -> bytes:
            # 단건 /pdf 와 동일: 서명본은 매번 합성(캐시 안 함), 비서명본은 캐시 재사용.
            if sign:
                return dhr_pdf.build_scanned_dhr_pdf(record, sign=True)
            data = dhr_cache.get(record)
            if data is None:
                data = dhr_pdf.build_scanned_dhr_pdf(record, sign=False)
                dhr_cache.put(record, data)
            return data

        buf = io.BytesIO()
        used: dict[tuple[str, str], int] = {}
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for record in records:
                folder = _sanitize_zip_name(
                    _family_folder_name(connection, record.get("product_name")), "기타"
                )
                base = _sanitize_zip_name(record.get("product_lot"), f"blend_{record['id']}")
                key = (folder, base)
                n = used.get(key, 0) + 1
                used[key] = n
                name = base if n == 1 else f"{base}_{n}"
                zf.writestr(f"{folder}/{name}.pdf", _pdf_bytes(record))
            if missing:
                note = "누락된 기록 id (해당 기록을 찾을 수 없어 제외됨):\n" + "\n".join(
                    str(m) for m in missing
                ) + "\n"
                zf.writestr("누락.txt", note.encode("utf-8"))
            if canceled_lots:
                note = (
                    "취소된 기록이라 제외했습니다 (배합일지 출력 대상 아님):\n"
                    + "\n".join(canceled_lots) + "\n"
                )
                zf.writestr("제외-취소된기록.txt", note.encode("utf-8"))
        buf.seek(0)

        from datetime import date as _date
        from urllib.parse import quote
        ascii_name = f"dhr-{_date.today().strftime('%Y%m%d')}.zip"
        utf8_name = quote(f"배합일지-{len(records)}건.zip")
        disposition = (
            f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_name}"
        )
        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": disposition},
        )

    @router.get("/blend/records/{record_id}")
    def blend_record_detail(
        record_id: int,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        record["viscosity"] = viscosity_service.list_readings_for_blend(connection, record_id)
        # 공정 설명 줄 — 기록 당시 레시피(recipe_id 원본 보존)의 설명을 함께 표시
        record["steps"] = []
        if record.get("recipe_id"):
            try:
                rows = connection.execute(
                    "SELECT position, note FROM recipe_steps WHERE recipe_id = ? "
                    "ORDER BY position, id",
                    (record["recipe_id"],),
                ).fetchall()
                record["steps"] = [
                    {"position": int(r["position"]), "note": r["note"]} for r in rows
                ]
            except sqlite3.OperationalError:
                pass
        # 취소된 기록 — 사유·취소자·시각·자동 삭제 예정일을 상세에 실어준다(F15).
        # 취소 시 기록엔 status/updated_at 만 남고 사유·행위자는 감사 로그가 원본이라,
        # 스키마 변경 없이 로그에서 읽는다(과거 취소분도 소급 표시).
        if record.get("status") == "canceled":
            reason = None
            actor = None
            canceled_at = record.get("updated_at")
            row = connection.execute(
                "SELECT actor_display_name, actor_username, details_json, created_at "
                "FROM audit_logs WHERE action = 'blend_record_cancel' "
                "AND target_type = 'blend_record' AND target_id = ? "
                "ORDER BY id DESC LIMIT 1",
                (str(record_id),),
            ).fetchone()
            if row:
                try:
                    reason = (json.loads(row["details_json"] or "{}") or {}).get("reason")
                except (ValueError, TypeError):
                    reason = None
                actor = row["actor_display_name"] or row["actor_username"]
                canceled_at = row["created_at"] or canceled_at
            purge_at = None
            if CANCELED_RETENTION_DAYS > 0 and record.get("updated_at"):
                # 자동 정리 기준은 updated_at(취소가 마지막 쓰기) — purge_expired_canceled 와 동일.
                try:
                    base = datetime.strptime(record["updated_at"], "%Y-%m-%dT%H:%M:%SZ")
                    purge_at = (base + timedelta(days=CANCELED_RETENTION_DAYS)).strftime(
                        "%Y-%m-%dT%H:%M:%SZ"
                    )
                except ValueError:
                    purge_at = None
            record["cancel_info"] = {
                "reason": reason,
                "actor": actor,
                "canceled_at": canceled_at,
                "retention_days": CANCELED_RETENTION_DAYS,
                "purge_at": purge_at,
            }
        return _mask_manual_entry(request, record)

    @router.post("/blend/records/{record_id}/viscosity")
    def blend_add_viscosity(
        record_id: int,
        body: BlendViscosityBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """점도 등록(배합 실적 연계). UI 는 점도 관리 화면 한 곳 — 이 라우트가 그 화면의
        저장 경로다(blend_record_id 연계 포함). 배합/기록 화면에는 입력 폼을 두지 않는다."""
        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        current_user = get_current_user(request, required=False)
        actor = actor_name(current_user) if current_user else "현장"
        now = utc_now_text()
        # 점도 화면이 선택한 반제품(product_id)이 들어오면 그 제품으로 귀속한다(F13).
        # 없으면(레거시 호환) 배합 기록의 제품(레시피)명으로 자동 확보한다. product_id 경로가
        # 있어야 화면이 고른 반제품(스펙 있음)으로 들어가고, ensure_product_by_code 가 화면
        # 선택을 무시하고 유령 반제품(스펙 없음)을 조용히 만드는 사고가 막힌다.
        product = None
        if body.product_id is not None:
            product = viscosity_service.get_product(connection, body.product_id)
            if not product or not product.get("is_active"):
                raise HTTPException(
                    status_code=400,
                    detail="선택한 반제품을 찾을 수 없습니다 — 반제품을 다시 선택하세요.",
                )
        else:
            product = viscosity_service.ensure_product_by_code(
                connection, record["product_name"], record["product_name"], now
            )
            if not product:
                raise HTTPException(status_code=400, detail="제품명이 없어 점도를 등록할 수 없습니다.")
        # '사용한 PB' — 화면이 보정한 값(body.material_lot)이 있으면 그것을, 없으면
        # 배합 상세에서 자재명이 PB 인 행을 찾아 감지한다(detect_source_pb_lot).
        # 종전의 "무조건 첫 행" 가정은 PB 가 첫 계량 자재가 아닌 레시피에서 엉뚱한
        # 자재 LOT 을 박았다(2026-08-13 검토).
        if body.material_lot is not None and body.material_lot.strip():
            source_pb_lot, pb_method = body.material_lot.strip(), "manual"
        else:
            source_pb_lot, pb_method = viscosity_service.detect_source_pb_lot(
                record.get("details") or []
            )
        try:
            viscosity_service.add_reading(
                connection,
                product_id=product["id"],
                lot_no=record["product_lot"],
                viscosity=body.viscosity,
                measured_date=record["work_date"],
                memo=body.memo,
                recipe_material=record["product_name"],
                material_lot=source_pb_lot,
                created_by=actor,
                created_at=now,
                blend_record_id=record_id,
                reactor=record.get("reactor"),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(
                status_code=409,
                detail=f"이미 등록된 점도(LOT {record['product_lot']})가 있습니다.",
            )
        write_audit_log(
            connection,
            action="blend_viscosity_link",
            actor=current_user,
            target_type="blend_record",
            target_id=str(record_id),
            target_label=record["product_lot"],
            details={
                "product_code": product["code"],
                "viscosity": body.viscosity,
                "used_pb_lot": source_pb_lot,
                "used_pb_method": pb_method,
            },
        )
        connection.commit()
        record["viscosity"] = viscosity_service.list_readings_for_blend(connection, record_id)
        # 화면이 등록 직후 "사용한 PB: xxx (감지/수동)" 을 보여줄 수 있게 싣는다.
        record["used_pb"] = {"lot": source_pb_lot, "method": pb_method}
        return _mask_manual_entry(request, record)

    # ── 증량(rescale) 책임자 현장 인증 — 세션 생성 없이 자격 증명만 확인 ──
    # 저장 시 approval_id 로 소비되는 1회용 승인 토큰을 발급한다. 비밀번호 검증은
    # 기존 authenticate_user 를 재사용(해시 로직 중복 금지). management-login 과 동일
    # slowapi 레이트리밋(5/분) 으로 무차별 대입을 막는다.
    #
    # purpose: "rescale"(기본) | "manual".
    #   - "rescale": 초과 계량 증량 승인. 성공 감사=blend_rescale_approved.
    #   - "manual": 저울 전용 모드에서 이 배합에 한해 수기 입력 허용 승인.
    #     성공 감사=blend_manual_entry_approved. 승인 토큰(approval_id)은 반환하되
    #     증량처럼 저장 시 소비(used=1)를 강제하지 않는다 — 서버는 실제 입력이 저울인지
    #     손인지 구분할 수 없으므로, 통제는 '책임자 승인 + 기록의 수동 입력 표시(manual_entry)'
    #     로 이루어진다(기존 manual_entry 설계와 일관). 거부 감사는 증량과 동일하게
    #     blend_rescale_approve_denied 를 쓰되 details 에 purpose 를 남긴다.
    @router.post("/blend/manager-verify")
    @limiter.limit("5/minute")
    def blend_manager_verify(
        request: Request,
        body: dict[str, Any],
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        username = str(body.get("username") or "").strip()
        password = str(body.get("password") or "")
        # 알려진 두 목적으로 좁힌다('manual' 외 전부 'rescale' — 옛 분기와 동일 결과).
        # 이 값은 이제 감사 로그뿐 아니라 토큰 행(blend_rescale_approvals.purpose)에도
        # 기록되어, 소비 시 목적이 다른 토큰(수기입력 승인)을 증량으로 쓰지 못하게 한다.
        purpose = blend_service.normalize_approval_purpose(body.get("purpose"))
        is_manual = purpose == blend_service.MANUAL_PURPOSE
        # management-login 과 동일한 순서로 책임자 자격을 검증한다(해시 로직 중복 금지):
        # 이름 기반 책임자(workers.is_manager) 우선, 없으면 레거시 admin(users) 폴백.
        user = authenticate_manager_worker(username, password)
        non_manager = False
        denied_actor: dict[str, Any] | None = None
        if user is None:
            legacy = authenticate_user(username, password)
            if legacy is not None:
                if has_access_level(legacy, "manager"):
                    user = legacy
                else:
                    # 유효한 계정이지만 책임자 권한이 아님 → 403.
                    non_manager = True
                    denied_actor = legacy
        if user is None:
            write_audit_log(
                connection,
                action="blend_rescale_approve_denied",
                actor=denied_actor,
                target_type="rescale_approval",
                target_label=username or "(빈 이름)",
                details={
                    "reason": "not_manager" if non_manager else "invalid_credentials",
                    "purpose": purpose,
                },
            )
            connection.commit()
            raise HTTPException(
                status_code=403 if non_manager else 401,
                detail="FORBIDDEN" if non_manager else "INVALID_CREDENTIALS",
            )
        # 통과 — 승인 토큰 발급(approver=표시명).
        approver = user.get("display_name") or user.get("username") or "책임자"
        result = blend_service.create_rescale_approval(connection, approver, purpose)
        write_audit_log(
            connection,
            action="blend_manual_entry_approved" if is_manual else "blend_rescale_approved",
            actor=user,
            target_type="rescale_approval",
            target_id=str(result["approval_id"]),
            target_label=approver,
            details={"purpose": purpose} if is_manual else {},
        )
        connection.commit()
        return result

    @router.post("/blend/batch-discards")
    def blend_batch_discard(
        body: BlendBatchDiscardBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """배치 전체 폐기 기록 — 과중량 폐기 권장·3회 증량 차단 뒤 협의 폐기의 흔적.

        저장을 대신하는 별도 스트림: 제품 LOT 을 소비하지 않고 blend_records 를 만들지
        않는다(목록·집계·DHR 불변, ERP 이관 이중 차감 방지). 저장 없이 화면을 떠나면
        실물 소모 최대의 폐기가 무기록이던 구멍(2026-08-05)의 마개다. 사유 필수.
        """
        worker = require_blend_worker(request)
        current_user = get_current_user(request, required=False)
        actor = actor_name(current_user) if current_user else "현장"
        discard_id = blend_service.create_batch_discard(
            connection,
            recipe_id=body.recipe_id,
            product_name=body.product_name,
            worker=worker,
            work_date=body.work_date,
            total_amount=body.total_amount,
            reason=body.reason,
            source=body.source,
            details=[d.model_dump() for d in body.details],
            created_by=actor,
            created_at=utc_now_text(),
        )
        write_audit_log(
            connection,
            action="blend_batch_discarded",
            actor=current_user,
            target_type="blend_batch_discard",
            target_id=str(discard_id),
            target_label=body.product_name,
            details={
                "source": body.source,
                "reason": body.reason.strip()[:300],
                "materials": len(body.details),
                "discarded_g": round(sum(d.actual_amount for d in body.details), 2),
            },
        )
        connection.commit()
        return {"id": discard_id}

    @router.post("/blend/records")
    def blend_create(
        body: BlendCreateBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        # 멱등 재시도 먼저 — 저장이 커밋된 뒤 응답만 유실된 재전송은 그때 만든 기록을
        # 그대로 돌려준다(검증보다 앞에 둔다 — 이미 저장된 요청을 뒤늦은 검증으로
        # 400 내면 작업자가 또 저장을 눌러 진짜 중복이 생긴다).
        request_id = (body.request_id or "").strip()
        if request_id:
            prior = blend_service.lookup_save_request(connection, request_id, _SAVE_EP_SINGLE)
            if prior:
                existing = blend_service.get_blend_record(connection, prior[0])
                if existing:
                    _log_duplicate_save(request_id, prior, existing["product_lot"])
                    return _mask_manual_entry(request, existing)
        if not body.details:
            raise HTTPException(status_code=400, detail="배합 상세가 비어 있습니다.")
        # 반응기 진행 반제품은 실적 저장 시 반응기(1~4) 지정 필수.
        if blend_service.product_uses_reactor(connection, body.product_name) and body.reactor is None:
            raise HTTPException(status_code=400, detail="반응기를 선택하세요.")

        # 비율·이론량은 서버가 레시피에서 직접 산출한다 — 클라이언트 값은 쓰지 않는다(감사 F-5).
        # 레시피 없이 저장되는 경로(옛 데이터 이관·수동 입력)는 대조할 근거가 없어 그대로 둔다.
        details = [d.model_dump() for d in body.details]
        total_amount = body.total_amount
        # 반응기 이월(carry-over) 검증·강제 채움 — derive 보다 먼저. 이월 행의 actual_amount
        # 를 1차 배합 총량으로 덮어쓰므로, 그 뒤 derive 가 올바른 기준 실측값으로 이론·총량을
        # 산출하게 한다(잘못된 클라이언트 값이 편차·총량에 스며드는 것을 막는다).
        try:
            blend_service.enforce_carry_over(
                connection, body.recipe_id, body.product_name, details
            )
        except blend_service.CarryOverError as exc:
            raise HTTPException(status_code=400, detail=exc.detail) from exc
        if body.recipe_id:
            # 화면이 열려 있는 사이 개정됐으면 옛 배합비다 — 조용히 저장하지 않고 되돌린다.
            if blend_service.resolve_chain_tip(connection, body.recipe_id) != body.recipe_id:
                raise HTTPException(
                    status_code=409,
                    detail="레시피가 개정되었습니다. 화면을 새로고침한 뒤 다시 확인하세요.",
                )
            try:
                details, total_amount = blend_service.derive_details_from_recipe(
                    connection, body.recipe_id, body.total_amount, details
                )
            except blend_service.RecipeMismatchError as exc:
                raise HTTPException(status_code=400, detail=exc.detail) from exc

        # 전 자재 계량 완료 — 실제량이 빈 자재가 하나라도 있으면 저장 거부.
        # 편차 검사는 이 결손을 못 잡는다(actual is None 이면 건너뛴다). 그대로 저장되면
        # 자재 사용량 SUM 이 그 자재를 빼고 집계하고 DHR 에 빈 줄이 남는다.
        # enforce_carry_over 이후라 반응기 이월 행은 이미 1차 총량으로 채워져 있다.
        missing_actuals = blend_service.missing_actual_names(details)
        if missing_actuals:
            shown = missing_actuals[:5]
            suffix = " …" if len(missing_actuals) > 5 else ""
            raise HTTPException(
                status_code=400,
                detail="실제량이 입력되지 않은 자재가 있습니다: " + ", ".join(shown) + suffix,
            )

        # 자재 LOT 필수 — 추적성 핵심. enforce_carry_over·derive 이후 최종 행 상태로 검사.
        missing_lots = blend_service.missing_lot_names(details)
        if missing_lots:
            shown = missing_lots[:5]
            suffix = " …" if len(missing_lots) > 5 else ""
            raise HTTPException(
                status_code=400,
                detail="자재 LOT 를 입력하세요: " + ", ".join(shown) + suffix,
            )

        # 앞 단계 배합 기록에 없는 반제품 LOT — **저장을 막지 않는다**(2026-08-04).
        # 1차를 만들고 곧바로 2차에 투입하는 정당한 경우에도 매번 400 이 나서 작업자가
        # 사유란에 아무 글자나 치고 넘어갔다(통제의 형해화). 지금은 서버가 스스로 감지해
        # blend_lot_acks 에 대사용 기록만 남긴다 — 저장 성공 후(record_id 확보 뒤) 기록.
        lot_acks = blend_service.collect_lot_acks(
            connection, details, body.lot_overrides
        )

        # 자재별 허용 편차 검사 — 합계 편차는 제한 없음. 편차는 레시피에서 결정
        # (recipe_id 가 없으면 기본값 0.05g). 메시지는 실제 적용된 편차를 표시.
        tolerance = blend_service.recipe_tolerance_g(connection, body.recipe_id)
        offenders = blend_service.weighing_tolerance_violations(
            details, tolerance_g=tolerance
        )
        if offenders:
            raise HTTPException(
                status_code=400,
                detail=f"허용 편차(±{tolerance}g)를 초과한 자재: "
                + ", ".join(offenders),
            )
        # 증량(rescale) 이벤트 검증·승인 소비 — create 직전 확정(같은 트랜잭션).
        # 유효 승인 토큰(approval_id)은 used=1 로 소비되고, 부재 사유(absence_reason)는
        # 미확인으로 기록된다. 3회 이상·무효/재사용/만료 토큰은 400.
        try:
            rescale = blend_service.validate_rescale_events(connection, body.rescale_events)
        except blend_service.RescaleApprovalError as exc:
            raise HTTPException(status_code=400, detail=exc.detail) from exc
        worker = require_blend_worker(request)
        current_user = get_current_user(request, required=False)
        actor = actor_name(current_user) if current_user else "현장"
        record_id = blend_service.create_blend_record(
            connection,
            recipe_id=body.recipe_id,
            product_name=body.product_name,
            ink_name=body.ink_name,
            position=body.position,
            worker=worker,
            work_date=body.work_date,
            work_time=body.work_time,
            total_amount=total_amount,
            scale=body.scale,
            note=body.note,
            details=details,
            created_by=actor,
            created_at=utc_now_text(),
            worker_sign=body.worker_sign,
            reactor=body.reactor,
            manual_entry=body.manual_entry,
        )
        record = blend_service.get_blend_record(connection, record_id)
        # 증량 이벤트가 있으면 컬럼 기록 + 감사. 없으면(rescale=None) 기존 동작 유지(컬럼 기본값 0).
        if rescale is not None:
            blend_service.apply_rescale_to_record(connection, record_id, rescale)
            write_audit_log(
                connection,
                action="blend_rescale_saved",
                actor=current_user,
                target_type="blend_record",
                target_id=str(record_id),
                target_label=record["product_lot"],
                details={
                    "count": rescale["count"],
                    "unapproved": rescale["unapproved"],
                    "totals": rescale["totals"],
                },
            )
        # 계량 중 자재 폐기 — '처음부터 다시' 재계량에서 실제로 버린 자재의 흔적.
        # 저장을 막지 않는 순수 기록(편차 강제라 최종 수치엔 안 나타나는 소모를 남긴다).
        discard_json = blend_service.apply_discard_events_to_record(
            connection,
            record_id,
            [ev.model_dump() for ev in body.discard_events] if body.discard_events else None,
        )
        if discard_json is not None:
            write_audit_log(
                connection,
                action="blend_discard_saved",
                actor=current_user,
                target_type="blend_record",
                target_id=str(record_id),
                target_label=record["product_lot"],
                details={"events": discard_json[:1000]},
            )
        # 수기 입력 '책임자 부재 진행' — 사유를 기록에 남기고 미확인(ack 대기)으로 표시.
        # 증량 부재와 동일하게 책임자 확인 전까지 대시보드·트레이 알림에 남는다.
        if blend_service.apply_manual_absence_to_record(
            connection, record_id, body.manual_absence_reason
        ):
            write_audit_log(
                connection,
                action="blend_manual_absence_saved",
                actor=current_user,
                target_type="blend_record",
                target_id=str(record_id),
                target_label=record["product_lot"],
                details={"reason": (body.manual_absence_reason or "").strip()[:300]},
            )
        # 앞 단계 기록에 없는 LOT 진행 — 대사용 구조화 기록(사유가 비어도 반드시 남는다).
        blend_service.record_lot_acks(connection, record_id, lot_acks, utc_now_text())
        # 총량 플래그(2026-08-04) — 저장을 막지 않고 표시만 남긴다.
        #   oversize_total       현장 1회 상한(25,000 g) 초과 저장('그래도 증량' 경로)
        #   total_bypass_suspect 증량 이력 없이 총량만 키운 증량 승인 우회 의심
        total_flags = _flag_total_anomalies(
            connection,
            record_id=record_id,
            product_lot=record["product_lot"],
            total_amount=total_amount,
            recipe_id=body.recipe_id,
            rescale_count=(rescale["count"] if rescale else 0),
            current_user=current_user,
        )
        create_audit_details: dict[str, Any] = {
            "product_name": body.product_name,
            "total_amount": body.total_amount,
            "items": len(body.details),
            "manual_entry": body.manual_entry,
        }
        if total_flags["oversize_total"] or total_flags["total_bypass_suspect"]:
            create_audit_details["total_flags"] = total_flags
        # 감사에도 같은 항목을 구조화 보존한다(GAP-1 belt-and-braces). blend_lot_acks 가
        # 대사의 1차 소스이고, 감사는 삭제·수정에도 남는 원본 사본이다.
        if lot_acks:
            create_audit_details["lot_acks"] = lot_acks
        write_audit_log(
            connection,
            action="blend_record_create",
            actor=current_user,
            target_type="blend_record",
            target_id=str(record_id),
            target_label=record["product_lot"],
            details=create_audit_details,
        )
        # 멱등 기록 — 저장과 같은 트랜잭션으로 남긴다(400 으로 중단된 저장은 id 를
        # 소모하지 않는다). 동시 요청이 먼저 커밋했으면 IntegrityError → 이 저장을
        # 통째로 롤백하고 먼저 커밋된 결과를 돌려준다.
        if request_id:
            try:
                blend_service.remember_save_request(
                    connection, request_id, _SAVE_EP_SINGLE, [record_id]
                )
            except sqlite3.IntegrityError:
                connection.rollback()
                return _resolve_duplicate_save(
                    connection, request, request_id, _SAVE_EP_SINGLE
                )
        connection.commit()
        return _mask_manual_entry(request, record)

    @router.put(
        "/blend/records/{record_id}",
        dependencies=[Depends(require_access_level("manager"))],
    )
    def blend_update(
        record_id: int,
        body: BlendCreateBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """배합 실적 전체 수정 — 책임자 이상만(현장 무로그인은 401). 드문 정정용.

        product_lot·상태·생성정보·서명(담당/검토/승인)은 보존하고 헤더·상세만 교체.
        """
        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        # 제품명은 수정 불가 — product_lot 이 {제품명}{YYMMDD}{순번} 이라 제품명만 바꾸면
        # LOT 접두사가 옛 제품명으로 남아 기록과 어긋난다(감사 F-8). LOT 을 새로 채번하면
        # 이미 출력·보관된 DHR 문서와 어긋나므로(F-1 이 막으려던 바로 그 오염) 재채번도 안 한다.
        # 제품을 잘못 등록했으면 이 기록을 취소하고 새로 등록한다.
        if (body.product_name or "").strip() != (record.get("product_name") or "").strip():
            raise HTTPException(
                status_code=400,
                detail="제품명은 수정할 수 없습니다(제품 LOT 이 제품명으로 채번됩니다). "
                "잘못 등록했다면 이 기록을 취소하고 새로 등록하세요.",
            )
        if not body.details:
            raise HTTPException(status_code=400, detail="배합 상세가 비어 있습니다.")
        if blend_service.product_uses_reactor(connection, body.product_name) and body.reactor is None:
            raise HTTPException(status_code=400, detail="반응기를 선택하세요.")
        details = [d.model_dump() for d in body.details]
        total_amount = body.total_amount
        # 반응기 이월(carry-over) 검증·강제 채움 — create 경로와 대칭. carried_over=true 행은
        # 파생 레시피의 기준 자재 + 등록된 1차 LOT 이어야 하고, actual_amount 는 1차 총량으로
        # 강제 덮어써진다(변조 방지). 정정 저장에서도 이 불변식을 지킨다. derive 보다 먼저.
        try:
            blend_service.enforce_carry_over(
                connection, body.recipe_id, body.product_name, details
            )
        except blend_service.CarryOverError as exc:
            raise HTTPException(status_code=400, detail=exc.detail) from exc
        # 비율·이론량은 서버가 레시피에서 직접 재산출한다(F-5) — 클라이언트 값 불신. create 와
        # 대칭(GAP-2). recipe_id 가 없는 옛/수동 기록은 대조 근거가 없어 그대로 둔다.
        # ⚠ 수정은 생성과 달리 **그 기록이 만들어진 개정본**을 쓴다(resolve_revision=False).
        # 최신 개정판으로 귀결시키면 작업시간만 고쳐도 과거 기록의 배합비율·이론량이 새
        # 레시피 값으로 조용히 바뀌고(감사에는 "작업시간만 변경"으로 남는다), 개정 이후엔
        # 실측이 옛 비율 기준이라 편차 초과 400 으로 비고 오타 정정조차 막혔다.
        if body.recipe_id:
            try:
                details, total_amount = blend_service.derive_details_from_recipe(
                    connection, body.recipe_id, body.total_amount, details,
                    resolve_revision=False,
                )
            except blend_service.RecipeMismatchError as exc:
                raise HTTPException(status_code=400, detail=exc.detail) from exc
        # 전 자재 계량 완료 — 정정 저장이 채워져 있던 실제량을 비우지 못하게 한다.
        # 단, 이미 실제량이 비어 있던 옛 기록(이 통제가 없던 시절 저장분)은 그대로 둔다 —
        # 그런 기록의 비고 오타 정정까지 막으면 되돌릴 방법이 없어진다. 그래서 '새로
        # 생기는 결손'만 차단한다(기존 결손 자재명 집합과의 차집합).
        new_missing = sorted(
            set(blend_service.missing_actual_names(details))
            - set(blend_service.missing_actual_names(record.get("details") or []))
        )
        if new_missing:
            shown = new_missing[:5]
            suffix = " …" if len(new_missing) > 5 else ""
            raise HTTPException(
                status_code=400,
                detail="실제량을 비울 수 없습니다: " + ", ".join(shown) + suffix,
            )
        # 자재 LOT 필수 — 추적성 핵심(create 와 동일 통제). enforce_carry_over·derive 이후 검사.
        missing_lots = blend_service.missing_lot_names(details)
        if missing_lots:
            shown = missing_lots[:5]
            suffix = " …" if len(missing_lots) > 5 else ""
            raise HTTPException(
                status_code=400,
                detail="자재 LOT 를 입력하세요: " + ", ".join(shown) + suffix,
            )
        # 앞 단계 기록에 없는 반제품 LOT — create 와 동일하게 막지 않고 대사용 기록만.
        lot_acks = blend_service.collect_lot_acks(
            connection, details, body.lot_overrides
        )
        # 자재별 허용 편차 — 서버 재산출 상세 기준. 편차는 레시피(recipe_id)에서 결정, 없으면 0.05g.
        tolerance = blend_service.recipe_tolerance_g(connection, body.recipe_id)
        offenders = blend_service.weighing_tolerance_violations(
            details, tolerance_g=tolerance
        )
        if offenders:
            raise HTTPException(
                status_code=400,
                detail=f"허용 편차(±{tolerance}g)를 초과한 자재: "
                + ", ".join(offenders),
            )
        current_user = get_current_user(request, required=False)
        # 규제 보존(before-image): 변경을 적용하기 전에 현재 헤더(수정 가능 필드)와 상세 행을
        # 포착한다. audit_logs.details_json 이 그대로 담으므로 스키마 변경은 없다. 상세 행은
        # 용량을 아끼려 terse 배열([자재명, 실제량, LOT, 이월(0/1), 수동(0/1)])로 압축한다.
        _EDITABLE = (
            "worker", "work_date", "work_time", "total_amount",
            "scale", "note", "position", "ink_name", "reactor",
        )

        def _terse_rows(rec: dict[str, Any]) -> list[list[Any]]:
            # ratio·theory_amount 도 포함한다 — 이 둘은 DHR 에 인쇄되는 값인데 예전 terse
            # 행에 빠져 있어서, 서버 재산출로 값이 바뀌어도 before/after 가 동일해 보였고
            # 감사가 "작업시간만 변경"이라고 잘못 기록했다(무성 변경 + 허위 감사).
            return [
                [
                    d.get("material_name"),
                    d.get("actual_amount"),
                    d.get("material_lot"),
                    1 if d.get("carried_over") else 0,
                    1 if d.get("manual_entry") else 0,
                    d.get("ratio"),
                    d.get("theory_amount"),
                ]
                for d in (rec.get("details") or [])
            ]

        before_header = {k: record.get(k) for k in _EDITABLE}
        before_rows = _terse_rows(record)
        blend_service.update_blend_record(
            connection,
            record_id,
            product_name=body.product_name,
            ink_name=body.ink_name,
            position=body.position,
            worker=body.worker,
            work_date=body.work_date,
            work_time=body.work_time,
            total_amount=total_amount,
            scale=body.scale,
            note=body.note,
            details=details,
            reactor=body.reactor,
            updated_at=utc_now_text(),
        )
        updated = blend_service.get_blend_record(connection, record_id)
        # after_summary: 변경된 필드만 (신규 값). 상세 행이 바뀌었으면 새 terse 행을 함께 남긴다.
        after_summary: dict[str, Any] = {
            k: updated.get(k) for k in _EDITABLE if updated.get(k) != before_header[k]
        }
        after_rows = _terse_rows(updated)
        if after_rows != before_rows:
            after_summary["rows"] = after_rows
        # 대사용 기록도 수정 후 상태로 다시 쓴다(replace) — 수정으로 LOT 가 바뀌면
        # 옛 대사 대상이 유령으로 남는다.
        blend_service.record_lot_acks(
            connection, record_id, lot_acks, utc_now_text(), replace=True
        )
        # 총량 플래그도 정정 후 총량 기준으로 다시 판정한다 — 잘못 친 총량을 고치면
        # 플래그가 0 으로 되돌아가고, 정정으로 상한을 넘기면 새로 켜진다. 증량 이력은
        # 수정 경로에서 바뀌지 않으므로 저장돼 있던 rescale_count 를 그대로 쓴다.
        total_flags = _flag_total_anomalies(
            connection,
            record_id=record_id,
            product_lot=updated["product_lot"],
            total_amount=total_amount,
            recipe_id=body.recipe_id,
            rescale_count=int(record.get("rescale_count") or 0),
            current_user=current_user,
        )
        if total_flags["oversize_total"] or total_flags["total_bypass_suspect"]:
            after_summary["total_flags"] = total_flags
        if lot_acks:
            after_summary["lot_acks"] = lot_acks
        write_audit_log(
            connection,
            action="blend_record_update",
            actor=current_user,
            target_type="blend_record",
            target_id=str(record_id),
            target_label=updated["product_lot"],
            details={
                "before": {"header": before_header, "rows": before_rows},
                "after_summary": after_summary,
            },
        )
        connection.commit()
        updated["viscosity"] = viscosity_service.list_readings_for_blend(connection, record_id)
        return updated

    @router.post(
        "/blend/records/{record_id}/approve",
        dependencies=[Depends(require_access_level("manager"))],
    )
    def blend_approve(
        record_id: int,
        body: BlendApprovalBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        current_user = get_current_user(request, required=False)
        now = utc_now_text()
        col = "reviewed" if body.role == "review" else "approved"
        connection.execute(
            f"UPDATE blend_records SET {col}_by = ?, {col}_at = ?, {col}_sign = ?, updated_at = ? WHERE id = ?",
            (body.name.strip(), now, body.signature, now, record_id),
        )
        write_audit_log(
            connection,
            action=f"blend_record_{body.role}",
            actor=current_user,
            target_type="blend_record",
            target_id=str(record_id),
            target_label=record["product_lot"],
            details={"name": body.name},
        )
        connection.commit()
        result = blend_service.get_blend_record(connection, record_id)
        result["viscosity"] = viscosity_service.list_readings_for_blend(connection, record_id)
        return result

    @router.get("/blend/records/{record_id}/export")
    def blend_export(
        record_id: int,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        _audit_dhr_export(
            connection, request, fmt="xlsx", record_ids=[record_id],
            target_id=str(record_id), target_label=record["product_lot"],
        )
        xlsx_bytes = dhr_excel.build_official_dhr_xlsx(record)
        buf = io.BytesIO(xlsx_bytes)
        buf.seek(0)
        # 한글 product_lot 대응: ASCII 폴백 + RFC 5987 filename* (UTF-8)
        from urllib.parse import quote
        ascii_name = f"blend-{record_id}.xlsx"
        utf8_name = quote(f"원료배합일지-{record['product_lot']}.xlsx")
        disposition = (
            f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_name}"
        )
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": disposition},
        )

    @router.get("/blend/records/{record_id}/pdf")
    def blend_pdf(
        record_id: int,
        request: Request,
        sign: bool = Query(default=False),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> StreamingResponse:
        """배합일지 PDF. 기본은 서명 없이(빈 결재칸), sign=True 면 서명 합성."""
        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        _audit_dhr_export(
            connection, request, fmt="pdf_signed" if sign else "pdf",
            record_ids=[record_id], target_id=str(record_id),
            target_label=record["product_lot"],
        )
        if sign:
            # 서명본은 캐시하지 않음(기본 비서명본만 캐시)
            pdf_bytes = dhr_pdf.build_scanned_dhr_pdf(record, sign=True)
        else:
            # 캐시(레코드·서명설정 변경 시 자동 무효화) → 없으면 생성 후 저장
            pdf_bytes = dhr_cache.get(record)
            if pdf_bytes is None:
                pdf_bytes = dhr_pdf.build_scanned_dhr_pdf(record, sign=False)
                dhr_cache.put(record, pdf_bytes)
        from urllib.parse import quote
        utf8_name = quote(f"원료배합일지-{record['product_lot']}.pdf")
        disposition = (
            f"attachment; filename=\"blend-{record_id}.pdf\"; filename*=UTF-8''{utf8_name}"
        )
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": disposition},
        )

    @router.post("/blend/records/bulk")
    def blend_create_bulk(
        body: BlendBulkBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        if not body.entries:
            raise HTTPException(status_code=400, detail="생성할 항목이 없습니다.")
        worker = require_blend_worker(request)
        current_user = get_current_user(request, required=False)
        actor = actor_name(current_user) if current_user else "현장"
        now = utc_now_text()
        try:
            ids = blend_service.create_bulk(
                connection,
                recipe_id=body.recipe_id,
                worker=worker,
                scale=body.scale,
                entries=[e.model_dump() for e in body.entries],
                created_by=actor,
                created_at=now,
            )
        except ValueError as exc:
            raise HTTPException(status_code=404, detail=str(exc))
        lots = [blend_service.get_blend_record(connection, rid)["product_lot"] for rid in ids]
        write_audit_log(
            connection,
            action="blend_record_bulk_create",
            actor=current_user,
            target_type="blend_record",
            target_id=",".join(str(i) for i in ids),
            target_label=f"{len(ids)}건",
            details={"recipe_id": body.recipe_id, "count": len(ids)},
        )
        connection.commit()
        return {"created": len(ids), "ids": ids, "product_lots": lots}

    @router.post("/blend/records/continuous")
    def blend_create_continuous(
        body: BlendContinuousBody,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """이어서 계량: 한 레시피 · 동일 총량으로 N개 로트를 한 번에 저장.

        각 로트를 기존 단건 저장과 동일하게 서버 도출·편차검사한 뒤, 모두 통과하면 순차
        저장한다(하나라도 실패하면 아무것도 저장하지 않음). product_lot 은 로트마다 연속 채번.
        """
        # 멱등 재시도 먼저(단건과 동일) — 이미 저장된 회차의 재전송은 그때 만든 N로트를
        # 그대로 돌려준다. 재시도가 로트 N건을 두 벌 만드는 것을 막는다.
        request_id = (body.request_id or "").strip()
        if request_id:
            prior = blend_service.lookup_save_request(connection, request_id, _SAVE_EP_CONTINUOUS)
            if prior:
                result = _continuous_result(connection, prior)
                _log_duplicate_save(request_id, prior, f"{len(prior)}건")
                return result
        if not body.lots:
            raise HTTPException(status_code=400, detail="저장할 로트가 없습니다.")
        if any(not lot for lot in body.lots):
            raise HTTPException(status_code=400, detail="자재 상세가 비어 있는 로트가 있습니다.")
        # 반응기 이월(carry-over)은 단일 배합 화면 전용 — 연속(다중 로트) 화면에서는 거부.
        if any(d.carried_over for lot in body.lots for d in lot):
            raise HTTPException(
                status_code=400,
                detail="반응기 이월은 단일 배합 화면에서만 사용할 수 있습니다.",
            )
        # 반응기 진행 반제품은 반응기(1~4) 지정 필수 (전 로트 공통).
        if blend_service.product_uses_reactor(connection, body.product_name) and body.reactor is None:
            raise HTTPException(status_code=400, detail="반응기를 선택하세요.")
        # 화면이 열린 사이 개정됐으면 옛 배합비 — 저장 거부(감사 F-5 / 단건과 동일).
        if blend_service.resolve_chain_tip(connection, body.recipe_id) != body.recipe_id:
            raise HTTPException(
                status_code=409,
                detail="레시피가 개정되었습니다. 화면을 새로고침한 뒤 다시 확인하세요.",
            )
        tolerance = blend_service.recipe_tolerance_g(connection, body.recipe_id)
        # 저장 전 전 로트 도출·편차검사 (원자성: 하나라도 실패하면 중단, 저장 없음)
        # lot_totals 가 있으면 그 로트의 총량 오버라이드를 사용(초과 계량 증량).
        derived_lots: list[list[dict[str, Any]]] = []
        # 로트별 '앞 단계 기록에 없는 LOT' 대사 항목 — derived_lots 와 평행(인덱스 j).
        lot_acks_per_lot: list[list[dict[str, Any]]] = []
        lot_totals = body.lot_totals or []
        for lot_no, lot in enumerate(body.lots, start=1):
            lot_total = lot_totals[lot_no - 1] if lot_totals and lot_totals[lot_no - 1] else body.total_amount
            details = [d.model_dump() for d in lot]
            try:
                derived, _total = blend_service.derive_details_from_recipe(
                    connection, body.recipe_id, lot_total, details
                )
            except blend_service.RecipeMismatchError as exc:
                raise HTTPException(status_code=400, detail=f"로트 {lot_no}: {exc.detail}") from exc
            # 전 자재 계량 완료(단건과 동일 규칙) — 실제량이 빈 셀이 있으면 400.
            # 화면(blend_continuous.js)에도 같은 검사가 있으나 서버가 최종 방어선이다.
            missing_actuals = blend_service.missing_actual_names(derived)
            if missing_actuals:
                shown = missing_actuals[:5]
                suffix = " …" if len(missing_actuals) > 5 else ""
                raise HTTPException(
                    status_code=400,
                    detail=f"로트 {lot_no}: 실제량이 입력되지 않은 자재가 있습니다: "
                    + ", ".join(shown) + suffix,
                )
            # 자재 LOT 필수 — 추적성 핵심(단건과 동일 규칙).
            missing_lots = blend_service.missing_lot_names(derived)
            if missing_lots:
                shown = missing_lots[:5]
                suffix = " …" if len(missing_lots) > 5 else ""
                raise HTTPException(
                    status_code=400,
                    detail=f"로트 {lot_no}: 자재 LOT 를 입력하세요: " + ", ".join(shown) + suffix,
                )
            # 앞 단계 기록에 없는 반제품 LOT — 단건과 동일하게 막지 않고 대사용 기록만.
            # 로트별로 모아 두었다가 저장 후 각 record_id 에 남긴다.
            lot_acks_per_lot.append(
                blend_service.collect_lot_acks(connection, derived, body.lot_overrides)
            )
            offenders = blend_service.weighing_tolerance_violations(derived, tolerance_g=tolerance)
            if offenders:
                raise HTTPException(
                    status_code=400,
                    detail=f"로트 {lot_no}: 허용 편차(±{tolerance}g) 초과 — " + ", ".join(offenders),
                )
            derived_lots.append(derived)
        # 증량(rescale) 이벤트 검증·승인 소비 — 로트별(lot_rescale_events[j]). 단건 blend_create
        # 와 동일한 validate_rescale_events 를 로트마다 호출한다. 유효 approval_id 는 used=1 로
        # 소비되고, absence_reason 은 그 로트만 미확인(rescale_unacked=1)으로 기록된다. 무효·재사용·
        # 만료·3회 초과는 400. 커밋은 맨 끝 한 번뿐이라, 여기서 400 이 나면 앞 로트에서 소비한
        # 승인 UPDATE 도 함께 롤백된다(get_db 가 미커밋 연결을 close → 자동 롤백 → 원자성).
        lot_rescale_events = body.lot_rescale_events or []
        lot_rescales: list[dict[str, Any] | None] = []
        try:
            for lot_no in range(1, len(body.lots) + 1):
                events = (
                    lot_rescale_events[lot_no - 1]
                    if lot_no - 1 < len(lot_rescale_events)
                    else None
                )
                lot_rescales.append(
                    blend_service.validate_rescale_events(connection, events)
                )
        except blend_service.RescaleApprovalError as exc:
            raise HTTPException(status_code=400, detail=f"로트 {lot_no}: {exc.detail}") from exc
        worker = require_blend_worker(request)
        current_user = get_current_user(request, required=False)
        actor = actor_name(current_user) if current_user else "현장"
        ids = blend_service.create_continuous(
            connection,
            recipe_id=body.recipe_id,
            product_name=body.product_name,
            ink_name=body.ink_name,
            position=body.position,
            worker=worker,
            work_date=body.work_date,
            work_time=body.work_time,
            total_amount=body.total_amount,
            scale=body.scale,
            note=body.note,
            lots_details=derived_lots,
            created_by=actor,
            created_at=utc_now_text(),
            worker_sign=body.worker_sign,
            reactor=body.reactor,
            lot_totals=body.lot_totals,
        )
        # 증량 이벤트가 있는 로트만 그 로트의 record 에 컬럼 기록 + 감사(단건과 동일 규칙).
        # lot_rescales[j] 는 create_continuous 가 저장한 ids[j] 와 같은 로트를 가리킨다(둘 다
        # body.lots 순서). 이벤트 없는 로트(None)는 건너뛰어 컬럼 기본값 0 을 유지한다.
        for lot_idx, rescale in enumerate(lot_rescales):
            if rescale is None:
                continue
            record_id = ids[lot_idx]
            blend_service.apply_rescale_to_record(connection, record_id, rescale)
            record = blend_service.get_blend_record(connection, record_id)
            write_audit_log(
                connection,
                action="blend_rescale_saved",
                actor=current_user,
                target_type="blend_record",
                target_id=str(record_id),
                target_label=record["product_lot"],
                details={
                    "count": rescale["count"],
                    "unapproved": rescale["unapproved"],
                    "totals": rescale["totals"],
                },
            )
        # 수기 입력 '책임자 부재 진행' — 화면 단위 승인이라 이 회차의 전 로트에 동일 적용.
        if (body.manual_absence_reason or "").strip():
            for record_id in ids:
                blend_service.apply_manual_absence_to_record(
                    connection, record_id, body.manual_absence_reason
                )
            write_audit_log(
                connection,
                action="blend_manual_absence_saved",
                actor=current_user,
                target_type="blend_record",
                target_id=",".join(str(i) for i in ids),
                target_label=f"{body.product_name} ({len(ids)}로트)",
                details={
                    "reason": (body.manual_absence_reason or "").strip()[:300],
                    "records": len(ids),
                },
            )
        # 앞 단계 기록에 없는 LOT 진행 — 로트별 record_id 에 대사용 기록(사유가 비어도 남는다).
        saved_at = utc_now_text()
        all_lot_acks: list[dict[str, Any]] = []
        for lot_idx, acks in enumerate(lot_acks_per_lot):
            if not acks or lot_idx >= len(ids):
                continue
            blend_service.record_lot_acks(connection, ids[lot_idx], acks, saved_at)
            all_lot_acks.extend(acks)
        lots = [blend_service.get_blend_record(connection, rid)["product_lot"] for rid in ids]
        # 총량 플래그 — 로트마다 판정한다(로트별 총량 오버라이드 lot_totals 때문에
        # 한 회차 안에서도 로트마다 총량이 다를 수 있다). 단건과 동일하게 표시만 남긴다.
        flagged_lots: list[dict[str, Any]] = []
        for lot_idx, rid in enumerate(ids):
            lot_total = body.total_amount
            if body.lot_totals and lot_idx < len(body.lot_totals) and body.lot_totals[lot_idx]:
                lot_total = body.lot_totals[lot_idx]
            lot_rescale = lot_rescales[lot_idx] if lot_idx < len(lot_rescales) else None
            lot_flags = _flag_total_anomalies(
                connection,
                record_id=rid,
                product_lot=lots[lot_idx],
                total_amount=lot_total,
                recipe_id=body.recipe_id,
                rescale_count=(lot_rescale["count"] if lot_rescale else 0),
                current_user=current_user,
            )
            if lot_flags["oversize_total"] or lot_flags["total_bypass_suspect"]:
                flagged_lots.append({"product_lot": lots[lot_idx], **lot_flags})
        continuous_audit_details: dict[str, Any] = {
            "recipe_id": body.recipe_id, "count": len(ids), "total_amount": body.total_amount,
        }
        if flagged_lots:
            continuous_audit_details["total_flags"] = flagged_lots
        # 감사에도 같은 항목을 구조화 보존(GAP-1 belt-and-braces).
        if all_lot_acks:
            continuous_audit_details["lot_acks"] = all_lot_acks
        write_audit_log(
            connection,
            action="blend_record_continuous_create",
            actor=current_user,
            target_type="blend_record",
            target_id=",".join(str(i) for i in ids),
            target_label=f"{len(ids)}건",
            details=continuous_audit_details,
        )
        # 멱등 기록 — 단건과 동일(저장과 같은 트랜잭션, 경합 시 롤백 후 먼저 커밋된 결과).
        if request_id:
            try:
                blend_service.remember_save_request(
                    connection, request_id, _SAVE_EP_CONTINUOUS, ids
                )
            except sqlite3.IntegrityError:
                connection.rollback()
                return _resolve_duplicate_save(
                    connection, request, request_id, _SAVE_EP_CONTINUOUS
                )
        connection.commit()
        return {"created": len(ids), "ids": ids, "product_lots": lots}

    @router.delete("/blend/records/{record_id}")
    def blend_cancel(
        record_id: int,
        request: Request,
        hard: bool = Query(default=False),
        reason: str | None = Query(default=None, max_length=500),
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        # 감사 F-2: soft/hard 모두 책임자 전용. soft 취소도 DHR 기록을 목록·출력·
        # 대시보드에서 숨기는 행위다(종전에는 soft 가 무인증이었다).
        # 인증을 404 조회보다 먼저 — 비인증 호출자에게 기록 존재 여부를 흘리지 않는다.
        # 상태 코드는 기존 hard 분기 관례를 보존: 미로그인·비책임자 모두 403.
        # (get_current_user(required=True)의 401 로 바꾸면 기존
        #  test_blend_hard_delete_requires_manager 의 403 기대가 깨진다.)
        current_user = get_current_user(request, required=False)
        if current_user is None or not has_access_level(current_user, "manager"):
            raise HTTPException(status_code=403, detail="FORBIDDEN")
        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        if hard:
            # 물리 삭제는 되돌릴 수 없다 — 규제 기록을 흔적 없이 없애는 가장 강한 행위이므로
            # ①사유를 반드시 받고 ②기록 전체 스냅샷을 감사에 남긴다. 예전에는 수정 경로에만
            # before-image 가 있고 삭제에는 LOT 한 줄뿐이라, 가장 위험한 행위가 가장 약한
            # 증적을 남겼다(백업 롤백 외 복구 수단 없음).
            clean_reason = (reason or "").strip()
            if not clean_reason:
                raise HTTPException(
                    status_code=400,
                    detail="완전 삭제는 사유가 필요합니다. 되돌릴 수 없으니 취소를 먼저 검토하세요.",
                )
            snapshot = {
                "header": {
                    k: record.get(k)
                    for k in (
                        "id", "product_lot", "recipe_id", "product_name", "ink_name",
                        "position", "worker", "work_date", "work_time", "total_amount",
                        "scale", "status", "note", "reactor", "manual_entry",
                        "is_bulk_regenerated", "manual_absence_reason",
                        "reviewed_by", "reviewed_at", "approved_by", "approved_at",
                        "created_by", "created_at", "updated_at",
                    )
                },
                "rows": [
                    [
                        d.get("material_name"), d.get("material_code"), d.get("material_lot"),
                        d.get("ratio"), d.get("theory_amount"), d.get("actual_amount"),
                        1 if d.get("carried_over") else 0,
                        1 if d.get("manual_entry") else 0,
                    ]
                    for d in (record.get("details") or [])
                ],
            }
            result = record_delete_service.delete_blend_record(connection, record_id)
            if result is None:
                raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
            write_audit_log(
                connection,
                action="blend_record_deleted",
                actor=current_user,
                target_type="blend_record",
                target_id=str(result.record_id),
                target_label=result.product_lot,
                details={"reason": clean_reason, "snapshot": snapshot},
            )
            connection.commit()
            return {"deleted": result.record_id}

        connection.execute(
            "UPDATE blend_records SET status = 'canceled', updated_at = ? WHERE id = ?",
            (utc_now_text(), record_id),
        )
        write_audit_log(
            connection,
            action="blend_record_cancel",
            actor=current_user,
            target_type="blend_record",
            target_id=str(record_id),
            target_label=record["product_lot"],
            details={"reason": reason} if reason else None,
        )
        connection.commit()
        return {"canceled": record_id}

    @router.post(
        "/blend/records/{record_id}/restore",
        dependencies=[Depends(require_access_level("manager"))],
    )
    def blend_restore(
        record_id: int,
        request: Request,
        connection: sqlite3.Connection = Depends(get_db),
    ) -> dict[str, Any]:
        """소프트 취소된 배합 기록 복원 (책임자 전용, 감사 F-2)."""
        record = blend_service.get_blend_record(connection, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="배합 기록을 찾을 수 없습니다.")
        if record["status"] != "canceled":
            raise HTTPException(status_code=400, detail="취소 상태의 기록이 아닙니다.")
        current_user = get_current_user(request, required=False)
        connection.execute(
            "UPDATE blend_records SET status = 'completed', updated_at = ? WHERE id = ?",
            (utc_now_text(), record_id),
        )
        write_audit_log(
            connection,
            action="blend_record_restore",
            actor=current_user,
            target_type="blend_record",
            target_id=str(record_id),
            target_label=record["product_lot"],
        )
        connection.commit()
        return {"restored": record_id}

    return router
