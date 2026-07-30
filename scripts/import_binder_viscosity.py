"""바인더 점도 기록 → 배합 실적(DHR)에 점도 연계 (멱등).

「바인더 기록 (자동 저장됨).xlsx」의 연도별 시트(24/25/26년도)를 읽어, 각 바인더
점도를 **우리 배합 기록에 연계**한다 — 배합 화면에서 점도를 등록하는 것과 완전히
같은 형태(lot_no = 배합 product_lot, blend_record_id 연계)로. 그래야 그 배합
실적서(DHR)에서 점도가 보이고, 점도 관리의 그 반제품 추세에 함께 잡힌다.

매칭 규칙(사용자 확정 2026-07-29):
    배합 기록 LOT = 바인더종류(정규화) + 사용한PB(8자리)
    예: APB + 26060101 = APB26060101  → 그 product_lot 의 배합 기록을 찾아 연계.
배합 기록이 없으면(구 시스템의 24/25년 등) 배합 미연계로 점도만 보존한다 — 같은
바인더 반제품에 쌓여 과거~현재가 한 추세로 이어진다(측정일은 엑셀 일자, 사용한PB 유지).

사용한PB 는 material_lot 에도 저장해, 그 PB 의 점도(PB 반제품)와 상관 조회에 쓴다
(analyze_product 의 source_pb 연계).

바인더 종류 정규화: APB(17)→APB17 · CSBP→CSPB · 괄호숫자 제거 · PM/PM17/HSPU 유지 ·
APB(TEST)/점도 결측 제외.

⚠ 이전 버전(별도 반제품 + lot_no=사용한PB + 배합 미연계)으로 넣은 흔적은 임포트
전에 자동 정리한다(--no-clean 으로 끌 수 있음).

운영 DB 적재:
    .venv\\Scripts\\python.exe scripts\\import_binder_viscosity.py "바인더 기록 (자동 저장됨).xlsx"
"""

import re
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from src.db import get_connection, init_db, utc_now_text  # noqa: E402
from src.services import viscosity_service  # noqa: E402

DEFAULT_FILE = "바인더 기록 (자동 저장됨).xlsx"
COL_DATE, COL_BINDER, COL_PB, COL_VISC, COL_WORKER = 0, 1, 2, 3, 4
_YEAR_FROM_SHEET = re.compile(r"(\d{2,4})\s*년")
_DATE_KO_YMD = re.compile(r"(?:(\d{2,4})\s*년)?\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일")


def _normalize_binder(raw: str) -> str | None:
    s = str(raw).strip()
    if not s or "TEST" in s.upper():
        return None
    if s == "CSBP":
        return "CSPB"
    m = re.fullmatch(r"([A-Za-z]+)\((\d+)\)", s)
    if m:
        head, num = m.group(1), m.group(2)
        return f"{head}{num}" if num == "17" else head
    s = re.sub(r"\(.*?\)", "", s).strip()
    return s or None


def _sheet_year(sheet_name: str) -> int | None:
    m = _YEAR_FROM_SHEET.search(sheet_name)
    if not m:
        return None
    y = int(m.group(1))
    return 2000 + y if y < 100 else y


def _parse_date(value, sheet_year: int | None) -> str | None:
    """일자 셀 → ISO(YYYY-MM-DD). 배합 미연계(과거) 점도의 측정일에 쓴다."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    m = _DATE_KO_YMD.search(s)
    if m:
        yr, mo, dy = m.group(1), int(m.group(2)), int(m.group(3))
        if yr:
            year = int(yr)
            year = 2000 + year if year < 100 else year
        elif sheet_year:
            year = sheet_year
        else:
            return None
        try:
            return date(year, mo, dy).isoformat()
        except ValueError:
            return None
    return None


def _pb_lot(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    return s if re.fullmatch(r"\d{8}", s) else None


def _visc(value) -> float | None:
    return float(value) if isinstance(value, (int, float)) else None


def _find_blend_record(connection, product_lot: str):
    """product_lot 로 배합 기록 1건 조회. 취소분은 제외, 여럿이면 최신."""
    try:
        return connection.execute(
            "SELECT id, product_name, product_lot, work_date FROM blend_records "
            "WHERE product_lot = ? AND status != 'canceled' ORDER BY id DESC LIMIT 1",
            (product_lot,),
        ).fetchone()
    except sqlite3.OperationalError:
        return None


def _clean_previous_import(connection) -> int:
    """이전 버전 임포트 흔적 제거 — 배합 미연계(blend_record_id IS NULL)이면서
    lot_no 가 8자리 숫자이고 material_lot 과 같은 행(그때의 특징). 배합 화면으로
    등록한 정상 점도(blend_record_id 있음, lot_no=product_lot)는 건드리지 않는다.
    """
    cur = connection.execute(
        "DELETE FROM viscosity_readings "
        "WHERE blend_record_id IS NULL AND lot_no = material_lot "
        "AND lot_no GLOB '[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]'"
    )
    return cur.rowcount


def import_binder(paths: list[str], *, clean: bool = True) -> dict:
    init_db()
    now = utc_now_text()
    stats = {
        "read": 0, "linked": 0, "archived": 0, "dup": 0, "cleaned": 0,
        "skip_binder": 0, "skip_pb": 0, "skip_visc": 0,
        "by_product": {}, "no_record_samples": [],
    }
    with get_connection() as connection:
        if clean:
            stats["cleaned"] = _clean_previous_import(connection)

        for path in paths:
            p = Path(path)
            if not p.exists():
                print(f"[건너뜀] 파일 없음: {p}")
                continue
            wb = load_workbook(p, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                if "바인더" not in sheet_name:
                    continue
                ws = wb[sheet_name]
                sheet_year = _sheet_year(sheet_name)
                for row in ws.iter_rows(values_only=True):
                    if not row or len(row) <= COL_VISC:
                        continue
                    binder_raw = row[COL_BINDER]
                    if not isinstance(binder_raw, str) or binder_raw.strip() in ("", "바인더"):
                        continue
                    stats["read"] += 1
                    binder = _normalize_binder(binder_raw)
                    if not binder:
                        stats["skip_binder"] += 1
                        continue
                    visc = _visc(row[COL_VISC])
                    if visc is None:
                        stats["skip_visc"] += 1
                        continue
                    pb = _pb_lot(row[COL_PB])
                    if not pb:
                        stats["skip_pb"] += 1
                        continue

                    # 매칭: 배합 기록 LOT = 바인더종류 + 사용한PB
                    product_lot = f"{binder}{pb}"
                    worker = None
                    if len(row) > COL_WORKER and isinstance(row[COL_WORKER], str):
                        worker = row[COL_WORKER].strip() or None

                    rec = _find_blend_record(connection, product_lot)
                    if rec is not None:
                        # 배합 화면 점도 등록과 동일 형태 — 배합 제품명으로 반제품 확보,
                        # lot_no = product_lot, blend_record_id 연계, material_lot = 사용한PB.
                        product = viscosity_service.ensure_product_by_code(
                            connection, rec["product_name"], rec["product_name"], now
                        )
                        if not product:
                            stats["skip_no_record"] += 1
                            continue
                        lot_no = rec["product_lot"]
                        measured_date = rec["work_date"]
                        blend_record_id = int(rec["id"])
                        recipe_material = rec["product_name"]
                        bucket = "linked"
                    else:
                        # 배합 기록 없음(구 시스템 등) — 배합 미연계로 점도만 보존한다.
                        # 반제품은 바인더종류, lot_no=바인더+사용한PB(연계분과 겹치지 않음),
                        # 측정일은 엑셀 일자, PB는 material_lot 유지(PB 점도 상관 조회).
                        if len(stats["no_record_samples"]) < 10:
                            stats["no_record_samples"].append(product_lot)
                        product = viscosity_service.ensure_product_by_code(
                            connection, binder, binder, now
                        )
                        if not product:
                            stats["skip_binder"] += 1
                            continue
                        lot_no = product_lot
                        measured_date = _parse_date(row[COL_DATE], sheet_year)
                        blend_record_id = None
                        recipe_material = binder
                        bucket = "archived"

                    try:
                        viscosity_service.add_reading(
                            connection,
                            product_id=product["id"],
                            lot_no=lot_no,
                            viscosity=visc,
                            measured_date=measured_date,
                            memo=None,
                            recipe_material=recipe_material,
                            material_lot=pb,
                            created_by=worker,
                            created_at=now,
                            blend_record_id=blend_record_id,
                        )
                        stats[bucket] += 1
                        stats["by_product"][binder] = stats["by_product"].get(binder, 0) + 1
                    except sqlite3.IntegrityError:
                        stats["dup"] += 1
            wb.close()
        connection.commit()
    return stats


def main() -> int:
    args = [a for a in sys.argv[1:] if a != "--no-clean"]
    clean = "--no-clean" not in sys.argv
    paths = args or [DEFAULT_FILE]
    s = import_binder(paths, clean=clean)
    print("\n=== 바인더 점도 → 배합 기록 연계 결과 ===")
    if clean:
        print(f"이전 임포트 정리 : {s['cleaned']}건 삭제")
    print(f"읽은 데이터행    : {s['read']}")
    print(f"배합 기록에 연계 : {s['linked']}")
    print(f"과거(배합 없음) 보존: {s['archived']}")
    print(f"중복(멱등)       : {s['dup']}")
    print(f"기타 제외 — 바인더 {s['skip_binder']} · 점도결측 {s['skip_visc']} · PB형식 {s['skip_pb']}")
    print(f"제품별 등록      : {s['by_product']}")
    if s["no_record_samples"]:
        print(f"배합 없어 과거로 보존한 LOT 표본: {s['no_record_samples']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
