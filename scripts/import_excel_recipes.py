"""Bulk importer for legacy Excel recipe workbooks.

Usage:
    python scripts/import_excel_recipes.py <xlsx_path> [--dry-run] [--created-by NAME]

Reads an xlsx file with the IRMS field-standard layout
(`제품명 | 위치 | 잉크명 | <재료...> | 비고`) across one or more sections in a
sheet, normalises material names, parses mixed-content cells via
`cell_value_parser`, and writes recipes + recipe_items. In dry-run mode nothing
is committed; a report is printed instead.

The importer preserves Excel formulas verbatim in `value_text` (formulas start
with '=' and are rendered by the existing spreadsheet engine on read).
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from src.database import get_connection, init_db, utc_now_text, write_audit_log  # noqa: E402
from src.services.cell_value_parser import parse_cell  # noqa: E402
from src.services.material_resolver import (  # noqa: E402
    normalize_material_name,
    resolve_material,
)


REQUIRED_HEADERS = {"제품명", "위치", "잉크명"}
REMARK_HEADERS = {"비고", "REMARK", "NOTE"}


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header_row(ws, max_scan: int = 40) -> list[tuple[int, list[str]]]:
    """Return list of (row_index, header_cells) for each section detected."""
    sections: list[tuple[int, list[str]]] = []
    for row_idx in range(1, min(ws.max_row, 200) + 1):
        row_values = [_norm(c.value) for c in ws[row_idx]]
        if not any(row_values):
            continue
        upper_set = {v for v in row_values if v}
        if REQUIRED_HEADERS.issubset(upper_set):
            sections.append((row_idx, row_values))
    return sections


def _section_data_rows(ws, header_row: int, next_header_row: int | None) -> list[list[str]]:
    end = next_header_row - 1 if next_header_row else ws.max_row
    rows = []
    for r in range(header_row + 1, end + 1):
        values = [_norm(c.value) for c in ws[r]]
        if not any(values):
            continue
        rows.append(values)
    return rows


def _build_column_map(header_row: list[str]) -> dict[str, Any]:
    idx = {name: -1 for name in ("product_name", "position", "ink_name", "remark")}
    material_cols: list[tuple[int, str]] = []
    for i, val in enumerate(header_row):
        v = val.strip()
        if not v:
            continue
        if v == "제품명":
            idx["product_name"] = i
        elif v == "위치":
            idx["position"] = i
        elif v == "잉크명":
            idx["ink_name"] = i
        elif v.upper() in REMARK_HEADERS:
            idx["remark"] = i
        elif i > 0 and idx["ink_name"] != -1 and i > idx["ink_name"]:
            material_cols.append((i, v))
    return {"cols": idx, "materials": material_cols}


def _resolve_materials_all(connection, material_cols: list[tuple[int, str]]) -> tuple[dict[int, int], list[str]]:
    resolved: dict[int, int] = {}
    missing: list[str] = []
    for idx, raw_name in material_cols:
        mid = resolve_material(connection, raw_name)
        if mid is None:
            missing.append(raw_name)
        else:
            resolved[idx] = mid
    return resolved, missing


def _parse_section(
    section_rows: list[list[str]],
    col_map: dict[str, Any],
    material_id_map: dict[int, int],
) -> list[dict[str, Any]]:
    cols = col_map["cols"]
    recipes: list[dict[str, Any]] = []
    last_product = ""

    for row in section_rows:
        def get(key: str) -> str:
            i = cols[key]
            return row[i].strip() if 0 <= i < len(row) else ""

        product = get("product_name") or last_product
        if get("product_name"):
            last_product = get("product_name")
        position = get("position")
        ink = get("ink_name")
        remark = get("remark") or None

        if not product or not ink:
            continue

        items = []
        for mat_idx, mat_name in col_map["materials"]:
            if mat_idx not in material_id_map:
                continue
            raw = row[mat_idx] if mat_idx < len(row) else ""
            weight, text = parse_cell(raw)
            if weight is None and text is None:
                continue
            items.append({
                "material_id": material_id_map[mat_idx],
                "material_name": mat_name,
                "value_weight": weight,
                "value_text": text,
            })

        recipes.append({
            "product_name": product,
            "position": position or "-",
            "ink_name": ink,
            "remark": remark,
            "items": items,
        })
    return recipes


def import_workbook(path: Path, dry_run: bool, created_by: str) -> int:
    init_db()
    wb = load_workbook(path, data_only=False)
    print(f"\n=== Import: {path.name} ===")
    print(f"Sheets: {wb.sheetnames}")

    total_recipes = 0
    total_items = 0
    all_missing_materials: set[str] = set()
    all_parsed: list[dict[str, Any]] = []
    warnings: list[str] = []

    with get_connection() as connection:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sections = _find_header_row(ws)
            if not sections:
                print(f"  [SKIP] '{sheet_name}': no header row found")
                continue

            for section_idx, (header_row, header_cells) in enumerate(sections):
                next_header = (
                    sections[section_idx + 1][0] if section_idx + 1 < len(sections) else None
                )
                col_map = _build_column_map(header_cells)
                material_id_map, missing = _resolve_materials_all(connection, col_map["materials"])
                all_missing_materials.update(missing)
                if missing:
                    warnings.append(
                        f"  [WARN] '{sheet_name}' @row{header_row}: unknown materials: {missing}"
                    )
                data_rows = _section_data_rows(ws, header_row, next_header)
                parsed = _parse_section(data_rows, col_map, material_id_map)
                for r in parsed:
                    item_weights_over_limit = [
                        it for it in r["items"]
                        if it["value_text"] and it["value_text"].startswith("=") and len(it["value_text"]) > 200
                    ]
                    if item_weights_over_limit:
                        warnings.append(
                            f"  [WARN] {r['product_name']}/{r['position']}: formula >200 chars"
                        )
                all_parsed.extend(parsed)
                total_recipes += len(parsed)
                total_items += sum(len(r["items"]) for r in parsed)
                print(f"  [OK] '{sheet_name}' section@row{header_row}: {len(parsed)} recipes")

        if all_missing_materials:
            print("\n=== Missing materials (resolve before import) ===")
            for m in sorted(all_missing_materials):
                print(f"  - {m}   (normalized: {normalize_material_name(m)})")
            print("\nABORT: cannot import with unknown materials.")
            return 2

        for w in warnings:
            print(w)

        print(f"\n=== Summary ===")
        print(f"  Recipes: {total_recipes}")
        print(f"  Items:   {total_items}")
        print(f"  Mode:    {'DRY-RUN' if dry_run else 'COMMIT'}")

        if dry_run:
            print("\nDry-run complete. No changes committed.")
            return 0

        now = utc_now_text()
        created_ids: list[int] = []
        for rec in all_parsed:
            cursor = connection.execute(
                """
                INSERT INTO recipes (
                    product_name, position, ink_name, status, created_by, created_at,
                    completed_at, remark
                ) VALUES (?, ?, ?, 'pending', ?, ?, NULL, ?)
                """,
                (
                    rec["product_name"],
                    rec["position"],
                    rec["ink_name"],
                    created_by,
                    now,
                    rec["remark"],
                ),
            )
            rid = cursor.lastrowid
            created_ids.append(rid)
            for item in rec["items"]:
                connection.execute(
                    """
                    INSERT INTO recipe_items (recipe_id, material_id, value_weight, value_text)
                    VALUES (?, ?, ?, ?)
                    """,
                    (rid, item["material_id"], item["value_weight"], item["value_text"]),
                )

        write_audit_log(
            connection,
            action="recipes_imported",
            actor={"id": 0, "username": created_by, "display_name": created_by},
            target_type="recipe_batch",
            target_label=f"{len(created_ids)} recipes from {path.name}",
            details={
                "source_file": str(path),
                "created_count": len(created_ids),
                "created_ids": created_ids,
            },
        )
        connection.commit()
        print(f"\n=== Committed {len(created_ids)} recipes ===")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Bulk import IRMS recipes from legacy Excel")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--dry-run", action="store_true", help="Parse and report without writing to DB")
    parser.add_argument("--created-by", default="excel-import", help="Creator name for audit")
    args = parser.parse_args()

    if not args.xlsx_path.exists():
        print(f"ERROR: file not found: {args.xlsx_path}")
        return 1
    return import_workbook(args.xlsx_path, args.dry_run, args.created_by)


if __name__ == "__main__":
    sys.exit(main())
