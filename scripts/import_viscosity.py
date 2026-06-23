"""합성 점도 데이터 → viscosity_readings 임포트 (멱등, 다중 시트 형태 지원).

지원 파일/시트 형태 (자동 판별):
  1. long-LOT   : PB/SBCT/SCRA 시트 (합성 점도.xlsx). A=LOT B=점도 C/D=메모 E=레시피 F=원료LOT
  2. wide-date  : 'TOP 점도' 시트. 날짜 1행 × 제품 여러 열(N-TOP/S-TOP/6-1 TOP/K-TOP)
  3. journal    : 합성일지.xlsx 연도 시트(2024/2025/2026). 일자·종류·점도·1차·2차·비고

제품은 코드로 자동 생성(없으면 INSERT). (product_id, lot_no) UNIQUE 로 재실행 멱등.
날짜 기반(wide/journal)은 lot_no = 'YYYY-MM-DD-NN'(같은 제품·날짜 순번) 으로 결정적
생성 → 같은 순서로 재임포트 시 중복 건너뜀.

운영 DB 에 적재하려면 서버와 동일한 환경변수로 실행:
    set IRMS_DATA_DIR=...   (서버가 쓰는 데이터 디렉토리)
    python scripts/import_viscosity.py "합성 점도.xlsx" "합성일지.xlsx"

인자를 안 주면 위 두 파일을 기본 시도한다.
"""

import re
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from src.db import get_connection, init_db, utc_now_text  # noqa: E402
from src.services import viscosity_service  # noqa: E402

# 'TOP 점도' 가로형 시트는 연도 표기가 없다 → 해당 파일(2026 데이터) 기준 연도.
WIDE_DEFAULT_YEAR = 2026

_DATE_KO = re.compile(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일")
_CODE_ALIAS = {"6-1-TOP": "6-1TOP"}


def _as_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _norm_code(value) -> str | None:
    if value is None:
        return None
    code = re.sub(r"\s+", "", str(value).strip().upper())
    if not code:
        return None
    return _CODE_ALIAS.get(code, code)


def _iso_date(value, default_year: int | None = None) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    text = str(value).strip()
    m = _DATE_KO.search(text)
    if m and default_year:
        try:
            return date(default_year, int(m.group(1)), int(m.group(2))).isoformat()
        except ValueError:
            return None
    return viscosity_service.parse_lot_date(text)


def _ensure_product(connection, code: str, now: str) -> dict:
    product = viscosity_service.get_product_by_code(connection, code)
    if product:
        return product
    connection.execute(
        "INSERT INTO viscosity_products (code, name, sigma_k, is_active, created_at) "
        "VALUES (?, ?, 3, 1, ?)",
        (code, code, now),
    )
    return viscosity_service.get_product_by_code(connection, code)


def _seq_lot(counters: dict, product_id: int, iso: str) -> str:
    key = (product_id, iso)
    counters[key] = counters.get(key, 0) + 1
    return f"{iso}-{counters[key]:02d}"


def _add(connection, stats, code, **kw) -> None:
    rec = stats.setdefault(code, [0, 0])
    try:
        viscosity_service.add_reading(connection, **kw)
        rec[0] += 1
    except sqlite3.IntegrityError:
        rec[1] += 1


def _find_header(sheet, marker: str, max_scan: int = 10):
    for r in range(1, min(sheet.max_row, max_scan) + 1):
        for c in range(1, sheet.max_column + 1):
            if _as_text(sheet.cell(r, c).value) == marker:
                return r, c
    return None, None


def _detect_shape(sheet) -> str | None:
    markers = set()
    for r in range(1, min(sheet.max_row, 8) + 1):
        for c in range(1, min(sheet.max_column, 12) + 1):
            t = _as_text(sheet.cell(r, c).value)
            if t:
                markers.add(t)
    if "점도 측정 날짜" in markers:
        return "wide"
    if "일자" in markers and "종류" in markers:
        return "journal"
    if "LOT" in markers:
        return "long"
    return None


def _import_long(connection, sheet, now, counters, stats) -> None:
    code = sheet.title.strip()
    product = _ensure_product(connection, code, now)
    header_row, _ = _find_header(sheet, "LOT")
    start = (header_row or 3) + 1
    for row in range(start, sheet.max_row + 1):
        lot_no = _as_text(sheet.cell(row, 1).value)
        visc = sheet.cell(row, 2).value
        if not lot_no or not isinstance(visc, (int, float)):
            continue
        memo_parts = [
            _as_text(sheet.cell(row, c).value)
            for c in (3, 4)
            if sheet.max_column >= c and _as_text(sheet.cell(row, c).value)
        ]
        _add(
            connection, stats, code,
            product_id=product["id"], lot_no=lot_no, viscosity=float(visc),
            measured_date=None,
            memo=" / ".join(memo_parts) if memo_parts else None,
            recipe_material=_as_text(sheet.cell(row, 5).value) if sheet.max_column >= 5 else None,
            material_lot=_as_text(sheet.cell(row, 6).value) if sheet.max_column >= 6 else None,
            created_by="엑셀 임포트", created_at=now,
        )


def _import_wide(connection, sheet, now, counters, stats) -> None:
    hr, dc = _find_header(sheet, "점도 측정 날짜")
    if hr is None:
        return
    prod_cols = {}
    for c in range(dc + 1, sheet.max_column + 1):
        code = _norm_code(sheet.cell(hr, c).value)
        if code:
            prod_cols[c] = code
    for row in range(hr + 1, sheet.max_row + 1):
        iso = _iso_date(sheet.cell(row, dc).value, WIDE_DEFAULT_YEAR)
        if not iso:
            continue
        for col, code in prod_cols.items():
            v = sheet.cell(row, col).value
            if not isinstance(v, (int, float)):
                continue
            product = _ensure_product(connection, code, now)
            _add(
                connection, stats, code,
                product_id=product["id"],
                lot_no=_seq_lot(counters, product["id"], iso),
                viscosity=float(v), measured_date=iso,
                memo=None, recipe_material=None, material_lot=None,
                created_by="엑셀 임포트", created_at=now,
            )


def _import_journal(connection, sheet, now, counters, stats) -> None:
    hr, date_col = _find_header(sheet, "일자")
    _, visc_col = _find_header(sheet, "점도")
    if hr is None or visc_col is None:
        return
    kind_col = visc_col - 1  # 종류 is immediately left of 점도 (summary 종류 block 회피)
    _, note_col = _find_header(sheet, "비고")
    for row in range(hr + 1, sheet.max_row + 1):
        v = sheet.cell(row, visc_col).value
        if not isinstance(v, (int, float)):
            continue
        iso = _iso_date(sheet.cell(row, date_col).value)
        code = _norm_code(sheet.cell(row, kind_col).value)
        if not iso or not code:
            continue
        product = _ensure_product(connection, code, now)
        memo = _as_text(sheet.cell(row, note_col).value) if note_col else None
        _add(
            connection, stats, code,
            product_id=product["id"],
            lot_no=_seq_lot(counters, product["id"], iso),
            viscosity=float(v), measured_date=iso,
            memo=memo, recipe_material=None, material_lot=None,
            created_by="합성일지 임포트", created_at=now,
        )


def run(paths: list[str]) -> None:
    init_db()
    now = utc_now_text()
    counters: dict = {}
    stats: dict = {}

    with get_connection() as connection:
        for path in paths:
            if not Path(path).exists():
                print(f"[건너뜀] 파일 없음: {path}")
                continue
            wb = load_workbook(path, data_only=True)
            print(f"=== {path} ===")
            for sheet in wb.worksheets:
                shape = _detect_shape(sheet)
                if shape == "long":
                    _import_long(connection, sheet, now, counters, stats)
                elif shape == "wide":
                    _import_wide(connection, sheet, now, counters, stats)
                elif shape == "journal":
                    _import_journal(connection, sheet, now, counters, stats)
                else:
                    print(f"  [건너뜀] 형태 판별 불가: {sheet.title}")
        connection.commit()

    total_new = sum(v[0] for v in stats.values())
    total_dup = sum(v[1] for v in stats.values())
    print("--- 제품별 신규(중복) ---")
    for code in sorted(stats):
        print(f"  {code}: {stats[code][0]} ({stats[code][1]})")
    print(f"완료: 신규 {total_new}건, 중복 건너뜀 {total_dup}건")


if __name__ == "__main__":
    args = sys.argv[1:] or ["합성 점도.xlsx", "합성일지.xlsx"]
    run(args)
