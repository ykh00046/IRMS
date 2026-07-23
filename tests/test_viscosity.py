"""Unit + route tests for viscosity-analysis.

Design: docs/02-design/features/viscosity-analysis.design.md
"""

from __future__ import annotations

import sqlite3
from datetime import datetime

import pytest

from src.services import viscosity_service as vs


def _make_db() -> sqlite3.Connection:
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    connection.executescript(
        """
        CREATE TABLE viscosity_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            target REAL,
            lower_limit REAL,
            upper_limit REAL,
            sigma_k REAL NOT NULL DEFAULT 3,
            rpm REAL,
            temperature REAL,
            remind_daily INTEGER NOT NULL DEFAULT 0,
            use_reactor INTEGER NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );
        CREATE TABLE viscosity_readings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            lot_no TEXT NOT NULL,
            viscosity REAL NOT NULL,
            measured_date TEXT,
            memo TEXT,
            recipe_material TEXT,
            material_lot TEXT,
            reactor INTEGER,
            created_by TEXT,
            created_at TEXT NOT NULL,
            blend_record_id INTEGER
        );
        CREATE UNIQUE INDEX idx_visc_readings_product_lot
            ON viscosity_readings(product_id, lot_no);
        """
    )
    return connection


def _add_product(conn, code="PB", **kw) -> dict:
    conn.execute(
        "INSERT INTO viscosity_products (code, name, target, lower_limit, upper_limit, "
        "sigma_k, is_active, created_at) VALUES (?, ?, ?, ?, ?, ?, 1, '2026-01-01T00:00:00Z')",
        (
            code, kw.get("name", code), kw.get("target"),
            kw.get("lower_limit"), kw.get("upper_limit"), kw.get("sigma_k", 3),
        ),
    )
    return vs.get_product_by_code(conn, code)


def _seed(conn, product_id, values, start_seq=1):
    for i, v in enumerate(values):
        vs.add_reading(
            conn, product_id=product_id, lot_no=f"2601{start_seq + i:04d}",
            viscosity=v, measured_date=f"2026-01-{(i % 28) + 1:02d}",
            memo=None, recipe_material=None, material_lot=None,
            created_by="test", created_at="2026-01-01T00:00:00Z",
        )


# ── 반응기 ──────────────────────────────────────────────────────
def test_reactor_stored_and_filtered():
    """반응기별 저장 + analyze_product(reactor=)로 해당 반응기 표본만 분석."""
    conn = _make_db()
    p = _add_product(conn, "PB")
    vs.add_reading(
        conn, product_id=p["id"], lot_no="26010101", viscosity=49.0,
        measured_date="2026-01-01", memo=None, recipe_material=None,
        material_lot=None, created_by="t", created_at="2026-01-01T00:00:00Z",
        reactor=1,
    )
    vs.add_reading(
        conn, product_id=p["id"], lot_no="26010201", viscosity=90.0,
        measured_date="2026-01-02", memo=None, recipe_material=None,
        material_lot=None, created_by="t", created_at="2026-01-01T00:00:00Z",
        reactor=2,
    )
    assert vs.available_reactors(conn, p["id"]) == [1, 2]

    all_read = vs.analyze_product(conn, p)
    assert all_read["stats"]["n"] == 2
    assert all_read["available_reactors"] == [1, 2]

    only_r2 = vs.analyze_product(conn, p, reactor=2)
    assert only_r2["stats"]["n"] == 1
    assert only_r2["readings"][0]["viscosity"] == 90.0
    assert only_r2["readings"][0]["reactor"] == 2
    assert only_r2["reactor"] == 2


# ── LOT 날짜 파서 ────────────────────────────────────────────────
def test_parse_lot_date_8digit():
    assert vs.parse_lot_date(26010701) == "2026-01-07"  # PB YYMMDD+seq


def test_parse_lot_date_6digit():
    assert vs.parse_lot_date(260106) == "2026-01-06"  # SBCT YYMMDD


def test_parse_lot_date_datetime():
    assert vs.parse_lot_date(datetime(2026, 3, 14)) == "2026-03-14"


def test_parse_lot_date_iso_string():
    assert vs.parse_lot_date("2026-02-28 00:00:00") == "2026-02-28"


def test_parse_lot_date_invalid():
    assert vs.parse_lot_date("ABC") is None
    assert vs.parse_lot_date(None) is None


# ── 통계 관리한계 + 이상 판정 ───────────────────────────────────
def test_sigma_anomaly_detected():
    """평균에서 멀리 떨어진 값은 ±kσ 위반 → anomaly."""
    conn = _make_db()
    p = _add_product(conn, "PB", sigma_k=3)
    _seed(conn, p["id"], [49.0] * 10 + [60.0])  # 60은 명백한 outlier
    result = vs.analyze_product(conn, p)
    assert result["counts"]["anomaly"] == 1
    top = result["anomalies"][0]
    assert top["viscosity"] == 60.0
    assert "sigma_high" in top["reasons"]
    assert top["side"] == "high"


def test_spec_limit_anomaly():
    """관리 상/하한을 벗어나면 표준편차와 무관하게 anomaly."""
    conn = _make_db()
    p = _add_product(conn, "SBCT", lower_limit=195, upper_limit=215)
    _seed(conn, p["id"], [200, 205, 210, 218])  # 218 > 215
    result = vs.analyze_product(conn, p)
    assert result["counts"]["anomaly"] == 1
    assert "spec_high" in result["anomalies"][0]["reasons"]


def test_target_used_as_center_when_set():
    """target 설정 시 중심선은 평균이 아니라 target."""
    conn = _make_db()
    p = _add_product(conn, "PB", target=50)
    _seed(conn, p["id"], [49, 49, 49, 49])
    result = vs.analyze_product(conn, p)
    assert result["stats"]["center"] == 50.0


def test_warn_zone():
    """2σ 초과 ~ kσ 이하는 anomaly 가 아니라 warn."""
    conn = _make_db()
    p = _add_product(conn, "PB", sigma_k=3)
    # center≈50.2, std≈0.632 → uwl≈51.46, ucl≈52.10. 52.0 은 2σ~3σ 경고 구간.
    _seed(conn, p["id"], [50] * 9 + [52.0])
    result = vs.analyze_product(conn, p)
    assert result["counts"]["warn"] == 1
    assert result["counts"]["anomaly"] == 0


def test_normal_when_no_spec_and_low_variance():
    conn = _make_db()
    p = _add_product(conn, "PB")
    _seed(conn, p["id"], [49, 49, 49, 49, 49])
    result = vs.analyze_product(conn, p)
    assert result["counts"]["anomaly"] == 0
    assert result["counts"]["warn"] == 0


# ── 추세 룰 ─────────────────────────────────────────────────────
def test_run_up_trend():
    """연속 상승 5회 → run_up 추세 경보."""
    conn = _make_db()
    p = _add_product(conn, "PB")
    _seed(conn, p["id"], [48, 48.5, 49, 49.5, 50, 50.5])
    result = vs.analyze_product(conn, p)
    types = [t["type"] for t in result["trends"]]
    assert "run_up" in types


def test_run_down_trend():
    """연속 하락 5회 → run_down 추세 경보 (POLISH-1: 상승/하락 대칭 검출 회귀 방지)."""
    conn = _make_db()
    p = _add_product(conn, "PB")
    _seed(conn, p["id"], [52, 51.5, 51, 50.5, 50, 49.5])
    result = vs.analyze_product(conn, p)
    trends = {t["type"]: t for t in result["trends"]}
    assert "run_down" in trends
    assert "run_up" not in trends  # 단조 하락 tail 은 run_up 을 내지 않는다
    assert trends["run_down"]["length"] >= 5


# ── 제품 코드 정규화(GAP-4) ─────────────────────────────────────
def test_ensure_product_by_code_normalizes_case_and_space():
    """GAP-4: 대소문자/앞뒤 공백만 다른 코드는 같은 논리적 제품으로 귀결(중복 생성 방지).

    자동 생성(ensure_product_by_code)·조회(get_product_by_code)가 리마인더 쿼리와 같은
    strip+upper 정규화를 쓴다.
    """
    conn = _make_db()
    p1 = vs.ensure_product_by_code(conn, "PB", "PB", "2026-01-01T00:00:00Z")
    assert p1 is not None
    # 소문자 + 앞뒤 공백 → 새로 만들지 않고 같은 제품을 반환
    p2 = vs.ensure_product_by_code(conn, "  pb  ", "pb", "2026-01-02T00:00:00Z")
    assert p2 is not None and p2["id"] == p1["id"]
    # 제품 행은 하나만 존재
    assert conn.execute("SELECT COUNT(*) AS c FROM viscosity_products").fetchone()["c"] == 1
    # get_product_by_code 도 정규화 조회
    assert vs.get_product_by_code(conn, "Pb")["id"] == p1["id"]


# ── 등록 + 멱등성 ───────────────────────────────────────────────
def test_add_reading_derives_date_from_lot():
    conn = _make_db()
    p = _add_product(conn, "PB")
    vs.add_reading(
        conn, product_id=p["id"], lot_no="26010701", viscosity=49.0,
        measured_date=None, memo=None, recipe_material=None, material_lot=None,
        created_by="t", created_at="2026-06-23T00:00:00Z",
    )
    row = conn.execute("SELECT measured_date FROM viscosity_readings").fetchone()
    assert row["measured_date"] == "2026-01-07"


def test_duplicate_lot_rejected():
    conn = _make_db()
    p = _add_product(conn, "PB")
    kw = dict(
        product_id=p["id"], lot_no="26010701", viscosity=49.0, measured_date=None,
        memo=None, recipe_material=None, material_lot=None, created_by="t",
        created_at="2026-06-23T00:00:00Z",
    )
    vs.add_reading(conn, **kw)
    with pytest.raises(sqlite3.IntegrityError):
        vs.add_reading(conn, **kw)


# ── 기간별(분기/월) 집계 ────────────────────────────────────────
def test_period_key_quarter_and_month():
    assert vs._period_key("2026-03-14", "quarter") == "2026-Q1"
    assert vs._period_key("2026-04-01", "quarter") == "2026-Q2"
    assert vs._period_key("2026-03-14", "month") == "2026-03"
    assert vs._period_key(None, "quarter") is None
    assert vs._period_key("bad", "quarter") is None


def test_period_key_day_and_week():
    assert vs._period_key("2026-03-15", "day") == "2026-03-15"
    # ISO 주차 — 2026-03-15 은 11주차, 다음날은 12주차(주 경계)
    assert vs._period_key("2026-03-15", "week") == "2026-W11"
    assert vs._period_key("2026-03-16", "week") == "2026-W12"
    # 연 초 ISO 주차(=W01), 잘못된 날짜는 None
    assert vs._period_key("2026-01-01", "week") == "2026-W01"
    assert vs._period_key("2026-13-40", "day") is None
    # 알 수 없는 granularity 는 분기로 폴백(기존 기본 동작 보존)
    assert vs._period_key("2026-03-15", "bogus") == "2026-Q1"


def test_quarterly_summary_with_delta():
    """분기별 평균 + 전기대비(mean_delta) 계산."""
    conn = _make_db()
    p = _add_product(conn, "PB")
    # Q1: 1~3월, Q2: 4~6월 — 측정일을 직접 지정
    specs = [
        ("26010101", 48.0, "2026-01-10"),
        ("26020101", 50.0, "2026-02-10"),
        ("26040101", 52.0, "2026-04-10"),
        ("26050101", 54.0, "2026-05-10"),
    ]
    for lot, v, d in specs:
        vs.add_reading(
            conn, product_id=p["id"], lot_no=lot, viscosity=v, measured_date=d,
            memo=None, recipe_material=None, material_lot=None,
            created_by="t", created_at="2026-06-23T00:00:00Z",
        )
    result = vs.analyze_product(conn, p, granularity="quarter")
    periods = result["periods"]
    assert [x["period"] for x in periods] == ["2026-Q1", "2026-Q2"]
    assert periods[0]["mean"] == 49.0  # (48+50)/2
    assert periods[0]["count"] == 2
    assert periods[0]["mean_delta"] is None
    assert periods[1]["mean"] == 53.0  # (52+54)/2
    assert periods[1]["mean_delta"] == 4.0  # 53 - 49


def test_monthly_granularity():
    conn = _make_db()
    p = _add_product(conn, "PB")
    for lot, d in [("26010101", "2026-01-05"), ("26020101", "2026-02-05")]:
        vs.add_reading(
            conn, product_id=p["id"], lot_no=lot, viscosity=49.0, measured_date=d,
            memo=None, recipe_material=None, material_lot=None,
            created_by="t", created_at="2026-06-23T00:00:00Z",
        )
    periods = vs.analyze_product(conn, p, granularity="month")["periods"]
    assert [x["period"] for x in periods] == ["2026-01", "2026-02"]


# ── 기간 알림 (이상 급증 / 평균 이동) ──────────────────────────
def test_period_anomaly_spike_alert():
    """spec 위반 다수가 한 분기에 몰리면 anomaly_spike 경보."""
    conn = _make_db()
    p = _add_product(conn, "PB", upper_limit=50)
    # Q1 정상, Q2 에 상한 초과 3건
    rows = [
        ("26010101", 49, "2026-01-10"), ("26020101", 49, "2026-02-10"),
        ("26040101", 55, "2026-04-10"), ("26050101", 56, "2026-05-10"),
        ("26060101", 57, "2026-06-10"),
    ]
    for lot, v, d in rows:
        vs.add_reading(
            conn, product_id=p["id"], lot_no=lot, viscosity=v, measured_date=d,
            memo=None, recipe_material=None, material_lot=None,
            created_by="t", created_at="2026-06-23T00:00:00Z",
        )
    alerts = vs.analyze_product(conn, p, granularity="quarter")["period_alerts"]
    spikes = [a for a in alerts if a["type"] == "anomaly_spike"]
    assert spikes and spikes[0]["period"] == "2026-Q2"
    assert spikes[0]["anomaly_count"] == 3


def test_period_mean_shift_alert():
    """전기대비 평균이 전체 σ 이상 이동하면 mean_shift 경보."""
    conn = _make_db()
    p = _add_product(conn, "PB")
    rows = [
        ("26010101", 40, "2026-01-10"), ("26020101", 40, "2026-02-10"),
        ("26040101", 60, "2026-04-10"), ("26050101", 60, "2026-05-10"),
    ]
    for lot, v, d in rows:
        vs.add_reading(
            conn, product_id=p["id"], lot_no=lot, viscosity=v, measured_date=d,
            memo=None, recipe_material=None, material_lot=None,
            created_by="t", created_at="2026-06-23T00:00:00Z",
        )
    alerts = vs.analyze_product(conn, p, granularity="quarter")["period_alerts"]
    shifts = [a for a in alerts if a["type"].startswith("mean_shift")]
    assert shifts and shifts[0]["type"] == "mean_shift_up"


# ── 연도별 기준 ─────────────────────────────────────────────────
def _add_dated(conn, pid, lot, v, d):
    vs.add_reading(
        conn, product_id=pid, lot_no=lot, viscosity=v, measured_date=d,
        memo=None, recipe_material=None, material_lot=None,
        created_by="t", created_at="2026-06-23T00:00:00Z",
    )


def test_available_years_desc():
    conn = _make_db()
    p = _add_product(conn, "N-TOP")
    _add_dated(conn, p["id"], "a", 165, "2024-03-04")
    _add_dated(conn, p["id"], "b", 144, "2025-01-09")
    _add_dated(conn, p["id"], "c", 131, "2026-05-04")
    assert vs.available_years(conn, p["id"]) == [2026, 2025, 2024]


def test_year_filtered_baseline():
    """같은 제품이라도 연도별로 중심선/표본이 분리된다."""
    conn = _make_db()
    p = _add_product(conn, "N-TOP")
    # 2024 대역 ~165, 2026 대역 ~131
    for i, v in enumerate([160, 165, 170]):
        _add_dated(conn, p["id"], f"24{i}", v, f"2024-03-0{i+1}")
    for i, v in enumerate([130, 131, 132]):
        _add_dated(conn, p["id"], f"26{i}", v, f"2026-05-0{i+1}")

    a24 = vs.analyze_product(conn, p, year=2024)
    a26 = vs.analyze_product(conn, p, year=2026)
    assert a24["stats"]["n"] == 3 and a24["stats"]["mean"] == 165.0
    assert a26["stats"]["n"] == 3 and a26["stats"]["mean"] == 131.0
    assert a24["year"] == 2024 and a26["year"] == 2026
    # 전체(year=None)는 두 대역을 합쳐 σ가 매우 커진다
    allp = vs.analyze_product(conn, p)
    assert allp["stats"]["n"] == 6
    assert allp["stats"]["std"] > a24["stats"]["std"]


def test_year_granularity_buckets():
    conn = _make_db()
    p = _add_product(conn, "SBCT")
    _add_dated(conn, p["id"], "a", 200, "2024-03-04")
    _add_dated(conn, p["id"], "b", 210, "2024-05-04")
    _add_dated(conn, p["id"], "c", 198, "2025-01-09")
    periods = vs.analyze_product(conn, p, granularity="year")["periods"]
    assert [x["period"] for x in periods] == ["2024", "2025"]
    assert periods[0]["mean"] == 205.0
    assert periods[1]["mean_delta"] == -7.0  # 198 - 205


def test_day_and_week_granularity_buckets():
    conn = _make_db()
    p = _add_product(conn, "SBCT")
    _add_dated(conn, p["id"], "a", 200, "2026-03-15")
    _add_dated(conn, p["id"], "b", 210, "2026-03-15")  # 같은 날 → 한 버킷
    _add_dated(conn, p["id"], "c", 198, "2026-03-16")  # 다음 주(주 경계)
    days = vs.analyze_product(conn, p, granularity="day")["periods"]
    assert [x["period"] for x in days] == ["2026-03-15", "2026-03-16"]
    assert days[0]["count"] == 2 and days[0]["mean"] == 205.0
    weeks = vs.analyze_product(conn, p, granularity="week")["periods"]
    assert [x["period"] for x in weeks] == ["2026-W11", "2026-W12"]


# ── 신규 입력 즉시 판정 ─────────────────────────────────────────
def test_classify_value_for_new_input():
    """입력 전 표본 기준으로 신규 값을 판정 (등록 즉시 경고용)."""
    conn = _make_db()
    p = _add_product(conn, "PB", upper_limit=50)
    _seed(conn, p["id"], [49, 49, 49])
    verdict = vs.classify_value(conn, p, 55.0)
    assert verdict["status"] == "anomaly"
    assert "spec_high" in verdict["reasons"]
    normal = vs.classify_value(conn, p, 49.0)
    assert normal["status"] == "normal"


# ── overview 집계 ───────────────────────────────────────────────
def test_overview_aggregates_anomaly_count():
    conn = _make_db()
    p1 = _add_product(conn, "PB", sigma_k=3)
    _seed(conn, p1["id"], [49.0] * 10 + [60.0])
    _add_product(conn, "SBCT")
    ov = vs.overview(conn)
    assert ov["product_count"] == 2
    assert ov["total_anomaly"] == 1
    pb = next(it for it in ov["items"] if it["code"] == "PB")
    assert pb["anomaly_count"] == 1
    assert pb["latest_value"] == 60.0


# ── 라우트 인증 ─────────────────────────────────────────────────
def test_viscosity_route_is_public():
    """점도 조회는 로그인 없이 접근 가능(사내 공용 단말 운영 편의)."""
    import importlib

    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    client = TestClient(mainmod.app)
    res = client.get("/api/viscosity/overview")
    assert res.status_code == 200
    # 마이그레이션 시드 제품(PB/SBCT/SCRA)이 비로그인으로도 보인다
    codes = {it["code"] for it in res.json()["items"]}
    assert {"PB", "SBCT", "SCRA"} <= codes


def test_viscosity_page_open_without_login():
    """/viscosity 페이지가 로그인 리다이렉트 없이 200 으로 열린다."""
    import importlib

    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    client = TestClient(mainmod.app)
    res = client.get("/viscosity")
    assert res.status_code == 200
    assert "visc-period-chart" in res.text
    assert client.cookies.get("csrftoken")


# ── GAP-3: anomaly_spike 완화 게이트 ────────────────────────────
def test_period_anomaly_spike_gated_to_coarse():
    """일/주 단위에서는 anomaly_spike 가 뜨지 않고 월/분기/연에서만 뜬다(GAP-3).

    mean_shift 만 coarse 게이트에 있고 anomaly_spike 는 제한이 없어, 일/주 조회 시
    한 버킷에 이상 2건만 몰려도 경보가 뜨던 과민을 mean_shift 와 동일하게 막는다.
    """
    periods = [
        {"period": "2026-01-01", "anomaly_count": 0, "mean_delta": None},
        {"period": "2026-01-02", "anomaly_count": 3, "mean_delta": 0.0},
    ]
    for gran in ("day", "week"):
        alerts = vs._period_alerts(periods, 1.0, gran)
        assert not any(a["type"] == "anomaly_spike" for a in alerts), gran
    for gran in ("month", "quarter", "year"):
        alerts = vs._period_alerts(periods, 1.0, gran)
        assert any(a["type"] == "anomaly_spike" for a in alerts), gran


# ── POLISH-2: 경고 밴드 붕괴 ─────────────────────────────────────
def test_warn_band_collapses_when_sigma_k_le_warn_sigma():
    """sigma_k <= WARN_SIGMA(2) 이면 경고 밴드(uwl/lwl)가 None 이 되어 경고가 사라진다.

    kσ(UCL) 가 2σ(UWL) 안쪽/동일이라 경고가 이상보다 바깥에 놓이는 역전을 차단.
    """
    conn = _make_db()
    p2 = _add_product(conn, "PBK2", sigma_k=2)
    _seed(conn, p2["id"], [40.0, 42.0, 44.0, 46.0, 48.0, 50.0])
    res2 = vs.analyze_product(conn, p2)
    assert res2["stats"]["std"] > 0
    assert res2["stats"]["uwl"] is None
    assert res2["stats"]["lwl"] is None
    # 경고(warn) 판정 자체가 나오지 않는다 — 정상/이상만.
    assert all(r["status"] != "warn" for r in res2["readings"])

    # 대조군: sigma_k=3 이면 경고 밴드가 존재한다.
    p3 = _add_product(conn, "PBK3", sigma_k=3)
    _seed(conn, p3["id"], [40.0, 42.0, 44.0, 46.0, 48.0, 50.0], start_seq=100)
    res3 = vs.analyze_product(conn, p3)
    assert res3["stats"]["uwl"] is not None
    assert res3["stats"]["lwl"] is not None


# ── 정책 ⓑ: management 성격 점도 쓰기 서버 강제 ──────────────────
def _visc_client():
    import importlib

    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    return TestClient(mainmod.app)


def test_viscosity_manager_writes_denied_when_anonymous():
    """제품 생성/수정, 측정 삭제, CSV export 는 비로그인에서 거부(401/403)."""
    client = _visc_client()
    # CSRF 토큰만 확보(로그인 안 함) — CSRF 를 통과해도 인증 게이트에서 막혀야 한다.
    client.get("/viscosity")
    tok = client.cookies.get("csrftoken")
    headers = {"x-csrftoken": tok} if tok else {}

    r = client.post("/api/viscosity/products", json={"code": "ZZ", "name": "ZZ"}, headers=headers)
    assert r.status_code in (401, 403), r.text
    r = client.patch("/api/viscosity/products/1", json={"name": "x"}, headers=headers)
    assert r.status_code in (401, 403), r.text
    r = client.delete("/api/viscosity/readings/1", headers=headers)
    assert r.status_code in (401, 403), r.text
    # export 는 GET(CSRF 무관) — 인증만으로 막힌다.
    r = client.get("/api/viscosity/products/1/export")
    assert r.status_code in (401, 403), r.text


def test_viscosity_reads_and_registration_stay_open():
    """정책 ⓑ 이후에도 조회(overview/products)는 무로그인 200 유지."""
    client = _visc_client()
    assert client.get("/api/viscosity/overview").status_code == 200
    assert client.get("/api/viscosity/products").status_code == 200


def _seed_export_product(client, headers, product, dates):
    """레시피 임포트 → 점도 제품 생성 → dates 각각에 측정 1건 등록. product id 반환."""
    header_row = "\t".join(["반제품명", "원료A", "원료B"])
    data_row = "\t".join([product, "60", "40"])
    imported = client.post(
        "/api/recipes/import", json={"raw_text": "\n".join([header_row, data_row])},
        headers=headers,
    )
    assert imported.status_code == 200, imported.text
    created = client.post(
        "/api/viscosity/products", json={"code": product, "name": product}, headers=headers
    )
    assert created.status_code in (200, 201), created.text
    pid = created.json()["id"]
    for seq, d in enumerate(dates, start=1):
        res = client.post("/api/viscosity/readings", json={
            "product_id": pid, "lot_no": f"{product}L{seq}", "viscosity": 50.0 + seq,
            "measured_date": d,
        }, headers=headers)
        assert res.status_code == 200, res.text
    return pid


def _load_export_wb(resp):
    """export 응답(xlsx)을 openpyxl 워크북으로 로드. PK 매직 바이트 검증 포함."""
    import io as _io

    from openpyxl import load_workbook

    assert resp.status_code == 200, resp.text
    assert resp.content[:2] == b"PK", "xlsx(zip) 매직 바이트가 아님"
    return load_workbook(_io.BytesIO(resp.content))


def test_viscosity_export_status_uses_year_filter():
    """GAP-2: Excel export 가 year 파라미터를 analyze_product 로 넘겨 화면과 같은
    연도 표본으로 판정·필터한다 — year 지정 시 해당 연도 측정만 '측정 원본' 시트에 나온다."""
    import uuid

    client = _visc_client()
    login = client.post(
        "/api/auth/management-login", json={"username": "admin", "password": "admin"}
    )
    assert login.status_code == 200, login.text
    tok = client.cookies.get("csrftoken")
    headers = {"x-csrftoken": tok} if tok else {}

    product = "YRP" + uuid.uuid4().hex[:6].upper()
    # 2025 3건 + 2026 2건.
    pid = _seed_export_product(
        client, headers, product,
        ("2025-03-01", "2025-03-02", "2025-03-03", "2026-04-01", "2026-04-02"),
    )

    def _years(resp):
        wb = _load_export_wb(resp)
        ws = wb["측정 원본"]
        years = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            measured_date = row[1]  # 헤더 2번째 컬럼 = 측정일
            if measured_date:
                years.add(str(measured_date)[:4])
        return sorted(years)

    assert _years(client.get(f"/api/viscosity/products/{pid}/export?year=2026")) == ["2026"]
    assert _years(client.get(f"/api/viscosity/products/{pid}/export?year=2025")) == ["2025"]
    # year 미지정 = 전체 연도.
    assert _years(client.get(f"/api/viscosity/products/{pid}/export")) == ["2025", "2026"]


def test_viscosity_export_xlsx_structure_and_period_rows():
    """Excel export: PK 매직 바이트 + 두 시트 이름 + '기간 요약' 데이터 행 수가
    같은 필터의 analyze_product periods 수와 일치한다."""
    import uuid

    client = _visc_client()
    login = client.post(
        "/api/auth/management-login", json={"username": "admin", "password": "admin"}
    )
    assert login.status_code == 200, login.text
    tok = client.cookies.get("csrftoken")
    headers = {"x-csrftoken": tok} if tok else {}

    product = "XLS" + uuid.uuid4().hex[:6].upper()
    # 4개월에 걸친 측정 → 월 단위로 4개 기간 버킷.
    dates = ("2026-01-05", "2026-02-05", "2026-03-05", "2026-04-05")
    pid = _seed_export_product(client, headers, product, dates)

    # 화면과 같은 필터(월 단위)로 기간 수 확인.
    detail = client.get(f"/api/viscosity/products/{pid}?granularity=month")
    assert detail.status_code == 200, detail.text
    expected_periods = len(detail.json()["periods"])
    assert expected_periods == 4

    resp = client.get(f"/api/viscosity/products/{pid}/export?granularity=month")
    wb = _load_export_wb(resp)
    assert wb.sheetnames == ["측정 원본", "기간 요약"]

    ws_readings = wb["측정 원본"]
    # 헤더 1행 + 측정 4건.
    assert ws_readings.max_row == 1 + len(dates)
    assert ws_readings.cell(row=1, column=1).value == "LOT"

    ws_periods = wb["기간 요약"]
    assert ws_periods.cell(row=1, column=1).value == "기간"
    period_rows = ws_periods.max_row - 1  # 헤더 제외
    assert period_rows == expected_periods

    # Content-Disposition 파일명 규칙.
    disposition = resp.headers.get("content-disposition", "")
    assert f"viscosity_{product}_" in disposition and disposition.endswith('.xlsx"')


def test_direct_registration_stores_the_date_used_for_judgement(monkeypatch):
    """감사 F-9: 라우트가 판정에 쓴 측정일이 그대로 저장돼야 한다 (폴백 1회).

    옛 코드는 라우트에서 resolved_date 로 판정 연도를 정한 뒤, 서비스에는 원본(None)을
    넘겨 **같은 폴백을 다시** 돌렸다 — 자정 경계에서 판정 연도와 저장 연도가 갈릴 수 있는
    구조. 측정일 미지정 + LOT 에서 날짜 추론이 되는 경우, 저장된 measured_date 는
    LOT 이 가리키는 날짜여야 한다(등록일이 아니라).
    """
    import importlib
    import uuid

    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    from src.db import get_connection

    client = TestClient(mainmod.app)
    client.post("/api/auth/management-login", json={"username": "admin", "password": "admin"})
    tok = client.cookies.get("csrftoken")
    headers = {"x-csrftoken": tok} if tok else {}

    # 점도 제품 code 는 레시피 제품명과 연동 — 레시피부터 등록한다.
    product = "F9P" + uuid.uuid4().hex[:6].upper()
    header_row = "\t".join(["반제품명", "원료A", "원료B"])
    data_row = "\t".join([product, "60", "40"])
    raw_text = "\n".join([header_row, data_row])
    imported = client.post("/api/recipes/import", json={"raw_text": raw_text},
                           headers=headers)
    assert imported.status_code == 200, imported.text

    created = client.post("/api/viscosity/products",
                          json={"code": product, "name": product}, headers=headers)
    assert created.status_code in (200, 201), created.text
    pid = created.json()["id"]

    # 두 폴백이 갈리는 지점을 강제한다: 측정일 미지정 + LOT 에서 날짜 추론 불가 →
    # 라우트는 local_today_text(), 서비스는 date.today() 를 각각 쓴다. 라우트 쪽만
    # 다른 날짜로 바꿔치면, 옛 코드는 서비스가 '진짜 오늘'로 다시 폴백해 어긋난다.
    import src.routers.viscosity_routes as vroutes

    monkeypatch.setattr(vroutes, "local_today_text", lambda: "2026-12-31")

    lot = "NODATE" + uuid.uuid4().hex[:6].upper()   # parse_lot_date 가 못 읽는 LOT
    res = client.post("/api/viscosity/readings", json={
        "product_id": pid, "lot_no": lot, "viscosity": 12.3,
    }, headers=headers)
    assert res.status_code == 200, res.text

    with get_connection() as conn:   # 등록 응답은 분석 결과라 id 가 없다 — LOT 으로 조회
        row = conn.execute(
            "SELECT measured_date FROM viscosity_readings WHERE product_id = ? AND lot_no = ?",
            (pid, lot),
        ).fetchone()
    assert row is not None
    # 판정에 쓴 날짜(라우트가 정한 값)가 그대로 저장돼야 한다 — 서비스가 다시 폴백하면 안 된다
    assert row["measured_date"] == "2026-12-31"
