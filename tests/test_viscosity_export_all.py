"""GET /api/viscosity/export-all — 점도 전체 Excel 내보내기 라우트 테스트.

패턴은 tests/test_blend.py 의 importlib.reload(cfg/mainmod) + TestClient +
admin/admin 로그인 + uuid 제품명 시드를 따른다. 책임자 전용 통제(미로그인
401/403), 200 + xlsx content-type, 시드한 두 반제품의 측정 행이 모두 존재,
판정 열 값이 한글(정상/경고/이상) 또는 빈칸임을 openpyxl 로 단언.
"""

from __future__ import annotations

import io
import importlib
import uuid


def _reload_app():
    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    return TestClient(mainmod.app)


def _seed(client, headers, code):
    """레시피 1건 등록 → 점도 제품 생성(target/limit/sigma) → 측정 2건 등록.

    반환: (반제품 코드, 점도 제품 id). 판정이 정상으로 떨어지도록 target 근처의
    안정적인 값을 넣는다(시그마 한계 안).
    """
    materials = {"MatX": 60, "MatY": 40}
    header_line = "반제품명\t" + "\t".join(materials.keys())
    row_line = f"{code}\t" + "\t".join(str(v) for v in materials.values())
    raw = header_line + "\n" + row_line
    r = client.post(
        "/api/recipes/import",
        json={"raw_text": raw, "force": True},
        headers=headers,
    )
    assert r.status_code == 200, ("import", r.status_code, r.text)

    r = client.post(
        "/api/viscosity/products",
        json={
            "code": code, "name": code + "명",
            "target": 50.0, "lower_limit": 40.0, "upper_limit": 60.0, "sigma_k": 3,
        },
        headers=headers,
    )
    assert r.status_code == 200, ("product", r.status_code, r.text)
    product_id = r.json()["id"]

    for j, v in enumerate((50.0, 50.5)):
        body = {
            "product_id": product_id,
            "lot_no": f"{code[:6]}{2607010}{j}",
            "viscosity": v,
            "measured_date": "2026-07-01",
            "memo": "시드",
        }
        r = client.post("/api/viscosity/readings", json=body, headers=headers)
        assert r.status_code == 200, ("reading", code, v, r.status_code, r.text)
    return code, product_id



def test_viscosity_export_all_requires_manager():
    """미로그인(또는 현장 권한) 접근은 401 또는 403."""
    client = _reload_app()
    res = client.get("/api/viscosity/export-all")
    assert res.status_code in (401, 403), res.text


def test_viscosity_export_all_returns_flat_sheet_with_seed_data():
    """책임자 로그인 후 200 + xlsx, 시트 '전체 측정', 두 반제품 행 모두 존재,
    판정 열은 한글(정상/경고/이상) 또는 빈칸."""
    client = _reload_app()
    client.post("/api/auth/management-login", json={"username": "admin", "password": "admin"})
    tok = client.cookies.get("csrftoken")
    headers = {"x-csrftoken": tok} if tok else {}

    p1 = "VX" + uuid.uuid4().hex[:5]
    p2 = "VY" + uuid.uuid4().hex[:5]
    code1, _ = _seed(client, headers, p1)
    code2, _ = _seed(client, headers, p2)

    res = client.get("/api/viscosity/export-all")
    assert res.status_code == 200, res.text
    assert "spreadsheetml" in res.headers.get("content-type", ""), res.headers.get("content-type")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content))
    assert wb.sheetnames == ["전체 측정"], wb.sheetnames
    ws = wb["전체 측정"]
    headers_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers_row[0] == "반제품 코드"
    assert headers_row[6] == "판정"

    codes = set()
    verdict_values = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        codes.add(row[0])
        verdict_values.add(row[6])
    assert code1 in codes, (code1, codes)
    assert code2 in codes, (code2, codes)
    allowed = {"정상", "경고", "이상", "", None}
    assert verdict_values <= allowed, verdict_values
