"""바인더 점도 임포트 — 배합 기록 연계 검사.

바인더 점도는 별도로 저장하는 게 아니라, 매칭되는 배합 기록(product_lot =
바인더종류 + 사용한PB)에 blend_record_id 로 연계돼야 한다 — 배합 화면 점도 등록과
같은 형태. 배합 기록이 없으면 건너뛰고, 이전 버전이 남긴 미연계 행은 정리한다.
"""

import importlib

from openpyxl import Workbook


def _make_binder_file(path, pb_prefix="2606"):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("26년도 바인더 점도 기록")
    ws.append(["일자", "바인더", "사용한PB", "점도", "작업자"])
    ws.append(["26년6월1일", "APB", f"{pb_prefix}0101", 385.5, "이시현"])
    ws.append(["26년6월1일", "CSBP", f"{pb_prefix}0102", 78.6, "이재석"])   # 오타→CSPB, 배합기록 없음
    ws.append(["26년6월2일", "APB(17)", f"{pb_prefix}0201", 384, "이재석"])  # →APB17, 배합기록 없음
    wb.save(path)


def _seed_blend_record(conn, product_lot, product_name):
    from src.db import utc_now_text
    conn.execute(
        "INSERT INTO blend_records (product_lot, product_name, worker, work_date, "
        "total_amount, status, created_at) VALUES (?, ?, '홍길동', '2026-06-01', 1000, 'completed', ?)",
        (product_lot, product_name, utc_now_text()),
    )
    conn.commit()


def test_binder_links_to_matching_blend_record(tmp_path):
    import src.config as cfg

    importlib.reload(cfg)
    from src.db import get_connection
    from src.services import viscosity_service

    from src.db import init_db
    init_db()
    xlsx = tmp_path / "바인더.xlsx"
    _make_binder_file(xlsx, pb_prefix="2606")

    with get_connection() as conn:
        _seed_blend_record(conn, "APB26060101", "APB")   # 이 배합만 존재

    import scripts.import_binder_viscosity as imp

    stats = imp.import_binder([str(xlsx)])

    assert stats["linked"] == 1     # APB26060101 만 배합 기록에 연계
    assert stats["archived"] == 2   # CSPB/APB17 은 배합 없어 과거로 보존

    with get_connection() as conn:
        row = conn.execute(
            "SELECT r.lot_no, r.viscosity, r.material_lot, r.blend_record_id, b.product_lot "
            "FROM viscosity_readings r JOIN blend_records b ON b.id = r.blend_record_id "
            "WHERE b.product_lot = 'APB26060101'"
        ).fetchone()
        assert row is not None
        assert row["lot_no"] == "APB26060101"       # 배합 product_lot
        assert row["viscosity"] == 385.5
        assert row["material_lot"] == "26060101"    # 사용한PB(PB 연계 키)
        assert row["blend_record_id"] is not None


def test_binder_import_is_idempotent(tmp_path):
    import src.config as cfg

    importlib.reload(cfg)
    from src.db import get_connection

    from src.db import init_db
    init_db()
    xlsx = tmp_path / "바인더.xlsx"
    _make_binder_file(xlsx, pb_prefix="2607")
    with get_connection() as conn:
        _seed_blend_record(conn, "APB26070101", "APB")

    import scripts.import_binder_viscosity as imp

    first = imp.import_binder([str(xlsx)])
    second = imp.import_binder([str(xlsx)])
    assert first["linked"] == 1
    assert first["archived"] == 2
    assert second["linked"] == 0
    assert second["archived"] == 0
    assert second["dup"] == 3   # 연계 1 + 과거 2 모두 재실행 시 중복


def test_cleanup_scoped_to_import_marker(tmp_path):
    """정리는 이 스크립트가 넣은 행(created_by 마커)만 지운다. 같은 형태라도
    사용자가 수동 등록한 PB 점도(created_by 마커 아님, blend_record_id NULL)는 보존한다."""
    import src.config as cfg

    importlib.reload(cfg)
    from src.db import get_connection, init_db, utc_now_text
    from src.services import viscosity_service

    init_db()
    import scripts.import_binder_viscosity as imp
    with get_connection() as conn:
        p = viscosity_service.ensure_product_by_code(conn, "APB", "APB", utc_now_text())
        # (1) 이전 버전 임포트 흔적 — created_by 마커 + 옛 형태 → 지워져야 함
        viscosity_service.add_reading(
            conn, product_id=p["id"], lot_no="24990001", viscosity=400.0,
            measured_date="2024-09-01", memo=None, recipe_material=None,
            material_lot="24990001", created_by=imp.CREATED_BY_MARKER,
            created_at=utc_now_text(),
        )
        # (2) 사용자가 수동 등록한 look-alike — 마커 아님(created_by NULL), 형태는 동일
        #     → 절대 지우면 안 됨
        viscosity_service.add_reading(
            conn, product_id=p["id"], lot_no="24990002", viscosity=402.0,
            measured_date="2024-09-02", memo=None, recipe_material=None,
            material_lot="24990002", created_by=None, created_at=utc_now_text(),
        )
        # (3) 배합 화면으로 등록한 정상 연계 점도 — 지우면 안 됨
        conn.execute(
            "INSERT INTO blend_records (product_lot, product_name, worker, work_date, "
            "total_amount, status, created_at) VALUES ('APB26050101','APB','h','2026-05-01',1000,'completed',?)",
            (utc_now_text(),),
        )
        rid = conn.execute("SELECT id FROM blend_records WHERE product_lot='APB26050101'").fetchone()["id"]
        viscosity_service.add_reading(
            conn, product_id=p["id"], lot_no="APB26050101", viscosity=380.0,
            measured_date="2026-05-01", memo=None, recipe_material="APB",
            material_lot="26050101", created_by="h", created_at=utc_now_text(),
            blend_record_id=rid,
        )
        conn.commit()

    xlsx = tmp_path / "바인더.xlsx"
    _make_binder_file(xlsx, pb_prefix="2699")   # 매칭 배합 없음 — 정리만 검증
    stats = imp.import_binder([str(xlsx)])

    assert stats["cleaned"] == 1   # 마커 행 1건만 삭제(수동 look-alike 제외)
    with get_connection() as conn:
        # (1) 마커 옛 방식은 사라지고
        assert conn.execute(
            "SELECT COUNT(*) FROM viscosity_readings WHERE lot_no='24990001'"
        ).fetchone()[0] == 0
        # (2) 사용자 수동 look-alike 는 생존
        assert conn.execute(
            "SELECT COUNT(*) FROM viscosity_readings WHERE lot_no='24990002'"
        ).fetchone()[0] == 1
        # (3) 정상 연계 점도는 보존
        assert conn.execute(
            "SELECT COUNT(*) FROM viscosity_readings WHERE lot_no='APB26050101'"
        ).fetchone()[0] == 1


def test_binder_without_record_is_archived(tmp_path):
    """배합 기록이 없는 과거 행은 건너뛰지 않고, 같은 바인더 반제품에 배합 미연계로
    점도만 보존한다 — lot_no=바인더+사용한PB, blend_record_id NULL, 측정일=엑셀 일자."""
    import src.config as cfg

    importlib.reload(cfg)
    from src.db import get_connection, init_db

    init_db()
    xlsx = tmp_path / "바인더.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("24년도 바인더 점도 기록")
    ws.append(["일자", "바인더", "사용한PB", "점도", "작업자"])
    ws.append(["8월29일", "APB", "24082902", 373.8, "이재석"])   # 배합 기록 없음(과거)
    wb.save(xlsx)

    import scripts.import_binder_viscosity as imp

    stats = imp.import_binder([str(xlsx)])
    assert stats["linked"] == 0
    assert stats["archived"] == 1

    with get_connection() as conn:
        row = conn.execute(
            "SELECT lot_no, viscosity, material_lot, blend_record_id, measured_date "
            "FROM viscosity_readings WHERE lot_no = 'APB24082902'"
        ).fetchone()
        assert row is not None
        assert row["viscosity"] == 373.8
        assert row["material_lot"] == "24082902"       # PB 상관 키 유지
        assert row["blend_record_id"] is None          # 배합 미연계
        assert row["measured_date"] == "2024-08-29"    # 시트 연도 + 엑셀 월일


def test_binder_normalize_rules():
    import scripts.import_binder_viscosity as imp

    assert imp._normalize_binder("APB(17)") == "APB17"
    assert imp._normalize_binder("CSBP") == "CSPB"
    assert imp._normalize_binder("APB(1)") == "APB"
    assert imp._normalize_binder("PM17") == "PM17"
    assert imp._normalize_binder("APB(TEST)") is None
