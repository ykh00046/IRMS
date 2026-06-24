"""운영 대시보드 보고서 Excel 검증."""

import io
import sqlite3

import openpyxl

from src.services import dashboard_export


def _db() -> sqlite3.Connection:
    con = sqlite3.connect(":memory:")
    con.row_factory = sqlite3.Row
    con.executescript(
        """
        CREATE TABLE recipes (
            id INTEGER PRIMARY KEY, status TEXT, completed_at TEXT, created_by TEXT
        );
        CREATE TABLE materials (id INTEGER PRIMARY KEY, name TEXT, category TEXT);
        CREATE TABLE recipe_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT, material_id INTEGER,
            value_weight REAL, measured_at TEXT, measured_by TEXT
        );
        """
    )
    con.execute("INSERT INTO materials (id,name,category) VALUES (1,'HEMA','모노머')")
    con.execute(
        "INSERT INTO recipes (id,status,completed_at,created_by) VALUES (1,'completed','2026-06-10T00:00:00Z','김도현')"
    )
    con.execute(
        "INSERT INTO recipe_items (material_id,value_weight,measured_at,measured_by) "
        "VALUES (1,100,'2026-06-10T00:00:00Z','김도현')"
    )
    con.commit()
    return con


def test_dashboard_excel_has_sections():
    con = _db()
    xb = dashboard_export.build_dashboard_excel(
        con, from_date="2026-06-01", to_date="2026-06-25",
        from_ts="2026-06-01T00:00:00Z", to_ts="2026-06-25T23:59:59Z",
    )
    assert xb[:2] == b"PK"
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active
    titles = [ws.cell(r, 1).value for r in range(1, 30) if ws.cell(r, 1).value]
    assert "운영 대시보드 보고서" in titles
    assert "[요약]" in titles
    assert "[자재 사용량 TOP 10]" in titles
    assert "[작업자 통계]" in titles
