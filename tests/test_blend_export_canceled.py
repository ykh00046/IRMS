"""GET /api/blend/records/export-all?include_canceled= — 취소 포함 정합 테스트.

/status 화면의 '취소 포함' 체크와 Excel 이 어긋나던 회귀: export-all 은 항상
취소를 제외했다. include_canceled 파라미터로 화면과 맞추고, 취소 행이 섞이면
구분되도록 '상태' 열 값을 한글(완료/취소)로 표기한다.

패턴은 tests/test_blend.py 의 importlib.reload + TestClient + admin/admin +
worker 세션 + uuid 제품명.
"""

from __future__ import annotations

import io
import importlib
import uuid


def _reload_app():
    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    return TestClient(mainmod.app)


def _create_record(client, csrf, prod, worker):
    """배합 기록 1건 생성 — worker 세션 필요(test_blend.py 패턴)."""
    created = client.post("/api/blend/records", json={
        "product_name": prod, "worker": worker, "work_date": "2026-07-01",
        "total_amount": 100, "scale": "M-65",
        "details": [
            {"material_name": "A", "ratio": 60, "theory_amount": 60, "actual_amount": 60, "material_lot": "LA"},
            {"material_name": "B", "ratio": 40, "theory_amount": 40, "actual_amount": 40, "material_lot": "LB"},
        ],
    }, headers=csrf)
    assert created.status_code == 200, created.text
    return created.json()


def _export_lots(client, include_canceled):
    url = "/api/blend/records/export-all"
    if include_canceled:
        url += "?include_canceled=1"
    res = client.get(url)
    assert res.status_code == 200, res.text
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content))
    ws = wb["배합기록"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    lot_idx = header.index("제품LOT")
    status_idx = header.index("상태")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({"lot": r[lot_idx], "status": r[status_idx]})
    return rows


def test_export_all_respects_include_canceled():
    client = _reload_app()
    worker = "취소작업" + uuid.uuid4().hex[:5]

    def csrf_headers():
        tok = client.cookies.get("csrftoken")
        return {"x-csrftoken": tok} if tok else {}

    client.get("/api/blend/records")  # csrf 쿠키 확보
    client.post("/api/workers", json={"name": worker}, headers=csrf_headers())
    client.post("/api/blend/session/login", json={"worker": worker}, headers=csrf_headers())

    prod = "CX" + uuid.uuid4().hex[:5]
    rec1 = _create_record(client, csrf_headers(), prod, worker)
    rec2 = _create_record(client, csrf_headers(), prod, worker)
    lot1, lot2 = rec1["product_lot"], rec2["product_lot"]

    # 1건 soft 취소(책임자). 현장 세션만으로는 403 이므로 관리 로그인.
    client.post("/api/auth/management-login", json={"username": "admin", "password": "admin"})
    canceled = client.delete(f"/api/blend/records/{rec2['id']}", headers=csrf_headers())
    assert canceled.status_code == 200, canceled.text

    # 기본 — 취소 제외: lot2(canceled) 없고 lot1(completed) 있음
    default_rows = _export_lots(client, include_canceled=False)
    default_lots = {r["lot"] for r in default_rows}
    assert lot1 in default_lots, default_lots
    assert lot2 not in default_lots, default_lots
    # 상태 열은 한글
    by_lot = {r["lot"]: r["status"] for r in default_rows}
    assert by_lot[lot1] == "완료", by_lot

    # include_canceled=1 — 취소 포함: lot2 도 있고 상태 '취소'
    incl_rows = _export_lots(client, include_canceled=True)
    incl_lots = {r["lot"] for r in incl_rows}
    assert lot1 in incl_lots, incl_lots
    assert lot2 in incl_lots, incl_lots
    by_lot2 = {r["lot"]: r["status"] for r in incl_rows}
    assert by_lot2[lot1] == "완료", by_lot2
    assert by_lot2[lot2] == "취소", by_lot2
