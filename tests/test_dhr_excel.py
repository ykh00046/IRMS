"""원료배합일지(DHR) 공식 양식 출력 검증."""

import io
import json

import openpyxl

from src.services import dhr_excel


def _sample_record():
    return {
        "product_lot": "제품A260625",
        "worker": "홍길동",
        "work_date": "2026-06-25",
        "work_time": "10:00:00",
        "scale": "M-65",
        "total_amount": 1000,
        "details": [
            {"material_name": "HEMA", "material_lot": "MN-101", "ratio": 71.43,
             "theory_amount": 714.3, "actual_amount": 714.3},
            {"material_name": "NVP", "material_lot": "MN-102", "ratio": 28.57,
             "theory_amount": 285.7, "actual_amount": 285.7},
        ],
    }


def test_official_dhr_form_fills_template():
    xb = dhr_excel.build_official_dhr_xlsx(_sample_record())
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active

    # 양식 제목·헤더 보존
    assert ws["A1"].value == "원 료 배 합 일 지"
    assert ws["C5"].value == "배합원료명"

    # 메타 채움
    assert "2026-06-25" in ws["A3"].value
    assert "홍길동" in ws["C3"].value
    assert "M-65" in ws["A4"].value

    # 제품 LOT + 총량/100
    assert ws["A6"].value == "제품A260625"
    assert ws["B6"].value == 10  # 1000 / 100

    # 자재 데이터(6행~)
    assert ws["C6"].value == "HEMA"
    assert ws["D6"].value == "MN-101"
    assert ws["E6"].value == 71.43
    assert ws["F6"].value == 714.3
    assert ws["G6"].value == 714.3
    assert ws["C7"].value == "NVP"

    # A/B 데이터 행 병합
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "A6:A7" in merged
    assert "B6:B7" in merged


def test_official_dhr_form_handles_missing_optionals():
    rec = _sample_record()
    rec["scale"] = None
    rec["work_time"] = None
    xb = dhr_excel.build_official_dhr_xlsx(rec, include_work_time=False)
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active
    assert ws["A1"].value == "원 료 배 합 일 지"
    assert ws["A6"].value == "제품A260625"


def _all_text(xb: bytes) -> str:
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active
    return "\n".join(
        c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    )


def test_official_dhr_is_plain_even_with_rescale_history():
    """배합일지는 외부 제출 문서 — 증량 승인/부재 이력은 문서에 싣지 않는다(2026-08-27 재지시).

    2026-07 감사 수정(GAP-5)이 비고 줄로 넣었다가 첫 실제 출력에서 걸러졌다. 이력은
    /status 상세(rescaleBlock)에서만 보인다.
    """
    rec = _sample_record()
    rec["rescale_count"] = 2
    rec["rescale_unacked"] = 1
    rec["rescale_events_json"] = json.dumps([
        {"before_total": 1000, "after_total": 1050, "approver": "홍길동"},
        {"before_total": 1050, "after_total": 1100, "absence_reason": "야간 단독"},
    ], ensure_ascii=False)
    joined = _all_text(dhr_excel.build_official_dhr_xlsx(rec))
    # '승인'·작업자명은 양식 결재칸·작업자 칸에 정당하게 있으므로 보지 않는다.
    for banned in ("증량", "부재", "야간 단독", "1000→1050", "1050→1100"):
        assert banned not in joined, f"배합일지에 통제 표식이 실렸다: {banned!r}"


def test_official_dhr_is_plain_even_with_manual_absence():
    """수기 입력을 책임자 부재로 진행한 사정도 문서에는 없다 — 조회 화면·트레이·감사로그 전용."""
    rec = _sample_record()
    rec["manual_entry"] = True
    rec["manual_unacked"] = 1
    rec["manual_absence_reason"] = "야간 근무 · 책임자 부재"
    joined = _all_text(dhr_excel.build_official_dhr_xlsx(rec))
    for banned in ("수기", "부재", "야간 근무", "수동"):
        assert banned not in joined, f"배합일지에 통제 표식이 실렸다: {banned!r}"


def test_official_dhr_is_plain_even_when_bulk_regenerated():
    """일괄 재생성 기록의 배합일지도 표식 없이 나간다 — 그 기록은 애초에 문서용으로 만든 것."""
    rec = _sample_record()
    rec["is_bulk_regenerated"] = True
    joined = _all_text(dhr_excel.build_official_dhr_xlsx(rec))
    assert "일괄 재생성" not in joined


def test_official_dhr_has_no_note_row_for_normal_record():
    """정상 기록은 표 아래에 아무 줄도 붙지 않는다 — 인쇄 영역이 표 끝에서 끝난다."""
    xb = dhr_excel.build_official_dhr_xlsx(_sample_record())
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active
    # openpyxl 은 시트명·$ 를 붙여 돌려준다("'시트'!$A$1:$G$7") — 끝만 본다.
    assert ws.print_area.replace("$", "").endswith("A1:G7"), ws.print_area


def test_rescale_summary_helper_is_gone():
    """통제 표식을 문서에 실을 도우미가 남아 있으면 누군가 다시 부른다."""
    assert not hasattr(dhr_excel, "rescale_summary_line")


def test_official_dhr_marks_canceled_record():
    """취소된 기록을 단건 출력하면 비고 영역에 '(취소된 기록)' 표식이 실린다(POLISH-7b)."""
    rec = _sample_record()
    rec["status"] = "canceled"
    xb = dhr_excel.build_official_dhr_xlsx(rec)
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active
    joined = "\n".join(
        c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    )
    assert "(취소된 기록)" in joined


def test_official_dhr_no_canceled_marker_when_completed():
    """완료 기록(status=completed)이면 취소 표식이 생기지 않는다(회귀 가드)."""
    rec = _sample_record()
    rec["status"] = "completed"
    xb = dhr_excel.build_official_dhr_xlsx(rec)
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active
    joined = "\n".join(
        c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    )
    assert "취소된 기록" not in joined


def test_official_dhr_marks_signature_failure():
    """서명 합성 실패(sign_failed) 시 결재칸에 표식을 남긴다 — 무언의 미서명 출력 금지(POLISH-6)."""
    xb = dhr_excel.build_official_dhr_xlsx(_sample_record(), sign_failed=True)
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active
    assert ws["G2"].value == "(서명 합성 실패)"


def test_dhr_loss_comp_included_in_theory_no_extra_column():
    """투입 로스 보정(2026-08-05): 보정분은 theory_amount 에 이미 포함돼 그대로 출력되고,
    보정 관련 별도 열/헤더는 추가되지 않는다(배합일지·Excel 출력은 변경 없이 501 로).
    loss_comp_g 메타가 detail 에 있어도 양식 열 구조는 동일해야 한다."""
    record = _sample_record()
    # 첫 자재에 보정 스냅샷이 있어도(저장 시 theory_amount 에 이미 반영됨) 출력은 그대로.
    record["details"][0]["loss_comp_g"] = 1.0
    xb = dhr_excel.build_official_dhr_xlsx(record)
    ws = openpyxl.load_workbook(io.BytesIO(xb)).active
    # 기존 양식 열 구조 보존 — 보정 전용 열이 새로 생기지 않는다(헤더 변화 없음).
    assert ws["C5"].value == "배합원료명"  # 헤더 그대로
    # theory_amount(F열) 가 detail 의 값(714.3) 을 그대로 출력 — 보정이 이미 포함된 값.
    assert ws["F6"].value == 714.3
    # 보정량 자체가 별도 셀로 찍히지 않는다(메타는 출력에 등장하지 않음).
    # H열(비고 영역 이후) 에 '1.0' 같은 보정값이 단독으로 들어가지 않음을 확인.
    for row in range(6, 8):
        val = ws[f"H{row}"].value
        assert val != 1.0, "보정량이 별도 열로 출력되면 안 된다 — theory_amount 에 포함돼야"
