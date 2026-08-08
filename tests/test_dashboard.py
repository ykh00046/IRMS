"""운영 대시보드(배합·점도 기반 재구축) 검증.

- 단위: dashboard_export.build_dashboard_excel 이 배합/점도 스키마로 생성되는지
- 라우트: /api/dashboard/* 신규 엔드포인트가 무로그인으로 응답하고 키가 맞는지
"""

from __future__ import annotations

import io
import sqlite3

from openpyxl import load_workbook

from src.services import blend_service as bs
from src.services import dashboard_export


def _make_db() -> sqlite3.Connection:
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    connection.executescript(
        """
        CREATE TABLE blend_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_lot TEXT NOT NULL, recipe_id INTEGER, product_name TEXT NOT NULL,
            ink_name TEXT, position TEXT, worker TEXT NOT NULL, work_date TEXT NOT NULL,
            work_time TEXT, total_amount REAL NOT NULL, scale TEXT,
            status TEXT NOT NULL DEFAULT 'completed', note TEXT, reactor INTEGER,
            is_bulk_regenerated INTEGER NOT NULL DEFAULT 0,
            reviewed_by TEXT, reviewed_at TEXT, approved_by TEXT, approved_at TEXT,
            worker_sign TEXT, reviewed_sign TEXT, approved_sign TEXT,
            created_by TEXT, created_at TEXT NOT NULL, updated_at TEXT
        );
        CREATE TABLE viscosity_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT UNIQUE, name TEXT,
            target REAL, lower_limit REAL, upper_limit REAL, sigma_k REAL NOT NULL DEFAULT 3,
            rpm REAL, temperature REAL, remind_daily INTEGER NOT NULL DEFAULT 0,
            use_reactor INTEGER NOT NULL DEFAULT 0, is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );
        CREATE TABLE viscosity_readings (
            id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER NOT NULL,
            lot_no TEXT NOT NULL, viscosity REAL NOT NULL, measured_date TEXT,
            memo TEXT, recipe_material TEXT, material_lot TEXT, reactor INTEGER,
            created_by TEXT, created_at TEXT NOT NULL, blend_record_id INTEGER,
            excluded INTEGER NOT NULL DEFAULT 0, exclude_reason TEXT,
            excluded_by TEXT, excluded_at TEXT
        );
        """
    )
    connection.execute(
        "INSERT INTO blend_records (product_lot, product_name, worker, work_date, "
        "total_amount, status, approved_by, created_at) "
        "VALUES ('PB26070101', 'PB', '홍길동', '2026-07-01', 17400, 'completed', NULL, '2026-07-01')"
    )
    connection.execute(
        "INSERT INTO blend_records (product_lot, product_name, worker, work_date, "
        "total_amount, status, approved_by, created_at) "
        "VALUES ('PB26070102', 'PB', '홍길동', '2026-07-01', 17400, 'canceled', NULL, '2026-07-01')"
    )
    connection.execute(
        "INSERT INTO viscosity_products (code, name, remind_daily, created_at) "
        "VALUES ('PB', 'PB', 1, '2026-01-01')"
    )
    connection.commit()
    return connection


def test_dashboard_excel_builds_from_blend_records():
    conn = _make_db()
    xlsx = dashboard_export.build_dashboard_excel(
        conn, from_date="2026-07-01", to_date="2026-07-01"
    )
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    # 이 보고서는 운영 스냅샷만 담는다 — 반제품 TOP·작업자 통계는 배합 분석 리포트와
    # 같은 표라서 뺐고(2026-08-08), '결재 대기'는 결재 미사용으로 죽은 값이었다.
    assert "[요약]" in values
    assert "배합 건수" in values
    assert "[점도 현황 (최신 연도 기준)]" in values
    assert "[반제품별 배합 TOP 10]" not in values
    assert "[작업자 통계]" not in values
    assert "결재 대기(전체)" not in values
    idx = values.index("배합 건수")
    assert values[idx + 1] == 1  # canceled 제외


def test_dashboard_routes_respond_without_login():
    import importlib

    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    client = TestClient(mainmod.app)

    summary = client.get("/api/dashboard/summary")
    assert summary.status_code == 200
    body = summary.json()
    for key in (
        "blend_count", "total_weight_g", "product_count", "worker_count",
        "viscosity_anomaly", "viscosity_due_today",
    ):
        assert key in body
    # 결재 대기(approval_pending)는 죽은 값이라 페이로드에서 제거됨(2026-07-23). 회귀 가드.
    assert "approval_pending" not in body

    trend = client.get("/api/dashboard/trend?from=2026-07-01&to=2026-07-03")
    assert trend.status_code == 200
    assert len(trend.json()["points"]) == 3

    assert client.get("/api/dashboard/recent").status_code == 200

    # 반제품 TOP·작업자별 실적은 배합 분석(/insight)으로 일원화하며 제거했다(2026-08-08)
    # — 같은 표가 양쪽에 있었고 세는 기준마저 서로 달랐다. 되살아나지 않는지 확인한다.
    assert client.get("/api/dashboard/products").status_code == 404
    assert client.get("/api/dashboard/workers").status_code == 404

    # 잘못된 기간은 400
    assert client.get("/api/dashboard/summary?from=2026-07-05&to=2026-07-01").status_code == 400


# ── 분석 집계 상한(no silent truncation) ─────────────────────────────

def _make_blend_db() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.executescript(
        """
        CREATE TABLE blend_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_lot TEXT NOT NULL, product_name TEXT NOT NULL,
            worker TEXT NOT NULL, work_date TEXT NOT NULL, total_amount REAL NOT NULL,
            status TEXT NOT NULL DEFAULT 'completed',
            is_bulk_regenerated INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE blend_details (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            blend_record_id INTEGER NOT NULL, material_code TEXT,
            material_name TEXT NOT NULL, material_lot TEXT,
            ratio REAL, theory_amount REAL, actual_amount REAL,
            sequence_order INTEGER NOT NULL DEFAULT 0
        );
        """
    )
    rid = int(
        conn.execute(
            "INSERT INTO blend_records (product_lot, product_name, worker, work_date, total_amount) "
            "VALUES ('PB26070101', 'PB', '홍길동', '2026-07-01', 300)"
        ).lastrowid
    )
    for i, name in enumerate(("A", "B", "C")):
        conn.execute(
            "INSERT INTO blend_details (blend_record_id, material_code, material_name, "
            "theory_amount, actual_amount, sequence_order) VALUES (?, ?, ?, 100, 100, ?)",
            (rid, name, name, i),
        )
    conn.commit()
    return conn


def test_material_usage_periods_truncation_flag(monkeypatch):
    conn = _make_blend_db()
    # 정상: 상한 미도달 → truncated False, 전체 반환
    res = bs.material_usage_periods(conn, start_date="2026-07-01", end_date="2026-07-31")
    assert res["truncated"] is False
    assert res["total_item_count"] == 3
    assert len(res["items"]) == 3

    # 상한을 1로 낮추면 잘리고 표면화(조용한 절단 금지)
    monkeypatch.setattr(bs, "_MATERIAL_USAGE_MAX_ITEMS", 1)
    res2 = bs.material_usage_periods(conn, start_date="2026-07-01", end_date="2026-07-31")
    assert res2["truncated"] is True
    assert res2["total_item_count"] == 3
    assert len(res2["items"]) == 1


def test_batch_details_truncation_flag():
    conn = _make_blend_db()
    # 여유 있는 limit → 상한 미도달
    res = bs.batch_details(conn, limit=1000)
    assert res["truncated"] is False
    assert res["total"] == 3

    # limit 이 결과 수에 걸리면 truncated 로 표면화
    res2 = bs.batch_details(conn, limit=2)
    assert res2["truncated"] is True
    assert res2["limit"] == 2
    assert res2["total"] == 2


def test_dashboard_attention_is_independent_of_the_period_filter():
    """'지금 조치'는 기간 선택과 무관해야 한다.

    예전에는 '오늘 점도 미입력'이 '배합 건수'와 한 줄에 나란히 있어, 기간을 30일로
    바꿔도 안 변하는 값과 바뀌는 값이 같은 기준으로 읽혔다(2026-08-08 분리).
    """
    import importlib

    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    client = TestClient(mainmod.app)
    res = client.get("/api/dashboard/attention")
    assert res.status_code == 200
    body = res.json()
    for key in (
        "today", "today_blend_count", "last_blend_at",
        "viscosity_due_today", "unacked_count", "material_lot_file",
    ):
        assert key in body, key
    assert isinstance(body["viscosity_due_today"], list)
    assert isinstance(body["material_lot_file"], dict)
    for key in ("file_name", "file_date", "found", "stale_days"):
        assert key in body["material_lot_file"], key

    # 기간 파라미터를 붙여도 무시된다 — 이 엔드포인트에 기간 개념이 없어야 한다.
    same = client.get("/api/dashboard/attention?from=2000-01-01&to=2000-01-02")
    assert same.status_code == 200
    assert same.json() == body


def test_material_lot_file_summary_does_not_open_the_workbook(tmp_path, monkeypatch):
    """대시보드가 700행짜리 LOT 목록을 끌어오지 않고 파일명만으로 경과일을 낸다."""
    from openpyxl import Workbook

    from src.services import erp_lot_service

    erp_dir = tmp_path / "erp"
    erp_dir.mkdir()
    wb = Workbook()
    wb.active.append(["창고", "구분", "품목코드", "품목명"])
    wb.save(erp_dir / "ERP_2026-08-02.xlsx")
    monkeypatch.setenv("IRMS_ERP_EXCEL_DIR", str(erp_dir))

    def _boom(*_a, **_k):  # 워크북을 열면 실패시킨다
        raise AssertionError("latest_file_summary 는 엑셀을 열지 않아야 한다")

    monkeypatch.setattr(erp_lot_service, "load_workbook", _boom, raising=False)
    summary = erp_lot_service.latest_file_summary()
    assert summary == {
        "file_name": "ERP_2026-08-02.xlsx",
        "file_date": "2026-08-02",
        "found": True,
    }


def test_material_lot_file_summary_when_no_file(tmp_path, monkeypatch):
    monkeypatch.setenv("IRMS_ERP_EXCEL_DIR", str(tmp_path))
    from src.services import erp_lot_service

    assert erp_lot_service.latest_file_summary() == {
        "file_name": None, "file_date": None, "found": False,
    }
