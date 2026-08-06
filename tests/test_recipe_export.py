"""GET /api/recipes/export — 레시피 전체 Excel 내보내기 라우트 테스트.

패턴은 tests/test_blend.py 의 importlib.reload(cfg/mainmod) + TestClient +
admin/admin 로그인 + uuid 제품명 시드를 따른다. 책임자 전용 통제(미로그인 401/403),
200 + xlsx content-type, 시트 2개, 시드한 레시피·자재명이 각 시트에 보이는지 단언.
"""

from __future__ import annotations

import io
import importlib
import uuid


def _reload_app():
    import src.config as cfg
    import src.main as mainmod

    importlib.reload(cfg)
    importlib.reload(mainmod)
    from fastapi.testclient import TestClient

    return TestClient(mainmod.app)


def _seed_recipe(client, headers, prod, materials, revision_of=None):
    """TSV 로 레시피 한 건을 등록하고 생성된 id 를 돌려준다."""
    header = "반제품명\t" + "\t".join(materials.keys())
    row = f"{prod}\t" + "\t".join(str(v) for v in materials.values())
    raw = header + "\n" + row
    body = {"raw_text": raw, "force": True}
    if revision_of is not None:
        body["revision_of"] = revision_of
    res = client.post(
        "/api/recipes/import",
        json=body,
        headers=headers,
    )
    assert res.status_code == 200, res.text
    return res.json()["created_ids"][0], list(materials.keys())


def test_recipe_export_requires_manager():
    """미로그인(또는 현장 권한) 접근은 401 또는 403."""
    client = _reload_app()
    res = client.get("/api/recipes/export")
    assert res.status_code in (401, 403), res.text


def test_recipe_export_returns_two_sheets_with_seed_data():
    """책임자 로그인 후 200 + xlsx, 시트 2개 이상, 시드한 반제품명/자재명이 각 시트에 존재."""
    client = _reload_app()
    client.post("/api/auth/management-login", json={"username": "admin", "password": "admin"})
    tok = client.cookies.get("csrftoken")
    headers = {"x-csrftoken": tok} if tok else {}

    prod = "RE" + uuid.uuid4().hex[:6]
    materials = {"MatX": 60, "MatY": 40}
    _rid, material_names = _seed_recipe(client, headers, prod, materials)

    res = client.get("/api/recipes/export")
    assert res.status_code == 200, res.text
    assert "spreadsheetml" in res.headers.get("content-type", ""), res.headers.get("content-type")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content))
    assert wb.sheetnames[:2] == ["레시피 목록", "자재 구성"], wb.sheetnames

    ws1 = wb["레시피 목록"]
    names1 = [r[1].value for r in ws1.iter_rows(min_row=2)]
    assert prod in names1, names1

    ws2 = wb["자재 구성"]
    names2 = [r[3].value for r in ws2.iter_rows(min_row=2)]
    for mname in material_names:
        assert mname in names2, (mname, names2)


def test_recipe_export_lists_only_current_tip_after_revision():
    """개정하면 시트1엔 반제품명이 1행만(현재판), 옛 개정판은 '개정 이력' 시트로.

    종전엔 recipes 전 행을 쏟아 개정 수만큼 같은 반제품명이 반복됐다(현장 지적
    2026-08-06 — '중복으로 보이는 품목들'). 화면 목록과 같은 tip 규칙으로 고정.
    """
    client = _reload_app()
    client.post("/api/auth/management-login", json={"username": "admin", "password": "admin"})
    tok = client.cookies.get("csrftoken")
    headers = {"x-csrftoken": tok} if tok else {}

    prod = "RV" + uuid.uuid4().hex[:6]
    old_id, _ = _seed_recipe(client, headers, prod, {"MatA": 70, "MatB": 30})
    new_id, _ = _seed_recipe(
        client, headers, prod, {"MatA": 65, "MatB": 35}, revision_of=old_id
    )  # 개정
    assert new_id != old_id

    res = client.get("/api/recipes/export")
    assert res.status_code == 200, res.text
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content))

    ws1 = wb["레시피 목록"]
    rows1 = [(r[0].value, r[1].value) for r in ws1.iter_rows(min_row=2)]
    matching = [row for row in rows1 if row[1] == prod]
    assert len(matching) == 1, f"시트1에 현재판 1행만 있어야 한다: {matching}"
    assert matching[0][0] == new_id, matching

    # 자재 구성도 현재판 것만 — 옛 개정판 id 행이 없어야 한다.
    ws2 = wb["자재 구성"]
    ids2 = {r[0].value for r in ws2.iter_rows(min_row=2)}
    assert new_id in ids2 and old_id not in ids2, ids2

    # 옛 개정판은 '개정 이력' 시트에 보존된다.
    assert "개정 이력" in wb.sheetnames, wb.sheetnames
    ws3 = wb["개정 이력"]
    hist_ids = {r[0].value for r in ws3.iter_rows(min_row=2)}
    assert old_id in hist_ids, hist_ids
