"""ERP 품목 마스터(code*.xlsx) → item_code_master 임포트 도구.

item-code P1 의 마스터 적재기. materials.code / recipes.product_code 부여는
매칭 단계(P2)에서 별도 스크립트가 담당 — 이 스크립트는 마스터만 채운다.

소스 엑셀(루트, 읽기 전용 — 커밋·이동 금지):
  - code.xlsx  : 전 품목 마스터(품목코드/품목명/규격/기준단위/LOT/품목구분/대분류/중분류)
                 대분류=='원자재' 행만 kind='material' 로 적재(포장재/소모품 등은 skip).
  - code2~4.xlsx : 반제품(품목코드/품명/규격/단위/회계분류/제품구분/...).
                 전 행을 kind='product' 로 적재. category_hint 는 제품구분에서 매핑.

코드 정규화: strip + upper (bc0001 → BC0001). 이름은 strip 만.
빈 코드/빈 이름 행은 skip. 같은 code 재임포트 시 갱신(upsert, imported_at 갱신).

사용:
  python tools/import_item_codes.py --material code.xlsx
  python tools/import_item_codes.py --product code2.xlsx --product code3.xlsx --product code4.xlsx
  python tools/import_item_codes.py --material code.xlsx --product code2.xlsx --dry-run
  python tools/import_item_codes.py --material code.xlsx --product code2.xlsx \
      --product code3.xlsx --product code4.xlsx --retire-missing
      # 이번 파일들에 없는 기존 코드를 폐기(retired) 표시 - manual 행 제외, 재등장 시 부활
  python tools/import_item_codes.py --material code.xlsx --db 경로/rehearsal.db
      # 비관례 파일명도 그 파일에 직접 스키마(item_code_master 포함) 적용 후 임포트
"""

import argparse
import os
import sqlite3
import sys
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.db.time_utils import utc_now_text  # noqa: E402  (sys.path 조정 후 import)
# 원재료 코드 계열 상수 — 단일 진실 원천(src.services.erp_lot_service)에서 가져온다.
# 마스터 임포트의 원자재 필터와 check_lot 의 LOT 검사 대상 판정이 같은 기준을 써야
# 한쪽에서만 원자재로 분류되는 어긋남이 생기지 않는다.
from src.services.erp_lot_service import RAW_MATERIAL_CODE_PREFIXES  # noqa: E402
# DB 연결 + 스키마 마이그레이션(item_code_master 보장) 공용 헬퍼. tools/ 내 상호 import
# 관례(Tests 도 from tools.match_item_codes 로 쓴다)에 따른다. 과거 이 스크립트는
# init_db() 가 *관례 DB(irms.db)* 에만 스키마를 잡고 --db 비관례 파일명은 별도 연결(스키마
# 누락)을 열어 — 대상에 item_code_master 가 없거나 관례 DB 가 오염되는 버그가 있었다.
# _open_target_db 는 대상 연결에 직접 apply_schema_migrations 를 적용해 이를 고친다.
from tools.match_item_codes import _open_target_db  # noqa: E402


# 제품구분(엑셀 원문) → IRMS 레시피 분류(category_hint) 매핑.
# code2~4 의 '제품구분' 열은 '잉크코드'/'합성코드'/'약품코드' 이며, IRMS 분류는
# '잉크'/'합성'/'약품' 이다(용수는 ERP 에 없음 — IRMS 자체 분류). 그 외 값은 원문 유지.
_PRODUCT_CATEGORY_HINT = {
    "잉크코드": "잉크",
    "합성코드": "합성",
    "약품코드": "약품",
}


def _norm_code(value) -> str | None:
    """품목코드 정규화: strip + upper. 빈 값 → None."""
    if value is None:
        return None
    s = str(value).strip()
    return s.upper() if s else None


def _norm_text(value) -> str | None:
    """일반 텍스트 정규화: strip 만. 빈 값 → None."""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _read_xlsx(path: str):
    """엑셀 첫 시트를 (1-based row, 1-based col) cell 접근 객체로 반환."""
    import openpyxl

    return openpyxl.load_workbook(path, data_only=True).active


def _upsert_master(conn: sqlite3.Connection, *, code: str, name: str, spec, unit,
                   kind: str, category_hint, source: str, imported_at: str) -> None:
    """item_code_master upsert — 같은 code 면 갱신(imported_at 포함).

    재임포트에 등장한 코드는 status 를 active 로 되돌린다(부활) — 한 번 폐기됐던
    코드가 ERP 에 다시 나타나면 현행 코드다. retired_at 도 함께 지운다.
    """
    conn.execute(
        """
        INSERT INTO item_code_master
            (code, name, spec, unit, kind, category_hint, source, imported_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(code) DO UPDATE SET
            name = excluded.name,
            spec = excluded.spec,
            unit = excluded.unit,
            kind = excluded.kind,
            category_hint = excluded.category_hint,
            source = excluded.source,
            imported_at = excluded.imported_at,
            status = 'active',
            retired_at = NULL
        """,
        (code, name, _norm_text(spec), _norm_text(unit), kind,
         _norm_text(category_hint), source, imported_at),
    )


def retire_missing_codes(
    conn: sqlite3.Connection, *, kind: str, imported_codes: set[str],
    now: str, dry_run: bool = False,
) -> list[str]:
    """이번 임포트에 등장하지 않은 같은 kind 의 active 코드를 폐기(retired) 표시.

    - 화면 수동 입력(source='manual') 행은 건드리지 않는다 — ERP 파일이 근거가 아니다.
    - imported_codes 가 비어 있으면 아무것도 하지 않는다(빈 파일 사고 방어 —
      전 코드 일괄 폐기를 막는다).
    - 폐기는 삭제가 아니다: 코드·이력은 남고 status 만 바뀐다. 재등장하면
      _upsert_master 가 active 로 되살린다.
    반환: 폐기(예정) 코드 목록.
    """
    if not imported_codes:
        return []
    rows = conn.execute(
        "SELECT code FROM item_code_master "
        "WHERE kind = ? AND COALESCE(source, '') != 'manual' "
        "  AND COALESCE(status, 'active') = 'active'",
        (kind,),
    ).fetchall()
    to_retire = [r["code"] for r in rows if r["code"] not in imported_codes]
    if to_retire and not dry_run:
        conn.executemany(
            "UPDATE item_code_master SET status = 'retired', retired_at = ? "
            "WHERE code = ?",
            [(now, c) for c in to_retire],
        )
        conn.commit()
    return to_retire


# 배합 원료 계열 코드 prefix. 운영 스냅샷 리허설(2026-07-16) 결과 실제 배합 자재 20종이
# '소모품' 대분류에 있어(예: AIBN=AC0006, Dibutyltin dilaurate=AS0052) 대분류 필터로는
# 누락된다 — 코드 prefix(AS/AC/AH/AW = 원자재 117 + 소모품 83 = 200행)가 정확한 기준.
# AA(상품)·CB/CL/SP(포장재·소모품 잡류)는 배합과 무관 → 제외.
# 상수 자체는 src.services.erp_lot_service.RAW_MATERIAL_CODE_PREFIXES 가 단일 진실 원천.
# 별칭은 이 파일 안의 기존 참조(import_material_master)와 테스트가 그대로 동작하도록.
MATERIAL_CODE_PREFIXES = RAW_MATERIAL_CODE_PREFIXES


def import_material_master(
    conn: sqlite3.Connection, path: str, *, source: str = "code",
    dry_run: bool = False,
) -> dict:
    """code.xlsx 형식 → item_code_master (kind='material').

    코드 prefix 가 MATERIAL_CODE_PREFIXES(AS/AC/AH/AW)인 행만 적재(배합 원료 계열).
    category_hint = 대분류/중분류. 반환: {read, imported, skipped, skipped_non_material}.
    """
    ws = _read_xlsx(path)
    now = utc_now_text()
    read = imported = skipped_non_material = skipped_empty = 0
    codes: set[str] = set()   # 이번 파일에 등장한 코드(--retire-missing 대조용)

    for r in range(2, ws.max_row + 1):  # 1행은 헤더
        code_raw = ws.cell(r, 1).value
        name_raw = ws.cell(r, 2).value
        spec = ws.cell(r, 3).value
        unit = ws.cell(r, 4).value
        dae = ws.cell(r, 7).value   # 대분류
        joong = ws.cell(r, 8).value  # 중분류
        read += 1

        code = _norm_code(code_raw)
        # 배합 원료 계열 prefix 만. 포장재/상품/기타 소모품 잡류는 배합과 무관 → skip.
        if not code or not code.startswith(MATERIAL_CODE_PREFIXES):
            skipped_non_material += 1
            continue

        name = _norm_text(name_raw)
        if not name:
            skipped_empty += 1
            continue

        # 대분류/중분류를 함께 보존(원자재 vs 소모품 구분이 화면 안내에 유용)
        hint = "/".join(str(x).strip() for x in (dae, joong) if x and str(x).strip())
        if not dry_run:
            _upsert_master(
                conn, code=code, name=name, spec=spec, unit=unit,
                kind="material", category_hint=hint or None, source=source, imported_at=now,
            )
        imported += 1
        codes.add(code)

    if not dry_run:
        conn.commit()
    return {
        "read": read,
        "imported": imported,
        "skipped_non_material": skipped_non_material,
        "skipped_empty": skipped_empty,
        "codes": codes,
    }


def import_product_master(
    conn: sqlite3.Connection, path: str, *, source: str = "code2",
    dry_run: bool = False,
) -> dict:
    """code2~4.xlsx 형식 → item_code_master (kind='product').

    전 행 적재. category_hint = 제품구분에서 매핑(잉크코드→잉크 등, 그 외 원문).
    반환: {read, imported, skipped_empty, category_breakdown}.
    """
    ws = _read_xlsx(path)
    now = utc_now_text()
    read = imported = skipped_empty = 0
    cats: Counter = Counter()
    codes: set[str] = set()   # 이번 파일에 등장한 코드(--retire-missing 대조용)

    for r in range(2, ws.max_row + 1):  # 1행은 헤더
        code_raw = ws.cell(r, 1).value
        name_raw = ws.cell(r, 2).value
        spec = ws.cell(r, 3).value
        unit = ws.cell(r, 4).value
        prod_gubun = ws.cell(r, 6).value  # 제품구분
        read += 1

        code = _norm_code(code_raw)
        name = _norm_text(name_raw)
        if not code or not name:
            skipped_empty += 1
            continue

        # 제품구분 → IRMS 분류 매핑. 매핑표에 없으면 원문 유지(추적 가능).
        gubun_norm = _norm_text(prod_gubun) or ""
        category_hint = _PRODUCT_CATEGORY_HINT.get(gubun_norm, gubun_norm or None)
        cats[str(category_hint)] += 1

        if not dry_run:
            _upsert_master(
                conn, code=code, name=name, spec=spec, unit=unit,
                kind="product", category_hint=category_hint,
                source=source, imported_at=now,
            )
        imported += 1
        codes.add(code)

    if not dry_run:
        conn.commit()
    return {
        "read": read,
        "imported": imported,
        "skipped_empty": skipped_empty,
        "category_breakdown": dict(cats),
        "codes": codes,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="ERP 품목 마스터 → item_code_master 임포트")
    ap.add_argument("--material", action="append", default=[],
                    help="code.xlsx 형식 경로(원자재만). 복수 지정 가능.")
    ap.add_argument("--product", action="append", default=[],
                    help="code2~4 형식 경로(반제품 전체). 복수 지정 가능.")
    ap.add_argument("--db", default=None,
                    help="대상 DB 경로(기본: IRMS_DATA_DIR 의 개발 DB)")
    ap.add_argument("--dry-run", action="store_true",
                    help="변경 없이 요약만 출력")
    ap.add_argument("--retire-missing", action="store_true",
                    help="이번 파일들에 없는 같은 종류의 기존 코드를 폐기(retired) 표시. "
                         "manual 행은 제외. 재등장하면 자동 부활.")
    args = ap.parse_args()

    if not (args.material or args.product):
        ap.error("--material 또는 --product 중 하나 이상을 지정하세요.")

    # 대상 DB 연결 + 스키마(item_code_master 포함) 보장. match_item_codes._open_target_db
    # 가 공용 헬퍼 — 비관례 파일명(rehearsal.db 등)이라도 *그 파일 연결* 에 직접 마이그레이션을
    # 적용한다(과거에는 init_db() 가 관례 DB 에만 스키마를 잡아 대상이 누락되는 버그가 있었음).
    conn, db_label = _open_target_db(args.db)
    print(f"[db] 대상: {db_label}")

    try:
        totals: Counter = Counter()
        seen_codes: dict[str, set] = {"material": set(), "product": set()}
        if args.material:
            for f in args.material:
                r = import_material_master(conn, f, source="code", dry_run=args.dry_run)
                tag = "원자재" if not args.dry_run else "원자재(예정)"
                print(f"[{tag}] {f}: read={r['read']} imported={r['imported']} "
                      f"skipped(비원자재)={r['skipped_non_material']} skipped(빈값)={r['skipped_empty']}")
                totals["material"] += r["imported"]
                seen_codes["material"] |= r["codes"]
        if args.product:
            for i, f in enumerate(args.product, start=2):
                src = f"code{i}" if i - 2 < len(args.product) else f"code{i}"
                r = import_product_master(conn, f, source=src, dry_run=args.dry_run)
                tag = "반제품" if not args.dry_run else "반제품(예정)"
                print(f"[{tag}] {f}: read={r['read']} imported={r['imported']} "
                      f"skipped(빈값)={r['skipped_empty']} 분류={r['category_breakdown']}")
                totals["product"] += r["imported"]
                seen_codes["product"] |= r["codes"]
        if args.retire_missing:
            # 이번에 지정한 종류(kind)만 폐기 대조 — --material 만 줬으면 반제품은
            # 건드리지 않는다(파일이 없는 종류를 일괄 폐기하는 사고 방지).
            now = utc_now_text()
            for kind, given in (("material", bool(args.material)),
                                ("product", bool(args.product))):
                if not given:
                    continue
                retired = retire_missing_codes(
                    conn, kind=kind, imported_codes=seen_codes[kind],
                    now=now, dry_run=args.dry_run,
                )
                tag = "폐기" if not args.dry_run else "폐기(예정)"
                head = ", ".join(retired[:10])
                more = f" 외 {len(retired) - 10}건" if len(retired) > 10 else ""
                print(f"[{tag}] kind={kind}: {len(retired)}건"
                      f"{' - ' + head + more if retired else ''}")
        print(f"[총계] material={totals['material']} product={totals['product']}"
              f"{' [DRY-RUN - 변경 없음]' if args.dry_run else ''}")
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
