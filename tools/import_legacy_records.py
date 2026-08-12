"""구 배합프로그램_v1.2 실적서 → BRM 배합 기록 이관 (제품 LOT 누락분만).

전제(사용자 규칙 2026-08-06):
- 대조 기준은 **제품 LOT 존재 여부뿐**이다. 구 툴은 수기 입력이라 같은 LOT 의
  배합량 숫자가 BRM 과 달라도 같은 기록으로 보고 건드리지 않는다.
- 이미 BRM 에 있는 LOT 은 무조건 건너뛴다(실행 시점의 DB 를 다시 확인하므로
  낡은 백업으로 비교했어도 안전).

동작:
- 실적서 excel 폴더의 각 <제품LOT>.xlsx 를 파싱해 blend_records + blend_details
  로 삽입한다. 재고·LOT 소비 등 부수효과는 일으키지 않는다(과거 이력 백필).
- created_by 마커('구프로그램 이관')로 이관분을 식별할 수 있다(정리 한정 가능,
  바인더 점도 이관과 동일 패턴).
- 레시피 연결: 제품명(LOT 에서 날짜부 8자리 제거)이 BRM 레시피와 일치하면
  recipe_id 연결, 없으면 NULL(과거 이력 보존 — 바인더 이관과 동일 방침).
- 자재 연결(2026-08-11 추가): 자재명을 materials(이름·별칭)와 정규화 대조해
  material_id·material_code 를 채운다. 종전에는 이 단계가 없어 NULL/'' 로 넣었고,
  그래서 이관분이 자재 사용량 API 에서 품목코드 없이 나가 상위 재고 대시보드가
  통째로 버렸다(미매핑 71kg 이 전부 이 경로였다). 못 찾은 이름은 자재를 새로
  만들지 않고 요약에 이름째로 보고한다 — 구 프로그램 표기로 마스터가 오염되는 것을
  막고, 운영자가 '자재 관리 > 품목코드'에서 동의어로 이으면 소급 해석된다.
  미리보기(--apply 없이)에서도 연결 결과가 나오므로 삽입 전에 확인할 수 있다.

사용(운영 PC, 서버 정지 필요 없음 — 짧은 트랜잭션):
  # 미리보기(아무것도 안 씀):
  python tools/import_legacy_records.py --source "배합프로그램_v1.2/배합프로그램_v1.2/실적서/excel"
  # 실제 삽입:
  python tools/import_legacy_records.py --source ... --apply
IRMS_DATA_DIR 환경변수(또는 --data-dir)가 운영 데이터 폴더를 가리켜야 한다.
"""

from __future__ import annotations

import argparse
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

MARKER = "구프로그램 이관"
LOT_DATE_RE = re.compile(r"^(.*?)(\d{8})$")  # 제품명 + YYMMDDNN


def _norm(value: str | None) -> str:
    """자재명 매칭용 정규화 — 대문자화 후 영숫자·한글만 남긴다.

    src/db/queries.py 의 normalize_token 과 같은 규칙을 의도적으로 복제한다. 이 도구는
    운영 PC 에서 sqlite3 만으로 도는 단독 스크립트라 src 패키지를 임포트하지 않는다.
    """
    return "".join(ch for ch in (value or "").strip().upper() if ch.isalnum())


def _material_index(conn: sqlite3.Connection) -> dict[str, tuple[int, str]]:
    """정규화(자재명·별칭) → (materials.id, code) 색인.

    자재명이 먼저고 별칭은 그 자재를 못 찾았을 때만 채운다(자재명이 항상 우선).
    """
    index: dict[str, tuple[int, str]] = {}
    for r in conn.execute("SELECT id, name, code FROM materials"):
        key = _norm(r["name"])
        if key:
            index[key] = (int(r["id"]), (r["code"] or "").strip())
    for r in conn.execute(
        "SELECT a.alias_name AS alias, m.id AS mid, m.code AS code "
        "FROM material_aliases a JOIN materials m ON m.id = a.material_id"
    ):
        key = _norm(r["alias"])
        if key and key not in index:
            index[key] = (int(r["mid"]), (r["code"] or "").strip())
    return index


def parse_sheet(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # 2행: 날짜/작업자/시간, 3행: 저울 — 라벨 위치가 밀려도 견디게 문자열 스캔.
    meta = {"date": None, "worker": None, "time": None, "scale": None}
    for row in rows[:4]:
        for cell in row:
            if not isinstance(cell, str):
                continue
            if "날짜" in cell and ":" in cell:
                meta["date"] = cell.split(":", 1)[1].strip()
            elif "작업자" in cell:
                meta["worker"] = cell.split(":", 1)[1].strip()
            elif "작업시간" in cell:
                meta["time"] = cell.split(":", 1)[1].strip()
            elif "저울" in cell:
                meta["scale"] = cell.split(":", 1)[1].strip()
    # 헤더 행 찾기('약품번호'로 시작) — 그 다음부터 자재 행.
    header_idx = None
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and "약품번호" in row[0]:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("헤더(약품번호) 행을 찾지 못했다")
    details = []
    product_lot = None
    total_x100 = None
    for row in rows[header_idx + 1:]:
        if not row or all(v is None for v in row):
            continue
        cells = list(row) + [None] * (7 - len(row))
        # 첫 자재 행: A=제품LOT, B=배합량(100g), C~G=자재. 이후 행: A,B 비고 C~G 자재.
        if cells[0] is not None and product_lot is None:
            product_lot = str(cells[0]).strip()
            total_x100 = cells[1]
        name, lot, ratio, theory, actual = cells[2:7]
        if name is None and lot is None and theory is None:
            continue
        details.append({
            "material_name": str(name).strip() if name is not None else "",
            "material_lot": str(lot).strip() if lot is not None else "",
            "ratio": float(ratio) if ratio is not None else None,
            "theory_amount": float(theory) if theory is not None else 0.0,
            "actual_amount": float(actual if actual is not None else theory or 0.0),
        })
    if not product_lot:
        raise ValueError("제품 LOT 을 찾지 못했다")
    if not details:
        raise ValueError("자재 행이 없다")
    # 총량 = 자재 이론합. 구 툴의 '배합량(100g)' 칸은 제품마다 의미가 달라
    # (6-1 TOP=×100, APB·신규 S2+코팅=다른 단위/자유 표기) 신뢰할 수 없다 —
    # 실측 전수 확인(2026-08-06 dry-run 137건).
    sum_theory = round(sum(d["theory_amount"] for d in details), 2)
    total = sum_theory
    del total_x100  # 참고용으로만 파싱했다
    file_lot = path.stem.strip()
    if file_lot != product_lot:
        # 파일명이 기준(누락 대조가 파일명으로 이뤄졌으므로) — 내부 표기가 다르면 경고만.
        pass
    return {
        "product_lot": file_lot,
        "sheet_lot": product_lot,
        "work_date": meta["date"],
        "work_time": meta["time"],
        "worker": meta["worker"] or "미상",
        "scale": meta["scale"] or "M-65",
        "total_amount": total,
        "sum_theory": sum_theory,
        "details": details,
    }


def product_name_of(lot: str) -> str:
    m = LOT_DATE_RE.match(lot)
    return (m.group(1) if m else lot).strip()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, help="실적서 excel 폴더")
    ap.add_argument("--data-dir", default=os.environ.get("IRMS_DATA_DIR", "data"))
    ap.add_argument("--apply", action="store_true", help="실제 삽입(기본은 미리보기)")
    args = ap.parse_args()

    src = Path(args.source)
    db_path = Path(args.data_dir) / "irms.db"
    if not db_path.exists():
        print(f"DB 없음: {db_path}", file=sys.stderr)
        return 2
    files = sorted(f for f in src.iterdir() if f.suffix.lower() == ".xlsx")
    if not files:
        print(f"실적서 xlsx 없음: {src}", file=sys.stderr)
        return 2

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    existing = {r[0] for r in conn.execute("SELECT product_lot FROM blend_records")}
    recipes = {
        str(r["product_name"]).strip(): int(r["id"])
        for r in conn.execute(
            "SELECT id, product_name FROM recipes ORDER BY id"
        )
    }

    # 자재 색인 — 실적서의 자재명을 materials.id/code 로 잇는다. 종전에는 이 단계가
    # 통째로 빠져 material_id=NULL, material_code='' 로 넣었고, 그 결과 이관분이
    # 자재 사용량 API 에서 품목코드 없이 나가 상위 재고 대시보드가 통째로 버렸다
    # (2026-08-11 에 드러난 미매핑 71kg 이 전부 이 경로였다).
    # 못 찾은 이름은 자재를 새로 만들지 않는다 — 구 프로그램 표기로 마스터가 오염된다.
    # 대신 아래 요약에서 이름을 그대로 보고하고, 운영자가 '자재 관리 > 품목코드' 에서
    # 동의어로 이으면 과거 기록까지 소급 해석된다.
    mat_index = _material_index(conn)
    unresolved: dict[str, int] = {}

    parsed, skipped, inserted, errors = 0, 0, 0, []
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    for f in files:
        lot = f.stem.strip()
        if lot in existing:
            skipped += 1
            continue
        try:
            rec = parse_sheet(f)
        except Exception as e:  # noqa: BLE001 — 파일 단위 보고
            errors.append(f"{f.name}: {e}")
            continue
        parsed += 1
        pname = product_name_of(lot)
        rid = recipes.get(pname)
        # 자재 해석은 미리보기에서도 돌린다 — 무엇이 안 이어지는지 삽입 전에 보여야
        # 운영자가 동의어를 먼저 등록하고 다시 돌릴 수 있다.
        resolved = [mat_index.get(_norm(d["material_name"])) for d in rec["details"]]
        for d, hit in zip(rec["details"], resolved):
            if hit is None:
                name = str(d["material_name"] or "").strip()
                unresolved[name] = unresolved.get(name, 0) + 1
        linked = sum(1 for h in resolved if h is not None)
        print(f"{'[삽입]' if args.apply else '[예정]'} {lot} · {rec['work_date']}"
              f" · {rec['worker']} · {rec['total_amount']:g}g"
              f" · 자재 {len(rec['details'])}종 · 레시피 {'연결' if rid else '없음(NULL)'}"
              f" · 자재연결 {linked}/{len(resolved)}")
        if not args.apply:
            continue
        cur = conn.execute(
            "INSERT INTO blend_records (product_lot, recipe_id, product_name, ink_name,"
            " worker, work_date, work_time, total_amount, scale, status, note,"
            " created_by, created_at, manual_entry)"
            " VALUES (?,?,?,?,?,?,?,?,?,'completed',?,?,?,0)",
            (
                lot, rid, pname, "none", rec["worker"], rec["work_date"],
                rec["work_time"], rec["total_amount"], rec["scale"],
                None, MARKER, now,
            ),
        )
        rec_id = cur.lastrowid
        for seq, (d, hit) in enumerate(zip(rec["details"], resolved), start=1):
            # 못 찾으면 NULL/'' 로 둔다(자재를 새로 만들지 않는다). 자재명은 그대로
            # 남으므로, 나중에 동의어를 등록하면 읽는 시점에 품목코드가 붙는다.
            mat_id, mat_code = hit if hit else (None, "")
            conn.execute(
                "INSERT INTO blend_details (blend_record_id, material_id, material_code,"
                " material_name, material_lot, ratio, theory_amount, actual_amount,"
                " sequence_order, created_at, manual_entry, carried_over)"
                " VALUES (?,?,?,?,?,?,?,?,?,?,0,0)",
                (
                    rec_id, mat_id, mat_code, d["material_name"], d["material_lot"],
                    d["ratio"], d["theory_amount"], d["actual_amount"], seq, now,
                ),
            )
        inserted += 1
    if args.apply:
        conn.commit()
    conn.close()

    print(f"\n요약: 대상 {len(files)} · 이미 있음(건너뜀) {skipped}"
          f" · 파싱 {parsed} · 삽입 {inserted if args.apply else 0}"
          f"{' (미리보기 - --apply 로 실제 삽입)' if not args.apply else ''}")

    # 자재를 못 이은 이름은 반드시 이름째로 보고한다. 종전엔 이 단계 자체가 없어
    # 이관이 조용히 성공한 것처럼 보였고, 그 결과가 재고 대시보드의 71kg 누락이었다.
    if unresolved:
        # 콘솔 코드페이지가 949 인 운영 PC 에서 '⚠'·'—' 는 UnicodeEncodeError 로
        # 도구를 죽인다. 출력 문자열은 CP949 로 인코딩 가능한 것만 쓴다.
        print(f"\n[주의] 자재 마스터에서 못 찾은 자재명 {len(unresolved)}종 -"
              " 이 이름으로 들어간 실적은 품목코드 없이 집계됩니다.")
        for name, n in sorted(unresolved.items(), key=lambda kv: -kv[1]):
            print(f"   {name}  ({n}행)")
        print("   → 화면 '자재 관리 > 품목코드' 에서 해당 자재의 [동의어] 로 이 이름을"
              " 등록하면 과거 기록까지 소급 해석됩니다.")
    elif parsed:
        print("자재 연결: 모든 자재명이 마스터와 이어졌습니다.")
    if errors:
        print("\n파싱 실패:")
        for e in errors:
            print("  !", e)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
