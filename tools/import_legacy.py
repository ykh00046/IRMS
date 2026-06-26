"""구 잉크 데스크톱(Program-estimation) 데이터 → IRMS 이관 도구.

  - 레시피 xlsx(레시피·품목코드·품목명·배합비율, long 형식) → recipes + recipe_items
  - 배합기록.xlsx(제품LOT·작업자·레시피·배합량·품목·배합비율·원재료LOT·이론/실제·작업일시)
    또는 mixing_records.db → blend_records + blend_details (제품LOT로 묶음)

자재는 이름/품목코드(alias)로 해석하고 없으면 생성한다. 중복(레시피명·product_lot)은 건너뛴다.
대상 DB는 IRMS_DATA_DIR 의 운영 DB(get_connection). 운영 PC에서 실행한다.

사용:
  python tools/import_legacy.py --recipes 레시피.xlsx
  python tools/import_legacy.py --records mixing_records.db
  python tools/import_legacy.py --recipes 레시피.xlsx --records mixing_records.db
"""

import argparse
import os
import sqlite3
import sys
from collections import OrderedDict
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.db import get_connection, init_db  # noqa: E402  (sys.path 조정 후 import)

_IMPORT_ACTOR = "이관"


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _resolve_material(conn: sqlite3.Connection, name: str | None, code: str | None) -> int:
    """이름/품목코드로 자재 해석, 없으면 생성. material_id 반환."""
    name = (name or "").strip()
    code = (code or "").strip()
    if name:
        row = conn.execute("SELECT id FROM materials WHERE name = ?", (name,)).fetchone()
        if row:
            return row[0]
    for token in (code, name):
        if not token:
            continue
        row = conn.execute(
            "SELECT material_id FROM material_aliases WHERE alias_name = ?", (token,)
        ).fetchone()
        if row:
            return row[0]
    cur = conn.execute(
        "INSERT INTO materials (name, unit_type, unit, color_group, category, is_active) "
        "VALUES (?, 'weight', 'g', 'none', '기타', 1)",
        (name or code or "(미상)",),
    )
    mid = int(cur.lastrowid)
    if code and code != name:
        conn.execute(
            "INSERT INTO material_aliases (material_id, alias_name) VALUES (?, ?)", (mid, code)
        )
    return mid


def import_recipes(conn: sqlite3.Connection, xlsx_path: str, *, is_dhr: bool = False) -> tuple[int, int]:
    """레시피 xlsx → recipes + recipe_items. is_dhr=True 면 DHR 전용으로 적재. (생성, 중복) 반환."""
    import openpyxl

    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    groups: "OrderedDict[str, list]" = OrderedDict()
    for r in range(2, ws.max_row + 1):  # 1행은 헤더
        product = ws.cell(r, 1).value
        if product is None or str(product).strip() == "":
            continue
        groups.setdefault(str(product).strip(), []).append(
            (ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value)
        )

    now = _now()
    created = skipped = 0
    for product, items in groups.items():
        if conn.execute(
            "SELECT 1 FROM recipes WHERE product_name = ? AND revision_of IS NULL LIMIT 1", (product,)
        ).fetchone():
            skipped += 1
            continue
        cur = conn.execute(
            "INSERT INTO recipes (product_name, ink_name, status, is_dhr, created_by, created_at, effective_from) "
            "VALUES (?, ?, 'completed', ?, ?, ?, ?)",
            (product, product, 1 if is_dhr else 0, _IMPORT_ACTOR, now, now[:10]),  # 잉크명 없음→레시피명
        )
        recipe_id = int(cur.lastrowid)
        for code, mname, ratio in items:
            material_id = _resolve_material(conn, mname, code)
            try:
                value = float(ratio)
            except (TypeError, ValueError):
                value = None
            conn.execute(
                "INSERT INTO recipe_items (recipe_id, material_id, value_weight) VALUES (?, ?, ?)",
                (recipe_id, material_id, value),
            )
        created += 1
    conn.commit()
    return created, skipped


def _split_datetime(value) -> tuple[str, str]:
    """'2025-06-25 09:35:49'(또는 datetime) → (work_date, work_time)."""
    if value is None:
        return "", ""
    s = str(value).strip()
    if " " in s:
        d, t = s.split(" ", 1)
        return d[:10], t.strip()[:8]
    return s[:10], ""


def import_records_xlsx(
    conn: sqlite3.Connection, xlsx_path: str, *, default_scale: str = "M-65"
) -> tuple[int, int]:
    """배합기록 xlsx(제품LOT·작업자·레시피·배합량·품목코드·품목명·배합비율·원재료LOT·
    이론계량·실제배합·작업일시) → blend_records + blend_details. 제품LOT로 묶는다."""
    import openpyxl

    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    groups: "OrderedDict[str, dict]" = OrderedDict()
    for r in range(2, ws.max_row + 1):  # 1행은 헤더
        lot = ws.cell(r, 1).value
        if lot is None or str(lot).strip() == "":
            continue
        lot = str(lot).strip()
        grp = groups.setdefault(lot, {
            "worker": ws.cell(r, 2).value,
            "recipe": ws.cell(r, 3).value,
            "total": ws.cell(r, 4).value,
            "dt": ws.cell(r, 11).value,
            "details": [],
        })
        grp["details"].append((
            ws.cell(r, 5).value,   # 품목코드
            ws.cell(r, 6).value,   # 품목명
            ws.cell(r, 7).value,   # 배합비율
            ws.cell(r, 8).value,   # 원재료LOT
            ws.cell(r, 9).value,   # 이론계량
            ws.cell(r, 10).value,  # 실제배합
        ))

    now = _now()
    created = skipped = 0
    for lot, grp in groups.items():
        if conn.execute("SELECT 1 FROM blend_records WHERE product_lot = ?", (lot,)).fetchone():
            skipped += 1
            continue
        recipe_name = str(grp["recipe"] or "").strip()
        work_date, work_time = _split_datetime(grp["dt"])
        created_at = f"{work_date} {work_time}".strip() or now
        recipe_row = conn.execute(
            "SELECT id FROM recipes WHERE product_name = ? AND revision_of IS NULL ORDER BY id LIMIT 1",
            (recipe_name,),
        ).fetchone()
        try:
            total = float(grp["total"])
        except (TypeError, ValueError):
            total = 0.0
        cur = conn.execute(
            "INSERT INTO blend_records (product_lot, recipe_id, product_name, ink_name, worker, "
            "work_date, work_time, total_amount, scale, status, created_by, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'completed', ?, ?, ?)",
            (lot, recipe_row[0] if recipe_row else None, recipe_name, recipe_name,
             str(grp["worker"] or "").strip(), work_date, work_time, total, default_scale,
             _IMPORT_ACTOR, created_at, created_at),
        )
        blend_id = int(cur.lastrowid)
        for seq, (code, mname, ratio, mlot, theory, actual) in enumerate(grp["details"], start=1):
            material_id = _resolve_material(conn, mname, code)

            def _num(v):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return None

            conn.execute(
                "INSERT INTO blend_details (blend_record_id, material_id, material_code, "
                "material_name, material_lot, ratio, theory_amount, actual_amount, "
                "sequence_order, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (blend_id, material_id, (str(code).strip() if code else None),
                 str(mname or "").strip(), (str(mlot).strip() if mlot else None),
                 _num(ratio), _num(theory), _num(actual), seq, created_at),
            )
        created += 1
    conn.commit()
    return created, skipped


def import_records(conn: sqlite3.Connection, db_path: str) -> tuple[int, int]:
    """mixing_records.db → blend_records + blend_details. (생성, 중복건너뜀) 반환."""
    src = sqlite3.connect(db_path)
    src.row_factory = sqlite3.Row
    try:
        records = src.execute("SELECT * FROM mixing_records ORDER BY id").fetchall()
        now = _now()
        created = skipped = 0
        for rec in records:
            lot = rec["product_lot"]
            if conn.execute(
                "SELECT 1 FROM blend_records WHERE product_lot = ?", (lot,)
            ).fetchone():
                skipped += 1
                continue
            recipe_row = conn.execute(
                "SELECT id FROM recipes WHERE product_name = ? AND revision_of IS NULL ORDER BY id LIMIT 1",
                (rec["recipe_name"],),
            ).fetchone()
            recipe_id = recipe_row[0] if recipe_row else None
            created_at = rec["created_at"] or now
            cur = conn.execute(
                "INSERT INTO blend_records (product_lot, recipe_id, product_name, ink_name, worker, "
                "work_date, work_time, total_amount, scale, status, created_by, created_at, updated_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'completed', ?, ?, ?)",
                (lot, recipe_id, rec["recipe_name"], rec["recipe_name"], rec["worker"], rec["work_date"],
                 rec["work_time"], float(rec["total_amount"]), rec["scale"], _IMPORT_ACTOR,
                 created_at, created_at),
            )
            blend_id = int(cur.lastrowid)
            details = src.execute(
                "SELECT * FROM mixing_details WHERE mixing_record_id = ? ORDER BY sequence_order",
                (rec["id"],),
            ).fetchall()
            for d in details:
                material_id = _resolve_material(conn, d["material_name"], d["material_code"])
                conn.execute(
                    "INSERT INTO blend_details (blend_record_id, material_id, material_code, "
                    "material_name, material_lot, ratio, theory_amount, actual_amount, "
                    "sequence_order, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (blend_id, material_id, d["material_code"], d["material_name"], d["material_lot"],
                     d["ratio"], d["theory_amount"], d["actual_amount"], d["sequence_order"], created_at),
                )
            created += 1
        conn.commit()
        return created, skipped
    finally:
        src.close()


def main() -> None:
    ap = argparse.ArgumentParser(description="구 잉크 데스크톱 데이터 → IRMS 이관")
    ap.add_argument("--recipes", help="레시피 xlsx 경로")
    ap.add_argument("--records", help="배합기록.xlsx 또는 mixing_records.db 경로(확장자 자동 인식)")
    ap.add_argument("--dhr", action="store_true", help="레시피를 DHR 전용으로 적재")
    args = ap.parse_args()
    if not (args.recipes or args.records):
        ap.error("--recipes 또는 --records 중 하나 이상을 지정하세요.")
    init_db()  # 스키마 보장(기존 운영 DB면 no-op)
    with get_connection() as conn:
        if args.recipes:
            c, s = import_recipes(conn, args.recipes, is_dhr=args.dhr)
            kind = "DHR 전용 레시피" if args.dhr else "레시피"
            print(f"{kind}: 생성 {c} · 중복 건너뜀 {s}")
        if args.records:
            if args.records.lower().endswith(".xlsx"):
                c, s = import_records_xlsx(conn, args.records)
            else:
                c, s = import_records(conn, args.records)
            print(f"배합기록: 생성 {c} · 중복 건너뜀 {s}")


if __name__ == "__main__":
    main()
